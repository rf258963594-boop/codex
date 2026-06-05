from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"

TITLE_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FILL = PatternFill("solid", fgColor="1F4D78")
INPUT_FILL = PatternFill("solid", fgColor="FFF8D6")
KEY_FILL = PatternFill("solid", fgColor="EEF2F7")
NOTE_FILL = PatternFill("solid", fgColor="F8FAFC")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
KEY_FONT = Font(color="475569", size=9)
NOTE_FONT = Font(color="334155")
WRAP = Alignment(wrap_text=True, vertical="top")

YES_NO = '"No,Yes"'
SOURCE_OPTIONS = '"new,common"'
PERSON_ACTIONS = '"appoint_director,resign_director,appoint_secretary,resign_secretary"'
TRANSFER_MODES = '"cooperative,non_cooperative"'
SHAREHOLDER_TYPES = '"person,corporate"'
CONSIDERATION_BASIS = '"internal_paid_up_basis,acra_paid_up_capital_basis,stamp_duty_higher_of_price_or_nav"'
FIELD_LABELS = '"ID type,ID number,Residential address,Email,Phone,Nationality,Name,Other"'
ANNUAL_ROUTES = '"ordinary_agm,exempt_private_company,dormant_company,manual"'


def title(ws, cells: str) -> None:
    for row in ws[cells]:
        for cell in row:
            cell.fill = TITLE_FILL
            cell.font = Font(color="FFFFFF", bold=True, size=15)
            cell.alignment = WRAP


def style_header(ws, row_no: int = 1) -> None:
    for cell in ws[row_no]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = WRAP


def style_common(ws) -> None:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
    for col in range(1, ws.max_column + 1):
        max_len = 12
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)) + 2, 45))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def field_from_header(value: Any) -> str:
    text = str(value or "")
    return text.splitlines()[-1].strip()


def add_table_validation(ws, key: str, values: str, start_row: int = 2, end_row: int = 200) -> None:
    headers = [field_from_header(cell.value) for cell in ws[1]]
    if key not in headers:
        return
    col = headers.index(key) + 1
    dv = DataValidation(type="list", formula1=values, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}")


def add_vertical_validation(ws, key: str, values: str) -> None:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 3).value == key:
            dv = DataValidation(type="list", formula1=values, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"D{row}")
            return


def add_transposed_validation(ws, key: str, values: str, start_col: int = 5, end_col: int = 14) -> None:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 3).value == key:
            dv = DataValidation(type="list", formula1=values, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}")
            return


def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("填写说明")
    ws.merge_cells("A1:C1")
    ws["A1"] = "P2 维护/变更/年审快速生成表 v5"
    rows = [
        ("主题", "中文说明", "English notes"),
        ("核心逻辑", "每个详情页自己的“是否办理/是否变更”才是准绳；默认 No 不生成，避免两处冲突。", "Each detail sheet has its own Yes/No switch."),
        ("最少字段", "普通变更通常只需要公司名称、UEN、注册地址、日期、签字董事和本次变更字段。", "Minimal fields first; full database data can be added later."),
        ("公司资料变更", "营业范围支持主业务和副业务；旧资料可空，新资料按需要填写。", "Business activity supports primary and secondary SSIC/activity."),
        ("人员资料更新", "一个人多个资料更新时，在“个人资料变更”填多行。", "Use one row per changed particular."),
        ("系统规则", "DR 合并、文件组、审批路线默认由后台判断，普通填写人不用填。", "Grouping and approval route are system rules."),
        ("旧版兼容", "v4 旧表仍可上传；v5 是新的主入口。", "v4 remains compatible; v5 is the new main entry."),
        ("日期", "建议用 DD/MM/YYYY；正式文件会转为英文日月年格式。", "Use DD/MM/YYYY."),
    ]
    for row in rows:
        ws.append(row)
    title(ws, "A1:C1")
    style_header(ws, 2)
    style_common(ws)
    ws.freeze_panes = "A3"


def build_vertical_sheet(wb: Workbook, name: str, heading: str, fields: list[tuple], sample: dict[str, Any] | None = None) -> None:
    sample = sample or {}
    ws = wb.create_sheet(name)
    ws.merge_cells("A1:G1")
    ws["A1"] = heading
    ws.append([])
    ws.append(["项目", "English", "field_key", "填写内容 / Value", "必填", "适用场景", "说明"])
    for cn, en, key, default, required, scene, note in fields:
        ws.append([cn, en, key, sample.get(key, default), required, scene, note])
    title(ws, "A1:G1")
    style_header(ws, 3)
    for row in range(4, ws.max_row + 1):
        ws.cell(row, 3).fill = KEY_FILL
        ws.cell(row, 3).font = KEY_FONT
        ws.cell(row, 4).fill = INPUT_FILL
        ws.cell(row, 7).fill = NOTE_FILL
        ws.cell(row, 7).font = NOTE_FONT
    style_common(ws)
    ws.freeze_panes = "A4"


def build_transposed_sheet(wb: Workbook, name: str, heading: str, fields: list[tuple], prefix: str, sample_rows: list[dict[str, Any]] | None = None, count: int = 10) -> None:
    sample_rows = sample_rows or []
    ws = wb.create_sheet(name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + count)
    ws["A1"] = heading
    ws.append([])
    ws.append(["项目", "English", "field_key", "说明", *[f"{prefix}{i}" for i in range(1, count + 1)]])
    for cn, en, key, note in fields:
        values = [sample_rows[i].get(key, "") if i < len(sample_rows) else "" for i in range(count)]
        ws.append([cn, en, key, note, *values])
    title(ws, f"A1:{get_column_letter(4 + count)}1")
    style_header(ws, 3)
    for row in range(4, ws.max_row + 1):
        ws.cell(row, 3).fill = KEY_FILL
        ws.cell(row, 3).font = KEY_FONT
        ws.cell(row, 4).fill = NOTE_FILL
        ws.cell(row, 4).font = NOTE_FONT
        for col in range(5, 5 + count):
            ws.cell(row, col).fill = INPUT_FILL
    style_common(ws)
    ws.freeze_panes = "E4"


def build_table_sheet(wb: Workbook, name: str, headers: list[tuple[str, str]], rows: list[dict[str, Any]] | None = None) -> None:
    rows = rows or []
    ws = wb.create_sheet(name)
    ws.append([f"{label}\n{key}" for label, key in headers])
    for row in rows:
        ws.append([row.get(key, "") for _, key in headers])
    style_header(ws, 1)
    for row in range(2, max(ws.max_row, 200) + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).fill = INPUT_FILL
    style_common(ws)
    ws.freeze_panes = "A2"


COMPANY_FIELDS = [
    ("任务类型", "Task type", "task_type", "maintenance", "Yes", "系统", "固定值，不要改。"),
    ("模板版本", "Template version", "template_version", "p2_v5", "Yes", "系统", "固定值，不要改。"),
    ("业务单编号", "Business order ID", "business_order_id", "", "Optional", "后台预留", "以后对接客户系统/电子签名使用。"),
    ("数据来源", "Source type", "source_type", "AI", "Optional", "后台预留", "AI / BizFile / Excel / Manual。"),
    ("来源文件编号", "Source file ID", "source_file_id", "", "Optional", "后台预留", "可填 BizFile 文件名、旧表名或备注。"),
    ("公司名称", "Company name", "company_name", "", "Yes", "全部文件", "英文法定名称。"),
    ("UEN", "UEN", "uen", "", "Yes", "现有公司", "现有公司变更/年审建议必填。"),
    ("当前注册地址", "Current registered office", "registered_office_address", "", "Recommended", "多数文件", "地址变更时可作为旧地址。"),
    ("默认文件日期", "Default document date", "default_document_date", "", "Recommended", "全部文件", "DD/MM/YYYY；留空可由网页端补当天。"),
    ("当前秘书公司", "Current secretary firm", "current_secretary_company", "", "Optional", "转入", "转入交接时使用。"),
    ("新秘书公司", "New secretary firm", "new_secretary_company", "RSIN GROUP PTE. LTD.", "Optional", "转入", "留空默认 RSIN GROUP PTE. LTD.。"),
    ("联系人", "Contact person_id", "contact_person_id", "", "Optional", "后台预留", "引用人员信息 person_id。"),
    ("客户方默认签字人", "Default client signer person_id", "client_signatory_person_id", "", "Optional", "服务协议/授权", "留空默认第一个客户董事/股东。"),
    ("董事默认签字人", "Default director signer person_id", "director_signer_person_id", "", "Optional", "M01", "留空默认所有董事签 M01。"),
    ("经办人", "Prepared by", "prepared_by", "", "Optional", "后台记录", ""),
    ("备注", "Notes", "notes", "", "Optional", "内部", "不进入正式文件。"),
]

PEOPLE_FIELDS = [
    ("人员编号", "Person ID", "person_id", "建议 P001/P002，用于其他页面引用。"),
    ("来源", "Source", "source", "new=手填；common=调用后台常用人员。"),
    ("常用人员名称", "Common person", "common_person_name", "source=common 时填写后台保存的名称。"),
    ("英文姓名", "Full name", "full_name", "文件签字名/人员名。"),
    ("证件类型", "ID type", "id_type", "Passport / NRIC / FIN。"),
    ("证件号", "ID number", "id_number", ""),
    ("国籍", "Nationality", "nationality", "任免或资料更新需要时填写。"),
    ("住址", "Residential address", "residential_address", "任免或资料更新需要时填写。"),
    ("邮箱", "Email", "email", "可空。"),
    ("电话", "Phone", "phone", "可空。"),
    ("当前身份", "Current roles", "current_roles", "仅供人工/AI 理解。"),
    ("是否董事", "Director", "is_director", "M01 签字董事建议 Yes。"),
    ("是否秘书", "Secretary", "is_secretary", "常用秘书可 Auto。"),
    ("是否股东", "Shareholder", "is_shareholder", "可空/No；股权文件再补。"),
    ("是否客户签字人", "Client signer", "is_client_signatory", "服务协议/授权类文件使用。"),
    ("备注", "Remarks", "remarks", ""),
]

SHAREHOLDER_FIELDS = [
    ("股东编号", "Shareholder ID", "shareholder_id", "普通 M01 可不填。"),
    ("股东类型", "Shareholder type", "shareholder_type", "person / corporate。"),
    ("关联人员编号", "Person ID", "person_id", "自然人股东可引用 P001。"),
    ("股东名称", "Shareholder name", "shareholder_name", ""),
    ("证件/注册号", "ID / Registration no.", "id_or_reg_no", ""),
    ("股份类别", "Share class", "share_class", "留空默认 Ordinary。"),
    ("当前股数", "Shares", "shares", ""),
    ("实缴金额", "Paid amount", "paid_amount", ""),
    ("币种", "Currency", "currency", "留空默认 SGD。"),
    ("证书编号", "Certificate no.", "certificate_no", ""),
    ("是否当前股东", "Current shareholder", "is_current", "可空。"),
    ("备注", "Remarks", "remarks", ""),
]

COMPANY_CHANGE_FIELDS = [
    ("是否变更注册地址", "Change registered office", "change_registered_office_required", "No", "Optional", "地址变更", "Yes 时读取本段地址字段。"),
    ("旧注册地址", "Old registered office", "old_registered_office_address", "", "Optional", "地址变更", "可空；默认用 P2公司信息 当前注册地址。"),
    ("新注册地址", "New registered office", "new_registered_office_address", "", "If Yes", "地址变更", "办理地址变更时填写。"),
    ("是否变更营业范围", "Change business activity", "change_business_activity_required", "No", "Optional", "营业范围", "Yes 时读取主/副 SSIC 和业务范围字段。"),
    ("旧主 SSIC", "Old primary SSIC", "old_primary_ssic", "", "Optional", "营业范围", "可从 BizFile 带出；可空。"),
    ("旧主营业范围", "Old primary activity", "old_primary_activity", "", "Optional", "营业范围", "可从 BizFile 带出；可空。"),
    ("新主 SSIC", "New primary SSIC", "new_primary_ssic", "", "If Yes", "营业范围", "办理营业范围变更时建议填写。"),
    ("新主营业范围", "New primary activity", "new_primary_activity", "", "If Yes", "营业范围", "办理营业范围变更时填写。"),
    ("旧副 SSIC", "Old secondary SSIC", "old_secondary_ssic", "", "Optional", "营业范围", "没有副业务可空。"),
    ("旧副营业范围", "Old secondary activity", "old_secondary_activity", "", "Optional", "营业范围", "没有副业务可空。"),
    ("新副 SSIC", "New secondary SSIC", "new_secondary_ssic", "", "Optional", "营业范围", "需要第二业务时填写。"),
    ("新副营业范围", "New secondary activity", "new_secondary_activity", "", "Optional", "营业范围", "需要第二业务时填写。"),
    ("是否变更 FYE", "Change FYE", "change_fye_required", "No", "Optional", "FYE", "Yes 时读取 FYE 字段。"),
    ("旧 FYE", "Old FYE", "old_fye", "", "If Yes", "FYE", "DD/MM/YYYY。"),
    ("新 FYE", "New FYE", "new_fye", "", "If Yes", "FYE", "DD/MM/YYYY。"),
    ("下个账期开始", "Next accounts period start", "next_accounts_period_start", "", "Optional", "FYE", "特殊 FYE 变更时填写。"),
    ("下个账期结束", "Next accounts period end", "next_accounts_period_end", "", "Optional", "FYE", "留空默认新 FYE。"),
    ("是否变更办公时间", "Change office hours", "change_office_hours_required", "No", "Optional", "办公时间", "很少使用。"),
    ("新办公时间", "New office hours", "new_office_hours", "", "Optional", "办公时间", "很少使用；普通情况不用填。"),
]

PERSON_ACTION_HEADERS = [
    ("是否办理", "generate"),
    ("动作", "action_type"),
    ("人员编号", "target_person_id"),
    ("人员姓名", "target_name"),
    ("生效日期", "effective_date"),
    ("是否出辞职信", "resignation_letter"),
    ("备注", "remarks"),
]

PARTICULAR_HEADERS = [
    ("是否办理", "generate"),
    ("人员编号", "target_person_id"),
    ("人员姓名", "target_name"),
    ("变更字段", "field_label"),
    ("原资料", "old_value"),
    ("新资料", "new_value"),
    ("生效日期", "effective_date"),
    ("备注", "remarks"),
]

TRANSFER_IN_FIELDS = [
    ("是否转入", "Transfer-in required", "transfer_in_required", "No", "Optional", "转入", "也可在办理事项页选 Yes。"),
    ("转入模式", "Transfer-in mode", "transfer_in_mode", "cooperative", "Optional", "转入", "cooperative / non_cooperative。"),
    ("生效日期", "Effective date", "effective_date", "", "If Yes", "转入", "DD/MM/YYYY。"),
    ("旧秘书公司", "Old secretary firm", "old_secretary_company", "", "Optional", "转入", "可空；默认用 P2公司信息 当前秘书公司。"),
    ("新秘书公司", "New secretary firm", "new_secretary_company", "RSIN GROUP PTE. LTD.", "Optional", "转入", "留空默认 RSIN GROUP PTE. LTD.。"),
    ("是否出交接信", "Generate handover letter", "generate_handover_letter", "Yes", "Optional", "转入", "M02 接入时使用。"),
    ("是否出辞职信", "Generate resignation letter", "generate_resignation_letter", "No", "Optional", "转入", "不配合转入通常不要出旧人员辞职信。"),
    ("客户签字人", "Client signer person_id", "client_signatory_person_id", "", "Optional", "转入", "留空默认公司信息的客户签字人。"),
    ("备注", "Remarks", "remarks", "", "Optional", "内部", ""),
]

TRANSFER_HEADERS = [
    ("是否办理", "generate"),
    ("转让方", "transferor_name"),
    ("受让方", "transferee_name"),
    ("股份类别", "share_class"),
    ("转让股数", "shares_transferred"),
    ("转让日期", "transfer_date"),
    ("对价口径", "consideration_basis"),
    ("现金对价", "cash_consideration"),
    ("旧证书编号", "old_certificate_no"),
    ("新证书编号", "new_certificate_no"),
    ("印花税复核", "stamp_duty_review"),
    ("备注", "remarks"),
]

ALLOTMENT_HEADERS = [
    ("是否办理", "generate"),
    ("认购人编号", "allottee_person_id"),
    ("认购人名称", "allottee_name"),
    ("股份类别", "share_class"),
    ("配发股数", "shares_allotted"),
    ("每股实缴", "amount_paid_per_share"),
    ("总实缴", "total_paid"),
    ("币种", "currency"),
    ("配股日期", "allotment_date"),
    ("授权日期", "authority_date"),
    ("是否 Form 24", "form24_required"),
    ("证书编号", "certificate_no"),
    ("备注", "remarks"),
]

ANNUAL_FIELDS = [
    ("是否做年审", "Annual review required", "annual_review_required", "No", "Optional", "年审", "Yes 时后续生成年审包。"),
    ("财政年结日", "Financial year end", "fye_date", "", "If Yes", "年审", "DD/MM/YYYY。"),
    ("AGM 日期", "AGM date", "agm_date", "", "Recommended", "年审", "DD/MM/YYYY。"),
    ("AGM 方式", "AGM route", "agm_route", "ordinary_agm", "Optional", "年审", "ordinary_agm / exempt_private_company / dormant_company / manual。"),
    ("AR 授权签字人", "AR signer person_id", "ar_authorized_signer_person_id", "", "Optional", "年审", "留空默认客户董事/股东。"),
    ("备注", "Remarks", "remarks", "", "Optional", "内部", ""),
]

OPTIONAL_FIELDS = [
    ("当前主 SSIC", "Current primary SSIC", "current_primary_ssic", "", "Optional", "资料库", "普通文件可不填。"),
    ("当前主营业范围", "Current primary activity", "current_primary_activity", "", "Optional", "资料库", "普通文件可不填。"),
    ("当前副 SSIC", "Current secondary SSIC", "current_secondary_ssic", "", "Optional", "资料库", "普通文件可不填。"),
    ("当前副营业范围", "Current secondary activity", "current_secondary_activity", "", "Optional", "资料库", "普通文件可不填。"),
    ("当前 FYE", "Current FYE", "current_fye", "", "Optional", "资料库", "普通文件可不填。"),
    ("当前办公时间", "Current office hours", "current_office_hours", "", "Optional", "资料库", "普通文件可不填。"),
]

OUTPUT_FIELDS = [
    ("文件包", "Package", "package", "AUTO_MAINTENANCE_V5", "System", "系统", "网站自动判断。"),
    ("输出格式", "Output format", "output_format", "pdf_zip", "System", "系统", "目前生成 PDF 包。"),
    ("生成内部核对表", "Generate internal checklist", "generate_internal_checklist", "Yes", "Optional", "后台", ""),
    ("输出备注", "Output notes", "notes", "", "Optional", "内部", ""),
]


def sample_company() -> dict[str, Any]:
    return {
        "business_order_id": "MAINT-V5-TEST-001",
        "source_type": "AI",
        "source_file_id": "BizFile + AI draft",
        "company_name": "SAMPLE V5 MAINTENANCE PTE. LTD.",
        "uen": "202612345A",
        "registered_office_address": "10 ANSON ROAD, #20-01 INTERNATIONAL PLAZA, SINGAPORE 079903",
        "default_document_date": "03/06/2026",
        "current_secretary_company": "OLD SECRETARIAL FIRM PTE. LTD.",
        "contact_person_id": "P001",
        "client_signatory_person_id": "P001",
        "director_signer_person_id": "P001",
        "prepared_by": "AI draft",
    }


def sample_people() -> list[dict[str, Any]]:
    return [
        {"person_id": "P001", "source": "new", "full_name": "ZHANG YI", "id_type": "Passport", "id_number": "E12345678", "nationality": "Chinese", "residential_address": "Shanghai, China", "email": "zhangyi@example.com", "phone": "+86 13800000001", "current_roles": "Director / Shareholder", "is_director": "Yes", "is_secretary": "No", "is_shareholder": "Yes", "is_client_signatory": "Yes", "remarks": "客户董事股东"},
        {"person_id": "P002", "source": "common", "common_person_name": "公司秘书 A", "full_name": "", "current_roles": "Secretary", "is_director": "No", "is_secretary": "Yes", "is_shareholder": "No", "is_client_signatory": "No", "remarks": "常用秘书"},
    ]


def sample_shareholders() -> list[dict[str, Any]]:
    return [
        {"shareholder_id": "SH001", "shareholder_type": "person", "person_id": "P001", "shareholder_name": "ZHANG YI", "id_or_reg_no": "E12345678", "share_class": "Ordinary", "shares": 1000, "paid_amount": 1000, "currency": "SGD", "certificate_no": "001", "is_current": "Yes"},
    ]


def sample_company_changes() -> dict[str, Any]:
    return {
        "change_registered_office_required": "Yes",
        "new_registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
        "change_business_activity_required": "Yes",
        "old_primary_ssic": "70201",
        "old_primary_activity": "MANAGEMENT CONSULTANCY SERVICES",
        "new_primary_ssic": "62011",
        "new_primary_activity": "DEVELOPMENT OF SOFTWARE AND APPLICATIONS",
        "old_secondary_ssic": "62019",
        "old_secondary_activity": "OTHER INFORMATION TECHNOLOGY AND COMPUTER SERVICE ACTIVITIES",
        "new_secondary_ssic": "73100",
        "new_secondary_activity": "ADVERTISING ACTIVITIES",
        "change_fye_required": "No",
        "change_office_hours_required": "No",
    }


def build_workbook(sample: bool) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb)
    build_vertical_sheet(wb, "P2公司信息", "P2公司信息 Company - 最少字段", COMPANY_FIELDS, sample_company() if sample else {})
    build_transposed_sheet(wb, "人员信息", "人员信息 People - 普通变更只填用得到的人", PEOPLE_FIELDS, "人员", sample_people() if sample else [])
    build_vertical_sheet(wb, "公司资料变更", "公司资料变更 Company Changes - 只填本次办理事项", COMPANY_CHANGE_FIELDS, sample_company_changes() if sample else {})
    build_table_sheet(wb, "人员任免", PERSON_ACTION_HEADERS, [
        {"generate": "Yes", "action_type": "appoint_secretary", "target_person_id": "P002", "target_name": "公司秘书 A", "effective_date": "03/06/2026", "resignation_letter": "No", "remarks": "示例：委任秘书"}
    ] if sample else [])
    build_table_sheet(wb, "个人资料变更", PARTICULAR_HEADERS, [
        {"generate": "Yes", "target_person_id": "P001", "target_name": "ZHANG YI", "field_label": "ID number", "old_value": "E12345678", "new_value": "G1234567A", "effective_date": "03/06/2026", "remarks": "示例：护照/FIN 更新"},
        {"generate": "Yes", "target_person_id": "P001", "target_name": "ZHANG YI", "field_label": "Residential address", "old_value": "Shanghai, China", "new_value": "8 MARINA BOULEVARD, #10-01, SINGAPORE 018981", "effective_date": "03/06/2026", "remarks": "示例：地址更新"},
    ] if sample else [])
    build_vertical_sheet(wb, "转入交接", "转入交接 Transfer-in - 不做可全空", TRANSFER_IN_FIELDS, {})
    build_table_sheet(wb, "股份转让", TRANSFER_HEADERS, [])
    build_table_sheet(wb, "增资配股", ALLOTMENT_HEADERS, [])
    build_vertical_sheet(wb, "年审信息", "年审信息 Annual Review - 不做可全空", ANNUAL_FIELDS, {})
    build_transposed_sheet(wb, "股东现状", "股东现状 Shareholders - 普通变更可不填", SHAREHOLDER_FIELDS, "股东", sample_shareholders() if sample else [])
    build_vertical_sheet(wb, "可选资料", "可选资料 Optional Master Data - 普通文件不用填", OPTIONAL_FIELDS, {})
    build_vertical_sheet(wb, "输出设置", "输出设置 Output Options", OUTPUT_FIELDS, {})
    apply_validations(wb)
    return wb


def apply_validations(wb: Workbook) -> None:
    for sheet in ["人员任免", "个人资料变更", "股份转让", "增资配股"]:
        if sheet in wb.sheetnames:
            add_table_validation(wb[sheet], "generate", YES_NO)
            add_table_validation(wb[sheet], "required", YES_NO)
    company_changes = wb["公司资料变更"]
    for key in ["change_registered_office_required", "change_business_activity_required", "change_fye_required", "change_office_hours_required"]:
        add_vertical_validation(company_changes, key, YES_NO)
    add_table_validation(wb["人员任免"], "action_type", PERSON_ACTIONS)
    add_table_validation(wb["人员任免"], "resignation_letter", YES_NO)
    add_table_validation(wb["个人资料变更"], "field_label", FIELD_LABELS)
    add_table_validation(wb["股份转让"], "consideration_basis", CONSIDERATION_BASIS)
    add_table_validation(wb["股份转让"], "stamp_duty_review", YES_NO)
    add_table_validation(wb["增资配股"], "form24_required", YES_NO)

    people = wb["人员信息"]
    add_transposed_validation(people, "source", SOURCE_OPTIONS)
    for key in ["is_director", "is_secretary", "is_shareholder", "is_client_signatory"]:
        add_transposed_validation(people, key, YES_NO)
    shareholders = wb["股东现状"]
    add_transposed_validation(shareholders, "shareholder_type", SHAREHOLDER_TYPES)
    add_transposed_validation(shareholders, "is_current", YES_NO)

    transfer_in = wb["转入交接"]
    add_vertical_validation(transfer_in, "transfer_in_required", YES_NO)
    add_vertical_validation(transfer_in, "transfer_in_mode", TRANSFER_MODES)
    add_vertical_validation(transfer_in, "generate_handover_letter", YES_NO)
    add_vertical_validation(transfer_in, "generate_resignation_letter", YES_NO)

    annual = wb["年审信息"]
    add_vertical_validation(annual, "annual_review_required", YES_NO)
    add_vertical_validation(annual, "agm_route", ANNUAL_ROUTES)


def save_workbook(wb: Workbook, name: str) -> Path:
    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / name
    wb.save(path)
    return path


def main() -> None:
    paths = [
        save_workbook(build_workbook(False), "AI适配_公司维护变更年审资料模板_v5_快速生成详情开关版_空白.xlsx"),
        save_workbook(build_workbook(True), "AI适配_公司维护变更年审资料模板_v5_快速生成详情开关版_示例.xlsx"),
    ]
    for path in paths:
        print(path)


if __name__ == "__main__":
    main()
