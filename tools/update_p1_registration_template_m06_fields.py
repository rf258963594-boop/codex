from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
IMPORT_DIR = ROOT / "templates" / "import"
OUTPUT_DIR = ROOT / "outputs"

PRIMARY_FILES = [
    IMPORT_DIR / "P1_registration_blank_v3_1.xlsx",
    IMPORT_DIR / "P1_registration_sample_v3_1.xlsx",
]

OUTPUT_COPIES = {
    PRIMARY_FILES[0]: [
        OUTPUT_DIR / "AI适配_新公司注册资料模板_v3.1_人性化竖排_Auto版_空白.xlsx",
        OUTPUT_DIR / "AI适配_新公司注册资料模板_v3_人性化竖排_Auto版_空白.xlsx",
    ],
    PRIMARY_FILES[1]: [
        OUTPUT_DIR / "AI适配_新公司注册资料模板_v3.1_人性化竖排_Auto版_示例.xlsx",
        OUTPUT_DIR / "AI适配_新公司注册资料模板_v3_人性化竖排_Auto版_示例.xlsx",
    ],
}

COMPANY_FIELDS = [
    ["挂名董事提名人姓名覆盖", "Nominee director nominator name override", "nominee_director_nominator_name", "", "Optional", "通常留空；系统默认第一个客户董事。特殊情况才填写。"],
    ["挂名董事提名人证件号覆盖", "Nominee director nominator ID override", "nominee_director_nominator_id_number", "", "Optional", "通常留空；与上方姓名一起用于 M06 register。"],
    ["挂名董事提名人地址覆盖", "Nominee director nominator address override", "nominee_director_nominator_address", "", "Optional", "通常留空；与上方姓名一起用于 M06 register。"],
]

SHAREHOLDER_FIELDS = [
    ["是否登记控制人覆盖", "Registrable controller override", "is_registrable_controller", "通常留空=Auto；系统按 25% 股权规则判断。特殊情况可填 Yes/No。"],
    ["控制人依据覆盖", "Controller basis override", "controller_basis", "通常留空；特殊控制权安排才填写。"],
    ["控制人登记日期", "Controller entry date", "controller_start_date", "通常留空=注册日期。"],
    ["是否挂名股东", "Nominee shareholder", "is_nominee_shareholder", "通常留空=No；只有 nominee shareholder 情况填 Yes。"],
    ["挂名股东提名人/受益人", "Nominee shareholder nominator / beneficial owner", "nominee_shareholder_nominator_name", "通常留空；只有挂名股东情况填写。"],
    ["挂名股东提名人证件号", "Nominee shareholder nominator ID", "nominee_shareholder_nominator_id_number", "通常留空。"],
    ["挂名股东提名人地址", "Nominee shareholder nominator address", "nominee_shareholder_nominator_address", "通常留空。"],
]


def find_key_row(ws, key: str) -> int | None:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 3).value or "").strip() == key:
            return row
    return None


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.fill = copy(source.fill)
        target.font = copy(source.font)
        target.border = copy(source.border)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def write_row(ws, row_no: int, values: list[str]) -> None:
    for col, value in enumerate(values, start=1):
        ws.cell(row_no, col).value = value
    for col in range(len(values) + 1, ws.max_column + 1):
        ws.cell(row_no, col).value = ""


def add_auto_yes_no_validation(ws, row_no: int) -> None:
    if ws.max_column < 5:
        return
    validation = DataValidation(type="list", formula1='"Auto,Yes,No"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(f"E{row_no}:{get_column_letter(ws.max_column)}{row_no}")


def upsert_rows_after(ws, anchor_key: str, rows: list[list[str]]) -> None:
    anchor = find_key_row(ws, anchor_key)
    if anchor is None:
        raise ValueError(f"Cannot find anchor field_key: {anchor_key}")
    existing = {str(ws.cell(row, 3).value or "").strip(): row for row in range(1, ws.max_row + 1)}
    insert_rows = [row for row in rows if row[2] not in existing]
    if insert_rows:
        ws.insert_rows(anchor + 1, len(insert_rows))
        for offset, row_values in enumerate(insert_rows, start=1):
            row_no = anchor + offset
            copy_row_style(ws, anchor, row_no)
            write_row(ws, row_no, row_values)

    for row_values in rows:
        row_no = find_key_row(ws, row_values[2])
        if row_no is not None:
            write_row(ws, row_no, row_values)
            if row_values[2] in {"is_registrable_controller", "is_nominee_shareholder"}:
                add_auto_yes_no_validation(ws, row_no)


def update_readme(ws) -> None:
    for row in range(1, ws.max_row + 1):
        cn = str(ws.cell(row, 2).value or "")
        en = str(ws.cell(row, 3).value or "")
        if "RORC 控制人、Form 24" in cn:
            ws.cell(row, 2).value = cn.replace("RORC 控制人、Form 24", "RORC 控制人、M06 registers、Form 24")
        if "RORC, Form 24" in en:
            ws.cell(row, 3).value = en.replace("RORC, Form 24", "RORC, M06 registers, Form 24")


def update_workbook(path: Path) -> None:
    wb = load_workbook(path)
    upsert_rows_after(wb["公司信息"], "authorized_representative_person_id", COMPANY_FIELDS)
    upsert_rows_after(wb["股东与股份"], "signing_required", SHAREHOLDER_FIELDS)
    if "填写说明" in wb.sheetnames:
        update_readme(wb["填写说明"])
    wb.save(path)


def main() -> None:
    for path in PRIMARY_FILES:
        update_workbook(path)
        for target in OUTPUT_COPIES.get(path, []):
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(path, target)
        print(path)


if __name__ == "__main__":
    main()
