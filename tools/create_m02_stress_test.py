from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "outputs" / "AI适配_公司维护变更年审资料模板_v7_一页式快速业务单含快速年审_空白.xlsx"
OUTPUT = ROOT / "outputs" / "M02_transfer_in_stress_input.xlsx"


def set_field(ws, key: str, value: object) -> None:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 3).value or "").strip() == key:
            ws.cell(row, 4).value = value
            return
    raise KeyError(key)


def main() -> None:
    wb = load_workbook(INPUT)
    ws = wb["P2快速业务单"]

    values = {
        "company_name": "M02 TRANSFER IN STRESS TEST PTE. LTD.",
        "uen": "202612345M",
        "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A, PENINSULA PLAZA, SINGAPORE 179098",
        "default_document_date": "04/06/2026",
        "director_signer_names": "ALICE DIRECTOR\nBOB DIRECTOR",
        "client_signatory_name": "ALICE SHAREHOLDER",
        "business_order_id": "M02-STRESS-001",
        "source_type": "Manual",
        "source_file_id": "BizFile sample 2026-06-04",
        "prepared_by": "RSIN TEST USER",
        "transfer_in_required": "Yes",
        "transfer_in_mode": "cooperative",
        "old_secretary_company": "OLD SECRETARIAL SERVICES PTE. LTD.",
        "new_secretary_company": "RSIN GROUP PTE. LTD.",
        "generate_resignation_letter": "Yes",
        "notes": "M02 pressure test: cooperative transfer-in with optional secretary resignation letter.",
    }
    for key, value in values.items():
        set_field(ws, key, value)

    rows = [
        ("Yes", "resign_secretary", "OLD SECRETARY PERSON", "04/06/2026", "Yes", "Optional resignation letter should be generated."),
        ("Yes", "appoint_secretary", "NEW SECRETARY PERSON", "04/06/2026", "No", "Appointment item remains available for M01 ordinary DR."),
    ]
    start_row = 40
    for offset, row_values in enumerate(rows):
        row_no = start_row + offset
        for col_no, value in enumerate(row_values, start=1):
            ws.cell(row_no, col_no).value = value

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
