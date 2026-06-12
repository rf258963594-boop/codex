from __future__ import annotations

import http.cookiejar
import json
import re
import shutil
import sqlite3
import subprocess
import sys
import time
import urllib.parse
import urllib.request
import uuid
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
OUT_DIR = ROOT / "outputs"
GENERATED_DIR = APP_DIR / "generated"
DB_PATH = APP_DIR / "data" / "secretary_files.db"
BASE_URL = "http://127.0.0.1:8088"

TEMPLATE = ROOT / "templates" / "import" / "P2_maintenance_annual_blank_v7.xlsx"
OUTPUT_XLSX = OUT_DIR / "EASTSEED_transfer_in_noncooperative_M01_M02_input_v1.xlsx"
AUDIT_JSON = OUT_DIR / "EASTSEED_transfer_in_noncooperative_audit_v1.json"
BUNDLE_ZIP = OUT_DIR / "EASTSEED_transfer_in_noncooperative_test_bundle_v1.zip"
EXTRACT_DIR = OUT_DIR / "EASTSEED_transfer_in_noncooperative_extracted"


COMPANY_FIELDS: dict[str, Any] = {
    "company_name": "EASTSEED PTE. LTD.",
    "uen": "201941693Z",
    "registered_office_address": "60 PAYA LEBAR ROAD, #11-53, PAYA LEBAR SQUARE, SINGAPORE 409051",
    "total_issued_shares": 30000,
    "issued_share_capital": 30000,
    "paid_up_capital": 30000,
    "currency": "SGD",
    "default_document_date": "",
    "director_signer_names": "XIE YU\nYAN ZHY WEI",
    "member_signer_names": "XIE YU",
    "client_signatory_name": "XIE YU",
    "business_order_id": "EASTSEED-TRANSFER-IN-NONCOOP-TEST-001",
    "source_type": "BizFile + manual test",
    "source_file_id": "17.3.25bizfile- EASTSEED PTE. LTD..pdf",
    "prepared_by": "Codex manual flow test",
    "change_registered_office_required": "",
    "new_registered_office_address": "",
    "transfer_in_required": "Yes",
    "transfer_in_mode": "",
    "old_secretary_company": "",
    "new_secretary_company": "RSIN GROUP PTE. LTD.",
    "generate_resignation_letter": "No",
    "annual_review_required": "No",
    "notes": (
        "Manual flow simulation for transfer-in with M01 execution documents and M02 fallback authority. "
        "Existing local director SITOH MON FUI RICKY and secretary ZHU YIMIN are replaced. "
        "New nominee/local resident director LE THI NGOC TRANG and secretary "
        "FENDI CHANDRA TING S ING EE are pulled from website common people. "
        "Document dates and transfer-in mode intentionally left blank to test simplified defaults."
    ),
}

PERSON_ACTION_ROWS = [
    {
        "generate": "Yes",
        "action_type": "resign_director",
        "target_name": "SITOH MON FUI RICKY",
        "effective_date": "",
        "resignation_letter": "No",
        "remarks": "Existing local/nominee director per BizFile; M01 outputs normally, no separate resignation letter requested.",
    },
    {
        "generate": "Yes",
        "action_type": "appoint_director",
        "target_name": "LE THI NGOC TRANG",
        "effective_date": "",
        "resignation_letter": "No",
        "remarks": "New nominee/local resident director; use backend common person record.",
    },
    {
        "generate": "Yes",
        "action_type": "resign_secretary",
        "target_name": "ZHU YIMIN",
        "effective_date": "",
        "resignation_letter": "No",
        "remarks": "Existing individual secretary per BizFile; M01 outputs normally, no separate resignation letter requested.",
    },
    {
        "generate": "Yes",
        "action_type": "appoint_secretary",
        "target_name": "FENDI CHANDRA TING S ING EE",
        "effective_date": "",
        "resignation_letter": "No",
        "remarks": "New individual secretary; use backend common person record.",
    },
]


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def header_field_key(value: Any) -> str:
    raw = clean(value)
    if not raw:
        return ""
    for part in reversed(re.split(r"[\r\n]+", raw)):
        candidate = part.strip().strip("()[] ")
        if re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", candidate):
            return candidate
    compact = re.sub(r"[^A-Za-z0-9_]", "", raw)
    return compact if re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", compact) else raw


def set_field(ws, key: str, value: Any) -> None:
    for row in range(1, ws.max_row + 1):
        if clean(ws.cell(row, 3).value) == key:
            ws.cell(row, 4).value = value
            ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="top")
            return
    raise KeyError(f"Field not found: {key}")


def find_person_action_header_row(ws) -> tuple[int, list[str]]:
    required = {"generate", "action_type", "target_name"}
    for row in range(1, ws.max_row + 1):
        keys = [header_field_key(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if required.issubset(set(keys)):
            return row, keys
    raise KeyError("Cannot find person action section")


def build_workbook() -> Path:
    wb = load_workbook(TEMPLATE)
    ws = wb.worksheets[0]

    for key, value in COMPANY_FIELDS.items():
        set_field(ws, key, value)

    header_row, headers = find_person_action_header_row(ws)
    next_section_row = ws.max_row + 1
    for row in range(header_row + 1, ws.max_row + 1):
        first_value = clean(ws.cell(row, 1).value)
        row_keys = {header_field_key(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        if row > header_row + 1 and first_value and {"generate", "field_label", "old_value", "new_value"}.issubset(row_keys):
            next_section_row = row - 1
            break
        if row > header_row + 1 and first_value and not any(clean(ws.cell(row, col).value) for col in range(2, ws.max_column + 1)):
            next_section_row = row
            break

    for row in range(header_row + 1, next_section_row):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None

    for offset, record in enumerate(PERSON_ACTION_ROWS, start=1):
        row = header_row + offset
        for col, key in enumerate(headers, start=1):
            if key in record:
                cell = ws.cell(row, col)
                cell.value = record[key]
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    return OUTPUT_XLSX


def common_people() -> dict[str, dict[str, Any]]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM common_people WHERE active = 1").fetchall()
    return {
        row["display_name"]: {
            "full_name": row["display_name"],
            "default_role": row["default_role"],
            "id_type": row["id_type"],
            "id_number": row["id_number"],
            "nationality": row["nationality"],
            "residential_address": row["residential_address"],
            "email": row["email"],
            "phone": row["phone"],
            "is_local_resident_director": bool(row["is_local_resident_director"]),
        }
        for row in rows
    }


def local_parse_check(path: Path) -> dict[str, Any]:
    sys.path.insert(0, str(APP_DIR))
    from excel_parser import parse_excel
    from rules import suggest_files

    parsed = parse_excel(path, common_people())
    suggestions = suggest_files(parsed)
    return {"parsed": parsed, "suggestions": suggestions}


def login_opener() -> urllib.request.OpenerDirector:
    cookies = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cookies))
    data = urllib.parse.urlencode({"username": "admin", "password": "admin123"}).encode()
    opener.open(
        urllib.request.Request(
            BASE_URL + "/login",
            data=data,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        ),
        timeout=30,
    ).read()
    return opener


def upload_workbook(opener: urllib.request.OpenerDirector, path: Path) -> int:
    boundary = "----codex" + uuid.uuid4().hex
    body = b"".join(
        [
            f"--{boundary}\r\n".encode(),
            f'Content-Disposition: form-data; name="file"; filename="{path.name}"\r\n'.encode(),
            b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n",
            path.read_bytes(),
            f"\r\n--{boundary}--\r\n".encode(),
        ]
    )
    response = opener.open(
        urllib.request.Request(
            BASE_URL + "/upload",
            data=body,
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        ),
        timeout=90,
    )
    response.read()
    match = re.search(r"id=(\d+)", response.geturl())
    if not match:
        raise RuntimeError(f"Cannot determine uploaded job id from {response.geturl()}")
    return int(match.group(1))


def job_row(job_id: int) -> dict[str, Any]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT * FROM generation_jobs WHERE id = ?", (job_id,)).fetchone()
    if not row:
        raise RuntimeError(f"Job not found: {job_id}")
    return dict(row)


def post_generate(opener: urllib.request.OpenerDirector, endpoint: str, job_id: int) -> None:
    data = urllib.parse.urlencode({"job_id": str(job_id)}).encode()
    opener.open(
        urllib.request.Request(
            BASE_URL + "/" + endpoint,
            data=data,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        ),
        timeout=60,
    ).read()


def wait_for_zip(job_id: int, expected_status: str, suffix: str, timeout_seconds: int = 240) -> Path:
    deadline = time.time() + timeout_seconds
    latest_status = ""
    zip_path = None
    while time.time() < deadline:
        row = job_row(job_id)
        latest_status = clean(row.get("status"))
        zip_path = GENERATED_DIR / f"{row['job_code']}_{suffix}_pdf_package.zip"
        if latest_status == "generation_failed":
            raise RuntimeError(f"Generation failed for {suffix}")
        if latest_status == expected_status and zip_path.exists():
            return zip_path
        time.sleep(2)
    raise TimeoutError(f"Timed out waiting for {suffix}; latest status={latest_status}, zip={zip_path}")


def extract_zip(zip_path: Path, target: Path) -> list[Path]:
    if target.exists():
        shutil.rmtree(target)
    target.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(target)
    return sorted(target.glob("*.pdf"))


def pdf_to_text(pdf_path: Path) -> str:
    pdftotext = shutil.which("pdftotext")
    if not pdftotext:
        return ""
    text_path = pdf_path.with_suffix(".txt")
    subprocess.run([pdftotext, "-layout", "-nopgbrk", str(pdf_path), str(text_path)], check=True)
    return text_path.read_text(encoding="utf-8", errors="ignore")


def normalized(text: str) -> str:
    return re.sub(r"\s+", " ", text).upper()


def audit_pdfs(m01_pdfs: list[Path], m02_pdfs: list[Path]) -> dict[str, Any]:
    m01_text = "\n".join(pdf_to_text(path) for path in m01_pdfs)
    m02_text = "\n".join(pdf_to_text(path) for path in m02_pdfs)
    all_text = normalized(m01_text + "\n" + m02_text)
    checks = {
        "company_name": "EASTSEED PTE. LTD." in all_text,
        "uen": "201941693Z" in all_text,
        "old_registered_address": "60 PAYA LEBAR ROAD" in all_text,
        "new_registered_address": "111 NORTH BRIDGE ROAD" in all_text,
        "new_nominee_director": "LE THI NGOC TRANG" in all_text,
        "new_secretary": "FENDI CHANDRA TING S ING EE" in all_text or "FENDI CHANDRA TING SING EE" in all_text,
        "old_local_director": "SITOH MON FUI RICKY" in all_text,
        "old_secretary": "ZHU YIMIN" in all_text,
        "client_director_signer_xie": "XIE YU" in all_text,
        "client_director_signer_yan": "YAN ZHY WEI" in all_text,
        "old_provider_default_phrase": "PREVIOUS CORPORATE SECRETARIAL SERVICE PROVIDER AND/OR THE FORMER COMPANY SECRETARY" in all_text,
        "new_provider": "RSIN GROUP PTE. LTD." in all_text,
        "no_outgoing_wording": "OUTGOING" not in all_text,
        "no_incoming_wording": "INCOMING" not in all_text,
        "no_non_cooperative_wording": "NON-COOPERATIVE" not in all_text and "NON_COOPERATIVE" not in all_text and "UNCOOPERATIVE" not in all_text,
        "no_placeholders": "{{" not in all_text and "}}" not in all_text and "[[" not in all_text and "]]" not in all_text,
    }
    return {
        "checks": checks,
        "missing_or_failed_checks": [key for key, ok in checks.items() if not ok],
        "m01_text_preview": m01_text[:2500],
        "m02_text_preview": m02_text[:2500],
    }


def create_bundle(paths: list[Path], audit_path: Path) -> Path:
    if BUNDLE_ZIP.exists():
        BUNDLE_ZIP.unlink()
    with zipfile.ZipFile(BUNDLE_ZIP, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in paths:
            if path.exists():
                zf.write(path, arcname=path.name)
        zf.write(audit_path, arcname=audit_path.name)
    return BUNDLE_ZIP


def main() -> None:
    workbook_path = build_workbook()
    parse_check = local_parse_check(workbook_path)
    opener = login_opener()
    job_id = upload_workbook(opener, workbook_path)
    uploaded_row = job_row(job_id)

    post_generate(opener, "generate-p2-m01", job_id)
    m01_zip = wait_for_zip(job_id, "p2_m01_pdf_generated", "P2_M01")
    post_generate(opener, "generate-p2-m02", job_id)
    m02_zip = wait_for_zip(job_id, "p2_m02_pdf_generated", "P2_M02")

    m01_pdfs = extract_zip(m01_zip, EXTRACT_DIR / "M01")
    m02_pdfs = extract_zip(m02_zip, EXTRACT_DIR / "M02")
    pdf_audit = audit_pdfs(m01_pdfs, m02_pdfs)
    audit = {
        "input_workbook": str(workbook_path),
        "job_id": job_id,
        "job_code": uploaded_row["job_code"],
        "job_url": f"{BASE_URL}/job?id={job_id}",
        "m01_zip": str(m01_zip),
        "m02_zip": str(m02_zip),
        "m01_pdfs": [str(path) for path in m01_pdfs],
        "m02_pdfs": [str(path) for path in m02_pdfs],
        "local_parse_summary": parse_check["suggestions"].get("summary", {}),
        "local_preview": parse_check["suggestions"].get("preview", []),
        "pdf_audit": pdf_audit,
    }
    AUDIT_JSON.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    bundle = create_bundle([workbook_path, m01_zip, m02_zip], AUDIT_JSON)
    audit["bundle_zip"] = str(bundle)
    AUDIT_JSON.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
