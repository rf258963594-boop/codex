from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
sys.path.insert(0, str(APP_DIR))

from doc_generator import generate_p1_pdf_package  # noqa: E402
from excel_parser import parse_excel  # noqa: E402


TEMPLATE = ROOT / "outputs" / "AI适配_新公司注册资料模板_v3.1_人性化竖排_Auto版_空白.xlsx"
OUTPUT_XLSX = ROOT / "outputs" / "P1_v31_share_capital_partly_paid_test.xlsx"
JOB_CODE = "P1_V31_SHARE_CAPITAL_PARTLY_PAID_TEST"


def clean(value: Any) -> str:
    return str(value or "").strip()


def find_header_row(ws, key_name: str = "field_key") -> tuple[int, list[str]]:
    for row_no in range(1, min(ws.max_row, 15) + 1):
        headers = [clean(cell.value) for cell in ws[row_no]]
        if key_name in headers:
            return row_no, headers
    raise ValueError(f"Cannot find {key_name} in {ws.title}")


def fill_single_value_sheet(wb, sheet_name: str, values: dict[str, Any]) -> None:
    ws = wb[sheet_name]
    header_row, headers = find_header_row(ws)
    key_col = headers.index("field_key") + 1
    value_col = next(
        idx + 1
        for idx, header in enumerate(headers)
        if "Value" in header or "填写内容" in header
    )
    for row_no in range(header_row + 1, ws.max_row + 1):
        key = clean(ws.cell(row_no, key_col).value)
        if key in values:
            ws.cell(row_no, value_col).value = values[key]


def fill_transposed_sheet(wb, sheet_name: str, prefix: str, objects: list[dict[str, Any]]) -> None:
    ws = wb[sheet_name]
    header_row, headers = find_header_row(ws)
    key_col = headers.index("field_key") + 1
    object_cols = [idx for idx, header in enumerate(headers, start=1) if clean(header).startswith(prefix)]
    for row_no in range(header_row + 1, ws.max_row + 1):
        key = clean(ws.cell(row_no, key_col).value)
        if not key:
            continue
        for obj_idx, col_no in enumerate(object_cols):
            ws.cell(row_no, col_no).value = objects[obj_idx].get(key, "") if obj_idx < len(objects) else ""


def main() -> None:
    wb = load_workbook(TEMPLATE)
    fill_single_value_sheet(
        wb,
        "公司信息",
        {
            "company_name": "ACRA SHARE CAPITAL TEST PTE. LTD.",
            "business_order_id": "INC-SHARE-CAPITAL-TEST",
            "source_type": "Excel",
            "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
            "incorporation_date": "01/06/2026",
            "business_activity_1": "Management consultancy services",
            "ssic_code_1": "70201",
            "currency": "SGD",
            "share_class_default": "Ordinary",
            "client_signatory_person_id": "P002",
            "remarks": "Stress test: 1,000,000 issued shares, issued share capital SGD 1,000,000, paid-up share capital SGD 1.",
        },
    )
    fill_transposed_sheet(
        wb,
        "人员信息",
        "人员",
        [
            {
                "person_id": "P001",
                "source": "new",
                "full_name": "TAN LOCAL NOMINEE",
                "id_type": "NRIC",
                "id_number": "S1234567A",
                "nationality": "Singaporean",
                "residential_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
                "email": "tan.local.nominee@example.com",
                "phone": "+65 9000 0001",
                "is_director": "Yes",
                "is_local_resident_director": "Yes",
                "is_nominee_director": "Yes",
            },
            {
                "person_id": "P002",
                "source": "new",
                "full_name": "ZHANG CLIENT DIRECTOR",
                "id_type": "Passport",
                "id_number": "P2000002B",
                "nationality": "Chinese",
                "date_of_birth": "02/02/1990",
                "residential_address": "ROOM 1808, BUILDING 1, INTERNATIONAL COMMERCE CENTRE, BEIJING 100020, CHINA",
                "email": "zhang.client.director@example.com",
                "phone": "+86 138 0000 0002",
                "is_director": "Yes",
            },
            {
                "person_id": "P003",
                "source": "new",
                "full_name": "LIM COMPANY SECRETARY",
                "id_type": "NRIC",
                "id_number": "S7654321B",
                "nationality": "Singaporean",
                "residential_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
                "email": "lim.secretary@example.com",
                "phone": "+65 9000 0003",
                "is_secretary": "Yes",
            },
        ],
    )
    fill_transposed_sheet(
        wb,
        "股东与股份",
        "股东",
        [
            {
                "shareholder_type": "person",
                "person_full_name": "ZHANG CLIENT DIRECTOR",
                "person_id_number": "P2000002B",
                "share_class": "Ordinary",
                "shares": 1_000_000,
                "issued_share_capital": 1_000_000,
                "paid_amount": 1,
                "currency": "SGD",
                "remarks": "Partly paid capital stress test.",
            }
        ],
    )
    fill_single_value_sheet(
        wb,
        "输出设置",
        {
            "output_format": "docx_pdf_zip",
            "prepared_by": "Codex share capital test",
            "client_signatory_logic": "shareholder_1",
        },
    )
    wb.save(OUTPUT_XLSX)
    parsed = parse_excel(OUTPUT_XLSX, {})
    zip_path = generate_p1_pdf_package(parsed, JOB_CODE)
    print(OUTPUT_XLSX)
    print(zip_path)
    print(parsed["shareholders"][0])


if __name__ == "__main__":
    main()
