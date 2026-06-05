from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"

TITLE_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FILL = PatternFill("solid", fgColor="1F4D78")
KEY_FILL = PatternFill("solid", fgColor="EEF2F7")
INPUT_FILL = PatternFill("solid", fgColor="FFF8D6")
NOTE_FILL = PatternFill("solid", fgColor="F8FAFC")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
KEY_FONT = Font(color="475569", size=9)
NOTE_FONT = Font(color="334155")
WRAP = Alignment(wrap_text=True, vertical="top")

AUTO_YES_NO = '"Auto,Yes,No"'
SOURCE_OPTIONS = '"new,common"'
SHAREHOLDER_TYPES = '"person,corporate"'
EVENT_TYPES = '"change_registered_office,change_office_hours,change_business_activity,change_fye,update_officer_particulars,appoint_director,resign_director,appoint_secretary,resign_secretary,transfer_in_cooperative,transfer_in_non_cooperative,share_transfer_approval,share_allotment_approval,bizfile_authorization,change_company_name,remove_director,strike_off"'
APPROVAL_ROUTES = '"auto,DR,EGM,WR,EGM+DR,manual"'
CONSIDERATION_BASIS = '"internal_paid_up_basis,acra_paid_up_capital_basis,stamp_duty_higher_of_price_or_nav"'
ANNUAL_ROUTES = '"ordinary_agm,exempt_private_company,dormant_company,manual"'


company_fields = [
    ("任务类型", "Task type", "task_type", "maintenance", "Yes", "固定值，不要修改。"),
    ("业务单编号", "Business order ID", "business_order_id", "", "Optional", "可空；以后对接客户系统/电子签名时使用。"),
    ("数据来源", "Source type", "source_type", "Excel", "Optional", "Excel / BizFile / AI / Manual。"),
    ("来源文件编号", "Source file ID", "source_file_id", "", "Optional", "可填 BizFile/护照/旧表等来源文件编号或备注。"),
    ("公司名称", "Company name", "company_name", "", "Yes", "BizFile 读取或人工填写。"),
    ("UEN", "UEN", "uen", "", "Recommended", "现有公司建议填写。"),
    ("当前注册地址", "Current registered office", "registered_office_address", "", "Recommended", ""),
    ("当前秘书公司", "Current secretary firm", "current_secretary_company", "", "Optional", "转入/交接时使用。"),
    ("默认文件日期", "Default document date", "default_document_date", "", "Recommended", "DD/MM/YYYY；正式文书自动转为 1st JUNE 2026 这类格式。"),
    ("联系人", "Contact person_id", "contact_person_id", "", "Optional", "可空；代理注册/维护时引用人员信息里的 person_id。"),
    ("代理/代办人", "Agent person_id", "agent_person_id", "", "Optional", "可空；联系人不是董事/股东时使用。"),
    ("客户方默认签字人", "Default client signer person_id", "client_signatory_person_id", "", "Optional", "留空默认第一个客户股东/董事。"),
    ("授权代表", "Authorised rep person_id", "authorized_representative_person_id", "", "Optional", "可空；企业股东或代理场景使用。"),
    ("董事默认签字人", "Default director signer person_id", "director_signer_person_id", "", "Optional", "留空默认所有董事签 M01；后续网页可覆盖。"),
    ("股东默认签字人", "Default shareholder signer person_id", "shareholder_signer_person_id", "", "Optional", "留空默认第一个股东。"),
    ("经办人", "Prepared by", "prepared_by", "", "Optional", ""),
    ("内部备注", "Internal notes", "notes", "", "Optional", "不进入正式文件。"),
]

people_fields = [
    ("人员编号", "Person ID", "person_id", "建议 P001/P002；给变更事项引用。"),
    ("来源", "Source", "source", "new=手填；common=调用后台常用人员。"),
    ("常用人员名称", "Common person", "common_person_name", "source=common 时填写后台保存的名称。"),
    ("英文姓名", "Full name", "full_name", "自然人/签字人英文姓名；调用常用人员时可空。"),
    ("证件类型", "ID type", "id_type", "Passport / NRIC / FIN。"),
    ("证件号", "ID number", "id_number", ""),
    ("国籍", "Nationality", "nationality", ""),
    ("住址", "Residential address", "residential_address", "长地址可换行。"),
    ("邮箱", "Email", "email", ""),
    ("电话", "Phone", "phone", ""),
    ("当前身份", "Current roles", "current_roles", "可填 Director / Secretary / Shareholder 等，便于 AI 和人工核对。"),
    ("变更后身份", "New roles", "new_roles", "任免或转入后身份。"),
    ("是否董事", "Director", "is_director", "Auto/Yes/No；客户董事建议明确 Yes。"),
    ("是否本地居民董事", "Local resident director", "is_local_resident_director", "Auto/Yes/No；常用挂名董事可 Auto。"),
    ("是否挂名董事", "Nominee director", "is_nominee_director", "Auto/Yes/No；常用挂名董事可 Auto。"),
    ("是否秘书", "Secretary", "is_secretary", "Auto/Yes/No；常用秘书可 Auto。"),
    ("是否股东", "Shareholder", "is_shareholder", "Auto/Yes/No；系统会按股东现状表自动判断。"),
    ("客户方代表签字", "Client representative signer", "is_client_signatory", "Auto/Yes/No；用于客户方代表，不控制法定角色文件。"),
    ("任命/生效日期", "Effective date", "effective_date", "DD/MM/YYYY；留空用默认文件日期。"),
    ("备注", "Remarks", "remarks", ""),
]

shareholding_fields = [
    ("股东编号", "Shareholder ID", "shareholder_id", "建议 SH001/SH002；给转股引用。"),
    ("股东类型", "Shareholder type", "shareholder_type", "person / corporate。"),
    ("关联人员编号", "Person ID", "person_id", "自然人股东建议填，对应人员信息。"),
    ("股东名称", "Shareholder name", "shareholder_name", ""),
    ("证件/注册号", "ID / Registration no.", "id_or_reg_no", ""),
    ("股份类别", "Share class", "share_class", "留空默认 Ordinary。"),
    ("当前股数", "Shares", "shares", ""),
    ("实缴金额", "Paid amount", "paid_amount", "通常按 ACRA 登记价值。"),
    ("币种", "Currency", "currency", "留空默认 SGD。"),
    ("证书编号", "Certificate no.", "certificate_no", ""),
    ("是否当前股东", "Current shareholder", "is_current", "Auto/Yes/No。"),
    ("备注", "Remarks", "remarks", ""),
]

change_fields = [
    ("事项编号", "Event ID", "event_id", "建议 EV001/EV002。"),
    ("事项类型", "Event type", "event_type", "从下拉选择；普通 DR/M01 事项可合并。"),
    ("中文事项名", "Chinese event name", "event_name_cn", "给人工复核显示用。"),
    ("是否生成", "Generate", "generate", "Auto/Yes/No；Auto 表示按填写内容判断。"),
    ("生效日期", "Effective date", "effective_date", "DD/MM/YYYY。"),
    ("审批方式", "Approval route", "approval_route", "auto/DR/EGM/WR/EGM+DR/manual。"),
    ("文件组", "Document group", "document_group", "同组普通 DR 会合并，例如 DR-001。"),
    ("合并进 DR", "Combine in DR", "combine_in_dr", "Auto/Yes/No；普通 M01 事项可 Auto。"),
    ("目标人员编号", "Target person ID", "target_person_id", "任免/资料更新时引用 P001。"),
    ("目标人员名称", "Target name", "target_name", "没有 person_id 时填写。"),
    ("原信息", "Old value", "old_value", ""),
    ("新信息", "New value", "new_value", ""),
    ("是否出辞职信", "Resignation letter", "resignation_letter", "Auto/Yes/No；转入包内可选。"),
    ("人工复核", "Manual review required", "manual_review_required", "Auto/Yes/No；高风险事项建议 Yes。"),
    ("备注", "Remarks", "remarks", ""),
]

transfer_fields = [
    ("转股编号", "Transfer ID", "transfer_id", "建议 TR001。"),
    ("是否生成", "Generate", "generate", "Auto/Yes/No。"),
    ("转出股东编号", "Transferor shareholder ID", "transferor_shareholder_id", "引用股东现状 SH001。"),
    ("转出方名称", "Transferor name", "transferor_name", ""),
    ("转入股东编号", "Transferee shareholder ID", "transferee_shareholder_id", "引用股东现状 SH002。"),
    ("转入方名称", "Transferee name", "transferee_name", ""),
    ("股份类别", "Share class", "share_class", "Ordinary。"),
    ("转让股数", "Shares transferred", "shares_transferred", ""),
    ("转让日期", "Transfer date", "transfer_date", "DD/MM/YYYY。"),
    ("对价口径", "Consideration basis", "consideration_basis", "默认 internal_paid_up_basis；旧公司/NAV/印花税场景人工复核。"),
    ("现金对价", "Cash consideration", "cash_consideration", "多数内部转让可空。"),
    ("旧证书编号", "Old certificate no.", "old_certificate_no", ""),
    ("新证书编号", "New certificate no.", "new_certificate_no", ""),
    ("印花税复核", "Stamp duty review", "stamp_duty_review", "Auto/Yes/No。"),
    ("备注", "Remarks", "remarks", ""),
]

allotment_fields = [
    ("配股编号", "Allotment ID", "allotment_id", "建议 AL001。"),
    ("是否生成", "Generate", "generate", "Auto/Yes/No。"),
    ("认购人员编号", "Allottee person ID", "allottee_person_id", "自然人可引用 P001。"),
    ("认购人名称", "Allottee name", "allottee_name", ""),
    ("股份类别", "Share class", "share_class", "Ordinary。"),
    ("配发股数", "Shares allotted", "shares_allotted", ""),
    ("每股实缴", "Amount paid per share", "amount_paid_per_share", ""),
    ("总实缴", "Total paid", "total_paid", ""),
    ("币种", "Currency", "currency", "SGD。"),
    ("配股日期", "Allotment date", "allotment_date", "DD/MM/YYYY。"),
    ("授权日期", "Authority date", "authority_date", "DD/MM/YYYY；S161/股东授权日期。"),
    ("是否 Form 24", "Form 24 required", "form24_required", "Auto/Yes/No；增资配股才需要。"),
    ("证书编号", "Certificate no.", "certificate_no", ""),
    ("备注", "Remarks", "remarks", ""),
]

annual_fields = [
    ("是否做年审", "Annual review required", "annual_review_required", "", "Optional", "Yes 时生成年审包；M01 先只做预览。"),
    ("财政年结日", "Financial year end", "fye_date", "", "Recommended", "DD/MM/YYYY。"),
    ("财年开始日", "Financial year start", "financial_year_start", "", "Optional", "可由系统/财报判断。"),
    ("AGM 日期", "AGM date", "agm_date", "", "Recommended", "DD/MM/YYYY。"),
    ("AGM 时间", "AGM time", "agm_time", "10.00 a.m.", "Optional", ""),
    ("AGM 地点", "AGM place", "agm_place", "", "Optional", "留空默认注册地址或书面决议逻辑。"),
    ("年审方式", "AGM route", "agm_route", "ordinary_agm", "Recommended", "ordinary_agm / exempt_private_company / dormant_company / manual。"),
    ("财报状态", "Accounts status", "accounts_status", "", "Optional", "dormant / non_dormant / unaudited / audited。"),
    ("审计豁免/小公司", "Audit exemption", "audit_exemption", "", "Optional", "审计师文件暂不生成。"),
    ("董事费", "Directors' fees", "directors_fee", "0", "Optional", "默认 0。"),
    ("董事薪酬", "Directors' remuneration", "directors_remuneration", "0", "Optional", "默认 0。"),
    ("AR 授权签字人", "AR authorised signer person_id", "ar_authorized_signer_person_id", "", "Optional", "留空默认客户董事/股东。"),
    ("是否生成 MRL", "Management representation letter", "management_rep_letter", "Yes", "Optional", "年审包后续接入。"),
    ("是否生成 CDD", "Client due diligence form", "cdd_form", "No", "Optional", "后续可做 KYC 模块。"),
    ("年审备注", "Annual review notes", "remarks", "", "Optional", ""),
]

output_fields = [
    ("文件包", "Package", "package", "AUTO_MAINTENANCE", "System", "网站自动判断多个文件包。"),
    ("输出格式", "Output format", "output_format", "docx_pdf_zip", "System", "目前生成 PDF 包。"),
    ("审批方式覆盖", "Approval route override", "approval_route_override", "auto", "Optional", "特殊才填 DR/EGM/WR/manual。"),
    ("转入模式", "Transfer-in mode", "transfer_in_mode", "auto", "Optional", "auto / cooperative / non_cooperative。"),
    ("股份动作", "Share action", "share_action", "none", "Optional", "none / share_transfer / share_allotment。"),
    ("生成内部核对表", "Generate internal checklist", "generate_internal_checklist", "Yes", "Optional", ""),
    ("完整包", "Include full pack", "include_full_pack", "Yes", "Optional", "No 时后续可只出签字页。"),
    ("输出备注", "Output notes", "notes", "", "Optional", ""),
]


def title(ws, cells: str) -> None:
    for row in ws[cells]:
        for cell in row:
            cell.fill = TITLE_FILL
            cell.font = Font(color="FFFFFF", bold=True, size=15)
            cell.alignment = WRAP


def style_header(ws, row_no: int = 3) -> None:
    for cell in ws[row_no]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = WRAP


def style_common(ws) -> None:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
    for col in range(1, ws.max_column + 1):
        max_len = 12
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)) + 2, 44))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("填写说明")
    ws.merge_cells("A1:C1")
    ws["A1"] = "公司维护/变更/年审资料导入模板 v4 - 字段竖排版"
    ws.append([])
    rows = [
        ("主题", "中文说明", "English notes"),
        ("核心逻辑", "字段放左边竖排；人员、股东、事项、转股、增资按列展开。", "Fields are vertical; objects are columns."),
        ("P2 入口", "上传本表后网站自动判断 M01/M02/M03/M04/年审等文件包；网页选项只做覆盖。", "Upload first; the website detects packages."),
        ("M01 当前能力", "普通董事决议 M01 已可生成；M02 转入、M03 转股、M04 增资和年审包先做预览。", "M01 can generate now; other packages are preview-first."),
        ("Auto/Yes/No", "Auto 或空白表示系统判断；Yes 强制生成/选中；No 明确排除。", "Auto/blank lets the system decide; No is an override."),
        ("日期格式", "表格建议用 DD/MM/YYYY；系统兼容 YYYY-MM-DD；正式文书自动转成 1st JUNE 2026。", "Use DD/MM/YYYY; formal documents render day-month-year wording."),
        ("常用人员", "秘书/挂名董事可填 source=common，再填 common_person_name，系统会从后台补资料。", "Use source=common for saved people."),
        ("DR 合并", "同一 document_group 且可合并的普通事项会合并进一份 M01。", "Ordinary events in the same document_group merge into one M01."),
        ("转入", "转入统一按 M02 包处理；交接信和辞职信是包内选项。", "Transfer-in belongs to M02 with optional letters."),
        ("转股", "股份转让不生成 Form 24；Form 24 只用于增资/配股。", "Share transfers do not generate Form 24; allotments do."),
        ("字段保护", "不要改 sheet 名称和 field_key；黄色区域是填写区。", "Do not rename sheets or field_key values."),
    ]
    for row in rows:
        ws.append(row)
    title(ws, "A1:C1")
    style_header(ws, 3)
    style_common(ws)
    ws.freeze_panes = "A3"


def build_vertical_sheet(wb: Workbook, name: str, heading: str, fields: list[tuple], sample: dict[str, Any] | None = None) -> None:
    sample = sample or {}
    ws = wb.create_sheet(name)
    ws.merge_cells("A1:F1")
    ws["A1"] = heading
    ws.append([])
    ws.append(["项目", "English", "field_key", "填写内容 / Value", "必填", "说明"])
    for cn, en, key, default, required, note in fields:
        ws.append([cn, en, key, sample.get(key, default), required, note])
    title(ws, "A1:F1")
    style_header(ws)
    for row in range(4, ws.max_row + 1):
        ws.cell(row, 3).fill = KEY_FILL
        ws.cell(row, 3).font = KEY_FONT
        ws.cell(row, 4).fill = INPUT_FILL
        ws.cell(row, 6).fill = NOTE_FILL
        ws.cell(row, 6).font = NOTE_FONT
    style_common(ws)
    ws.freeze_panes = "A4"


def build_transposed_sheet(wb: Workbook, name: str, heading: str, fields: list[tuple], prefix: str, count: int, sample_objects: list[dict[str, Any]] | None = None) -> None:
    sample_objects = sample_objects or []
    ws = wb.create_sheet(name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + count)
    ws["A1"] = heading
    ws.append([])
    ws.append(["项目", "English", "field_key", "说明", *[f"{prefix}{i}" for i in range(1, count + 1)]])
    for cn, en, key, note in fields:
        values = [sample_objects[i].get(key, "") if i < len(sample_objects) else "" for i in range(count)]
        ws.append([cn, en, key, note, *values])
    title(ws, f"A1:{get_column_letter(4 + count)}1")
    style_header(ws)
    for row in range(4, ws.max_row + 1):
        ws.cell(row, 3).fill = KEY_FILL
        ws.cell(row, 3).font = KEY_FONT
        ws.cell(row, 4).fill = NOTE_FILL
        ws.cell(row, 4).font = NOTE_FONT
        for col in range(5, 5 + count):
            ws.cell(row, col).fill = INPUT_FILL
    style_common(ws)
    ws.freeze_panes = "E4"


def add_list(ws, row_key: str, values: str, start_col: int = 5, end_col: int = 14) -> None:
    key_col = 3
    target_row = None
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, key_col).value or "") == row_key:
            target_row = row
            break
    if not target_row:
        return
    dv = DataValidation(type="list", formula1=values, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(start_col)}{target_row}:{get_column_letter(end_col)}{target_row}")


def add_vertical_list(ws, row_key: str, values: str) -> None:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 3).value or "") == row_key:
            dv = DataValidation(type="list", formula1=values, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"D{row}")
            return


def apply_validations(wb: Workbook) -> None:
    people = wb["人员信息"]
    add_list(people, "source", SOURCE_OPTIONS)
    for key in ["is_director", "is_local_resident_director", "is_nominee_director", "is_secretary", "is_shareholder", "is_client_signatory"]:
        add_list(people, key, AUTO_YES_NO)

    shareholders = wb["股东现状"]
    add_list(shareholders, "shareholder_type", SHAREHOLDER_TYPES)
    add_list(shareholders, "is_current", AUTO_YES_NO)

    events = wb["变更事项"]
    add_list(events, "event_type", EVENT_TYPES)
    add_list(events, "generate", AUTO_YES_NO)
    add_list(events, "approval_route", APPROVAL_ROUTES)
    for key in ["combine_in_dr", "resignation_letter", "manual_review_required"]:
        add_list(events, key, AUTO_YES_NO)

    transfers = wb["股份转让"]
    add_list(transfers, "generate", AUTO_YES_NO)
    add_list(transfers, "consideration_basis", CONSIDERATION_BASIS)
    add_list(transfers, "stamp_duty_review", AUTO_YES_NO)

    allotments = wb["增资配股"]
    add_list(allotments, "generate", AUTO_YES_NO)
    add_list(allotments, "form24_required", AUTO_YES_NO)

    annual = wb["年审信息"]
    add_vertical_list(annual, "annual_review_required", AUTO_YES_NO)
    add_vertical_list(annual, "agm_route", ANNUAL_ROUTES)
    add_vertical_list(annual, "management_rep_letter", AUTO_YES_NO)
    add_vertical_list(annual, "cdd_form", AUTO_YES_NO)


def sample_company() -> dict[str, Any]:
    return {
        "business_order_id": "MAINT-TEST-001",
        "source_type": "BizFile",
        "company_name": "SAMPLE MAINTENANCE PTE. LTD.",
        "uen": "202612345A",
        "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
        "current_secretary_company": "OLD SECRETARIAL FIRM PTE. LTD.",
        "default_document_date": "29/05/2026",
        "contact_person_id": "P001",
        "agent_person_id": "",
        "client_signatory_person_id": "P001",
        "authorized_representative_person_id": "",
        "director_signer_person_id": "P001",
        "shareholder_signer_person_id": "P001",
    }


def sample_people() -> list[dict[str, Any]]:
    return [
        {"person_id": "P001", "source": "new", "full_name": "ZHANG YI", "id_type": "Passport", "id_number": "E12345678", "nationality": "Chinese", "residential_address": "Shanghai, China", "email": "zhangyi@example.com", "phone": "+86 13800000001", "current_roles": "Director/Shareholder", "new_roles": "Director/Shareholder", "is_director": "Yes", "is_local_resident_director": "No", "is_nominee_director": "No", "is_secretary": "No", "is_shareholder": "Auto", "is_client_signatory": "Auto", "effective_date": "29/05/2026", "remarks": "客户董事股东"},
        {"person_id": "P002", "source": "new", "full_name": "ZHANG ER", "id_type": "Passport", "id_number": "E22345678", "nationality": "Chinese", "residential_address": "Beijing, China", "email": "zhanger@example.com", "phone": "+86 13800000002", "current_roles": "Shareholder", "new_roles": "Shareholder", "is_director": "No", "is_local_resident_director": "No", "is_nominee_director": "No", "is_secretary": "No", "is_shareholder": "Auto", "is_client_signatory": "Auto", "effective_date": "29/05/2026", "remarks": "新受让股东"},
        {"person_id": "P003", "source": "common", "common_person_name": "公司秘书 A", "current_roles": "", "new_roles": "Secretary", "is_director": "Auto", "is_local_resident_director": "Auto", "is_nominee_director": "Auto", "is_secretary": "Auto", "is_shareholder": "Auto", "is_client_signatory": "Auto", "effective_date": "29/05/2026", "remarks": "常用秘书"},
    ]


def sample_shareholdings() -> list[dict[str, Any]]:
    return [
        {"shareholder_id": "SH001", "shareholder_type": "person", "person_id": "P001", "shareholder_name": "ZHANG YI", "id_or_reg_no": "E12345678", "share_class": "Ordinary", "shares": 1000, "paid_amount": 1000, "currency": "SGD", "certificate_no": "001", "is_current": "Yes", "remarks": "当前股东"},
        {"shareholder_id": "SH002", "shareholder_type": "person", "person_id": "P002", "shareholder_name": "ZHANG ER", "id_or_reg_no": "E22345678", "share_class": "Ordinary", "shares": 0, "paid_amount": 0, "currency": "SGD", "certificate_no": "", "is_current": "No", "remarks": "转入后股东"},
    ]


def sample_events() -> list[dict[str, Any]]:
    return [
        {"event_id": "EV001", "event_type": "change_registered_office", "event_name_cn": "注册地址变更", "generate": "Auto", "effective_date": "29/05/2026", "approval_route": "DR", "document_group": "DR-001", "combine_in_dr": "Auto", "old_value": "Old address", "new_value": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098", "resignation_letter": "No", "manual_review_required": "No", "remarks": "M01 示例"},
        {"event_id": "EV002", "event_type": "appoint_secretary", "event_name_cn": "委任秘书", "generate": "Auto", "effective_date": "29/05/2026", "approval_route": "DR", "document_group": "DR-001", "combine_in_dr": "Auto", "target_person_id": "P003", "target_name": "公司秘书 A", "new_value": "Secretary", "resignation_letter": "No", "manual_review_required": "No", "remarks": "M01 示例"},
        {"event_id": "EV003", "event_type": "transfer_in_cooperative", "event_name_cn": "配合转入", "generate": "No", "effective_date": "29/05/2026", "approval_route": "EGM+DR", "document_group": "TAKEOVER-001", "combine_in_dr": "No", "old_value": "Old secretary firm", "new_value": "RSIN GROUP PTE. LTD.", "resignation_letter": "Auto", "manual_review_required": "Yes", "remarks": "M02 预览示例"},
    ]


def sample_transfers() -> list[dict[str, Any]]:
    return [
        {"transfer_id": "TR001", "generate": "No", "transferor_shareholder_id": "SH001", "transferor_name": "ZHANG YI", "transferee_shareholder_id": "SH002", "transferee_name": "ZHANG ER", "share_class": "Ordinary", "shares_transferred": 300, "transfer_date": "29/05/2026", "consideration_basis": "internal_paid_up_basis", "old_certificate_no": "001", "new_certificate_no": "002", "stamp_duty_review": "Auto", "remarks": "M03 预览示例"},
    ]


def sample_allotments() -> list[dict[str, Any]]:
    return [
        {"allotment_id": "AL001", "generate": "No", "allottee_person_id": "P002", "allottee_name": "ZHANG ER", "share_class": "Ordinary", "shares_allotted": 100, "amount_paid_per_share": 1, "total_paid": 100, "currency": "SGD", "allotment_date": "29/05/2026", "authority_date": "29/05/2026", "form24_required": "Auto", "certificate_no": "003", "remarks": "M04 预览示例"},
    ]


def sample_annual() -> dict[str, Any]:
    return {
        "annual_review_required": "No",
        "fye_date": "31/12/2025",
        "financial_year_start": "01/01/2025",
        "agm_date": "30/06/2026",
        "agm_place": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
        "accounts_status": "non_dormant",
        "ar_authorized_signer_person_id": "P001",
    }


def build_workbook(sample: bool) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb)
    build_vertical_sheet(wb, "维护公司信息", "维护公司信息 Company", company_fields, sample_company() if sample else {})
    build_transposed_sheet(wb, "人员信息", "人员信息 People", people_fields, "人员", 10, sample_people() if sample else [])
    build_transposed_sheet(wb, "股东现状", "股东现状 Shareholdings", shareholding_fields, "股东", 10, sample_shareholdings() if sample else [])
    build_transposed_sheet(wb, "变更事项", "变更事项 Change Events", change_fields, "事项", 12, sample_events() if sample else [])
    build_transposed_sheet(wb, "股份转让", "股份转让 Share Transfers", transfer_fields, "转股", 8, sample_transfers() if sample else [])
    build_transposed_sheet(wb, "增资配股", "增资配股 Share Allotments", allotment_fields, "增资", 8, sample_allotments() if sample else [])
    build_vertical_sheet(wb, "年审信息", "年审信息 Annual Review", annual_fields, sample_annual() if sample else {})
    build_vertical_sheet(wb, "输出设置", "输出设置 Output Options", output_fields, {})
    apply_validations(wb)
    return wb


def save_workbook(wb: Workbook, name: str) -> Path:
    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / name
    wb.save(path)
    return path


def main() -> None:
    paths = [
        save_workbook(build_workbook(False), "AI适配_公司维护变更年审资料模板_v4_竖排_Auto预留字段版_空白.xlsx"),
        save_workbook(build_workbook(True), "AI适配_公司维护变更年审资料模板_v4_竖排_Auto预留字段版_示例.xlsx"),
    ]
    for path in paths:
        print(path)


if __name__ == "__main__":
    main()
