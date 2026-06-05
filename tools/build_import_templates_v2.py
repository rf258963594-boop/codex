from __future__ import annotations

from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


YES_NO = '"Yes,No"'
TASK_TYPES = '"incorporation,change"'
SOURCE_TYPES = '"new,common"'
PERSON_ACTIONS = '"no_change,appoint,resign,remove,update_particulars"'
SH_TYPES = '"person,corporate"'
VALUE_BASIS = '"acra_paid_up_capital_basis,stamp_duty_higher_of_price_or_nav,manual_override"'


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            width = min(max(len(str(cell.value)) + 2, 10), 38)
            widths[cell.column] = max(widths.get(cell.column, 0), width)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_sheet(wb: Workbook, title: str, rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def add_validation(ws, col_name: str, values: str, start_row: int = 2, end_row: int = 200) -> None:
    headers = [cell.value for cell in ws[1]]
    if col_name not in headers:
        return
    col = headers.index(col_name) + 1
    dv = DataValidation(type="list", formula1=values, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}")


def add_common_validations(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for col in [
            "required",
            "change_required",
            "combine_in_same_dr",
            "manual_review_required",
            "signing_required",
            "is_director",
            "is_local_resident_director",
            "is_nominee_director",
            "is_secretary",
            "is_shareholder",
            "is_authorized_rep",
        ]:
            add_validation(ws, col, YES_NO)
        add_validation(ws, "task_type", TASK_TYPES)
        add_validation(ws, "source", SOURCE_TYPES)
        add_validation(ws, "position_action", PERSON_ACTIONS)
        add_validation(ws, "shareholder_type", SH_TYPES)
        add_validation(ws, "transferor_type", SH_TYPES)
        add_validation(ws, "transferee_type", SH_TYPES)
        add_validation(ws, "value_basis", VALUE_BASIS)


def build_registration(blank: bool) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(
        wb,
        "README",
        [
            ["项目 / Item", "中文说明", "English notes"],
            ["用途", "新公司注册资料导入；上传网站后生成 P1 注册签字文件。", "Import data for incorporation package generation."],
            ["正式模板", "正式上传版不要保留示例行；不知道的字段留空，不要编造。", "For live upload, keep unknown fields blank and do not invent values."],
            ["常用人员", "挂名董事/秘书可填 source=common，再填 common_person_name，网站会用后台常用人员库补资料。", "Use source=common for nominee directors/secretaries saved in website settings."],
            ["自动计算", "FYE、首个财年结束日、证书编号、总股数/总实缴可由网页端按数据自动计算。", "FYE, certificate numbers, and share totals can be calculated by the website."],
            ["签字默认", "客户方签字默认股东1；股权证书默认董事1+对应股东；Form 24 当前为公司一份列出全部股东。", "Client signer defaults to Shareholder 1; certificates use Director 1 plus the relevant shareholder."],
        ],
    )
    company_rows = [
        ["field_key", "中文名称", "English name", "value", "required", "AI/人工备注"],
        ["task_type", "任务类型", "Task type", "incorporation", "Yes", "固定填写 incorporation / Fixed value"],
        ["company_name", "拟注册公司名称", "Proposed company name", "", "Yes", "英文法定名称 / English legal name"],
        ["uen", "UEN/注册号", "UEN / Registration number", "", "No", "注册前可空；注册后补填可用于证书等文件。"],
        ["company_type", "公司类型", "Company type", "", "No", "留空默认 Private Company Limited by Shares"],
        ["registered_office_address", "注册地址", "Registered office address", "", "Yes", "英文地址；不要填中文地址。"],
        ["office_hours", "办公时间", "Office hours", "", "No", "留空默认 Monday to Friday, 9.00 a.m. to 5.00 p.m."],
        ["incorporation_date", "注册/文件日期", "Incorporation / document date", "", "Recommended", "建议填写 DD/MM/YYYY；系统兼容 YYYY-MM-DD；FYE 默认按此日期计算。"],
        ["document_date", "签字/文件日期", "Document date", "", "No", "P1 现统一使用 incorporation_date；此字段保留兼容旧表，但不会覆盖注册日期。"],
        ["first_fye", "首个财政年结日覆盖", "First FYE override", "", "No", "留空由网页端自动计算；特殊首个财年才填 DD/MM/YYYY。"],
        ["fye", "财政年度月份/年结日覆盖", "FYE override", "", "No", "留空由网页端自动计算；特殊 FYE 才填月份或 DD/MM/YYYY。"],
        ["business_activity_1", "主要业务描述", "Primary business activity", "", "Yes", ""],
        ["ssic_code_1", "主要 SSIC Code", "Primary SSIC code", "", "Recommended", ""],
        ["business_activity_2", "第二业务描述", "Secondary business activity", "", "No", ""],
        ["ssic_code_2", "第二 SSIC Code", "Secondary SSIC code", "", "No", ""],
        ["currency", "股本币种", "Share currency", "", "No", "留空默认 SGD"],
        ["total_issued_shares", "已发行股数备用", "Issued shares fallback", "", "No", "优先从 Shareholders.shares 合计；此处只作为校验/备用。"],
        ["paid_up_capital", "实缴股本备用", "Paid-up capital fallback", "", "No", "优先从 Shareholders.paid_amount 合计；此处只作为校验/备用。"],
        ["share_class_default", "默认股份类别", "Default share class", "", "No", "留空默认 Ordinary"],
        ["share_par_value", "每股登记价值覆盖", "Par/registered value per share override", "", "No", "留空通常按 paid_up_capital / shares 或 1 处理。"],
        ["register_location", "股东名册保存地址", "Register location", "", "No", "留空默认注册地址。"],
        ["remarks", "备注", "Remarks", "", "No", "内部备注，不进入正式文件。"],
    ]
    people_rows = [
        [
            "source",
            "common_person_name",
            "full_name",
            "id_type",
            "id_number",
            "nationality",
            "date_of_birth",
            "residential_address",
            "email",
            "phone",
            "is_director",
            "is_local_resident_director",
            "is_nominee_director",
            "is_secretary",
            "is_shareholder",
            "is_authorized_rep",
            "signing_required",
            "appointment_date",
            "remarks",
        ]
    ]
    shareholder_rows = [
        [
            "shareholder_type",
            "person_full_name",
            "person_id_number",
            "corporate_name",
            "corporate_registration_country",
            "corporate_registration_number",
            "corporate_registered_address",
            "authorized_rep_full_name",
            "share_class",
            "shares",
            "paid_amount",
            "currency",
            "certificate_no",
            "signing_required",
            "remarks",
        ]
    ]
    generation_rows = [
        ["option_key", "中文名称", "English name", "value", "说明 / Notes"],
        ["package", "文件包", "Package", "INCORPORATION_BASIC", "上传后网站自动判断；当前 P1 注册包。"],
        ["output_format", "输出格式", "Output format", "docx_pdf_zip", "docx / pdf / docx_pdf_zip"],
        ["prepared_by", "经办人", "Prepared by", "", "可空"],
        ["signing_mode", "签字模式", "Signing mode", "default", "默认即可；特殊流程再人工调整。"],
        ["client_signatory_logic", "客户方签字默认", "Client signer default", "shareholder_1", "说明项：默认股东1；后续网页可做覆盖选择。"],
        ["form24_mode", "Form 24 生成方式", "Form 24 mode", "one_company_file_all_shareholders", "说明项：当前一家公司一份，列出全部股东和签字块。"],
        ["share_certificate_signing_director", "股权证书董事签字", "Certificate director signer", "director_1", "说明项：当前默认第一个董事。"],
    ]
    if not blank:
        company_values = {
            "company_name": "LONGFIELD TEST HOLDINGS PTE. LTD.",
            "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
            "incorporation_date": "27/05/2026",
            "business_activity_1": "Management consultancy services",
            "ssic_code_1": "70201",
            "currency": "SGD",
        }
        for row in company_rows[1:]:
            row[3] = company_values.get(row[0], row[3])
        people_rows.extend(
            [
                ["common", "挂名董事 A", "", "", "", "", "", "", "", "", "Yes", "Yes", "Yes", "No", "No", "No", "Yes", "27/05/2026", "从网站常用人员读取"],
                ["common", "公司秘书 A", "", "", "", "", "", "", "", "", "No", "No", "No", "Yes", "No", "No", "Yes", "27/05/2026", "从网站常用人员读取"],
                ["new", "", "CLIENT SHAREHOLDER 1", "Passport", "P1000001A", "Chinese", "03/03/1988", "BEIJING, CHINA", "client1@example.com", "+86 13800000001", "Yes", "No", "No", "No", "Yes", "No", "Yes", "27/05/2026", "客户董事/股东"],
            ]
        )
        shareholder_rows.append(["person", "CLIENT SHAREHOLDER 1", "P1000001A", "", "", "", "", "", "Ordinary", 1000, 1000, "SGD", "", "Yes", "自然人股东示例；正式模板请删除示例行"])
    write_sheet(wb, "Company", company_rows)
    write_sheet(wb, "People", people_rows)
    write_sheet(wb, "Shareholders", shareholder_rows)
    write_sheet(wb, "Generation", generation_rows)
    add_common_validations(wb)
    return wb


def build_change(blank: bool) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(
        wb,
        "README",
        [
            ["项目 / Item", "中文说明", "English notes"],
            ["用途", "现有公司变更资料导入；先由 BizFile 读取当前资料，再填写要变更的新资料。", "Import current company data and intended changes."],
            ["DR 合并", "多个普通变更可用 combine_in_same_dr=Yes 放进同一份 Directors' Resolution。", "Use combine_in_same_dr=Yes to group ordinary changes into one DR."],
            ["SR/MR", "SR/MR 暂作人工复核字段；系统第一版只提示，不自动判断复杂法律争议。", "SR/MR remains a manual review flag in MVP."],
            ["转股价值", "ACRA 申报关注股数、币种、股份类别、paid-up share capital 等登记资料；IRAS/交易价值通常看实际价格或 NAV。", "ACRA filing data and IRAS/commercial consideration should not be mixed."],
        ],
    )
    write_sheet(
        wb,
        "Company_Current",
        [
            ["field_key", "中文名称", "English name", "current_value", "AI备注"],
            ["task_type", "任务类型", "Task type", "change", "固定填写 change"],
            ["company_name", "当前公司名称", "Current company name", "", "BizFile 读取"],
            ["uen", "UEN", "UEN", "", "BizFile 读取"],
            ["registered_office_address", "当前注册地址", "Current registered office", "", ""],
            ["office_hours", "当前办公时间", "Current office hours", "", "通常留空使用默认"],
            ["fye", "当前 FYE", "Current FYE", "", ""],
            ["business_activity_1", "当前主要业务描述", "Current primary activity", "", ""],
            ["ssic_code_1", "当前主要 SSIC Code", "Current primary SSIC", "", ""],
            ["business_activity_2", "当前第二业务描述", "Current secondary activity", "", ""],
            ["ssic_code_2", "当前第二 SSIC Code", "Current secondary SSIC", "", ""],
            ["currency", "币种", "Currency", "SGD", ""],
            ["total_issued_shares", "当前已发行股数", "Current issued shares", "", ""],
            ["paid_up_capital", "当前实缴股本总额", "Current paid-up capital", "", ""],
        ],
    )
    company_changes = [
        ["change_item", "中文项目", "English item", "current_value", "new_value", "effective_date", "change_required", "combine_in_same_dr", "resolution_group", "resolution_type_override", "manual_review_required", "AI备注"],
        ["company_name", "公司名称", "Company name", "", "", "", "No", "Yes", "DR-001", "", "No", "公司更名时填写 new_value。"],
        ["registered_office_address", "注册地址", "Registered office", "", "", "", "No", "Yes", "DR-001", "", "No", "注册地址变更。"],
        ["office_hours", "办公时间", "Office hours", "", "", "", "No", "Yes", "DR-001", "", "No", "通常留空使用默认工作时间。"],
        ["fye", "财政年结日", "Financial year end", "", "", "", "No", "Yes", "DR-001", "", "No", "FYE 变更。"],
        ["business_activity_1", "主要业务", "Primary activity", "", "", "", "No", "Yes", "DR-001", "", "No", "业务变更。"],
        ["ssic_code_1", "主要 SSIC", "Primary SSIC", "", "", "", "No", "Yes", "DR-001", "", "No", "SSIC 变更。"],
        ["business_activity_2", "第二业务", "Secondary activity", "", "", "", "No", "Yes", "DR-001", "", "No", ""],
        ["ssic_code_2", "第二 SSIC", "Secondary SSIC", "", "", "", "No", "Yes", "DR-001", "", "No", ""],
        ["director_change", "董事任免", "Director appointment/resignation", "", "", "", "No", "Yes", "DR-001", "", "Yes", "董事辞任/强制移除需人工复核。"],
        ["secretary_change", "秘书变更", "Secretary change", "", "", "", "No", "Yes", "DR-001", "", "Yes", "辞退旧秘书/争议情况人工复核。"],
    ]
    if not blank:
        for row in company_changes[1:7]:
            row[3] = "Current value"
            row[5] = "27/05/2026"
    write_sheet(wb, "Company_Changes", company_changes)
    people = [
        ["source", "common_person_name", "full_name", "id_type", "id_number", "nationality", "residential_address", "email", "phone", "current_roles", "new_roles", "position_action", "effective_date", "signing_required", "remarks"],
    ]
    if not blank:
        people.extend(
            [
                ["new", "", "CURRENT DIRECTOR", "Passport", "E12345678", "Chinese", "BEIJING, CHINA", "old@example.com", "+86...", "Director/Shareholder", "Director/Shareholder", "no_change", "27/05/2026", "Yes", "BizFile 当前人员"],
                ["common", "挂名董事 A", "", "", "", "", "", "", "", "", "Director/Nominee Director", "appoint", "27/05/2026", "Yes", "任命常用挂名董事"],
            ]
        )
    write_sheet(wb, "People", people)
    personal = [
        ["full_name", "id_number", "current_roles", "change_field", "old_value", "new_value", "effective_date", "change_required", "AI备注"],
        ["", "", "", "ID type", "", "", "", "No", "例：Passport 改 FIN"],
        ["", "", "", "ID number", "", "", "", "No", "例：护照号改 FIN No."],
        ["", "", "", "Residential address", "", "", "", "No", "住址变更"],
        ["", "", "", "Phone", "", "", "", "No", "电话变更"],
        ["", "", "", "Email", "", "", "", "No", "邮箱变更"],
    ]
    write_sheet(wb, "Personal_Changes", personal)
    write_sheet(
        wb,
        "Shareholders_Current",
        [
            ["shareholder_type", "person_full_name", "person_id_number", "corporate_name", "corporate_registration_country", "corporate_registration_number", "share_class", "current_shares", "current_paid_amount", "currency", "remarks"],
            *([] if blank else [["person", "CURRENT DIRECTOR", "E12345678", "", "", "", "Ordinary", 1000, 1000, "SGD", "BizFile 当前股东"]]),
        ],
    )
    write_sheet(
        wb,
        "Share_Changes",
        [
            ["change_required", "transferor_type", "transferor_name_or_company", "transferor_id_or_reg_no", "transferee_type", "transferee_name_or_company", "transferee_id_or_reg_no", "share_class", "shares_transferred", "effective_date", "value_basis", "cash_consideration_override", "estimated_company_net_asset_value", "remarks"],
            ["No", "", "", "", "", "", "", "Ordinary", "", "", "acra_paid_up_capital_basis", "", "", "默认用于 ACRA/BizFile 登记资料；如涉及真实交易价或印花税，改用 stamp_duty_higher_of_price_or_nav 并人工复核。"],
        ],
    )
    write_sheet(
        wb,
        "Generation",
        [
            ["option_key", "中文名称", "English name", "value", "说明 / Notes"],
            ["package", "文件包", "Package", "AUTO_CHANGE", "网站自动判断"],
            ["output_format", "输出格式", "Output format", "docx_pdf_zip", "docx / pdf / docx_pdf_zip"],
            ["prepared_by", "经办人", "Prepared by", "", "可空"],
            ["default_resolution_group", "默认决议合并", "Default DR grouping", "same_dr_by_default", "普通变更默认可合并；争议/强制/特殊事项人工复核。"],
            ["manual_review_note", "人工复核备注", "Manual review note", "", "有争议事项写这里"],
        ],
    )
    write_sheet(
        wb,
        "规则说明_Rules",
        [
            ["主题 / Topic", "中文说明", "English notes"],
            ["combine_in_same_dr", "Yes 表示可合并进同一份 DR；No 表示建议单独文件或人工判断。", "Yes means the item can be grouped into the same Directors' Resolution."],
            ["resolution_group", "同一组编号会进入同一份 DR，例如 DR-001。", "Items with the same group code can be drafted together."],
            ["resolution_type_override", "不确定 MR/SR 时留空；需要 SR/MR 时人工填写并复核。", "Leave blank unless a specific MR/SR route is confirmed."],
            ["value_basis: acra_paid_up_capital_basis", "用于 ACRA/BizFile 登记资料：按当前持股、转让股数及 paid-up share capital 记录填写，不代表真实买卖价。", "For ACRA/BizFile filing data: shares transferred and paid-up share capital records, not necessarily commercial price."],
            ["value_basis: stamp_duty_higher_of_price_or_nav", "涉及真实交易或印花税时，按实际价格或 NAV/股份价值中较高者复核；旧公司尤其要看资产/负债。", "For stamp duty/commercial review, check the higher of actual price and NAV/share value."],
        ],
    )
    add_common_validations(wb)
    return wb


def save_workbook(wb: Workbook, name: str) -> Path:
    path = OUT / name
    wb.save(path)
    return path


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    outputs = [
        save_workbook(build_registration(blank=True), "AI适配_新公司注册资料模板_v2_空白.xlsx"),
        save_workbook(build_registration(blank=False), "AI适配_新公司注册资料模板_v2_示例.xlsx"),
        save_workbook(build_change(blank=True), "AI适配_现有公司变更资料模板_v2_空白.xlsx"),
        save_workbook(build_change(blank=False), "AI适配_现有公司变更资料模板_v2_示例.xlsx"),
    ]
    for path in outputs:
        print(path)


if __name__ == "__main__":
    main()
