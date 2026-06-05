from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "outputs" / "AI适配_公司维护变更年审资料模板_v7_一页式快速业务单含快速年审_空白.xlsx"
OUTPUT = ROOT / "outputs" / "M04_share_allotment_stress_input.xlsx"


COMPANY_FIELDS = {
    "company_name": "RBIZ SHARE ALLOTMENT STRESS PTE. LTD.",
    "uen": "202613888N",
    "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A, PENINSULA PLAZA, SINGAPORE 179098",
    "total_issued_shares": 100000,
    "issued_share_capital": 100000,
    "paid_up_capital": 50000,
    "currency": "SGD",
    "default_document_date": "05/06/2026",
    "director_signer_names": "ZHANG FIRST CLIENT DIRECTOR\nLEE SECOND CLIENT DIRECTOR\nONG THIRD CLIENT DIRECTOR",
    "member_signer_names": "ZHANG FIRST CLIENT SHAREHOLDER\nLEE SECOND CLIENT SHAREHOLDER",
    "client_signatory_name": "ZHANG FIRST CLIENT SHAREHOLDER",
    "business_order_id": "M04-STRESS-001",
    "source_type": "AI",
    "source_file_id": "BizFile + allotment instruction draft",
    "prepared_by": "Codex test",
    "change_registered_office_required": "No",
    "change_business_activity_required": "No",
    "change_fye_required": "No",
    "transfer_in_required": "No",
    "annual_review_required": "No",
    "notes": "M04 pressure test: two allottees, long address, full paid and partly paid share capital, multi-director and multi-member signatures.",
}


ALLOTMENT_ROWS = [
    {
        "generate": "Yes",
        "allottee_name": "GAMMA VENTURES PTE. LTD.",
        "allottee_id_number": "202377773K",
        "allottee_address": "8 MARINA BOULEVARD, #33-01 MARINA BAY FINANCIAL CENTRE TOWER 1, SINGAPORE 018981",
        "shares_allotted": 40000,
        "share_class": "Ordinary",
        "issued_share_capital": 40000,
        "amount_paid_per_share": 1,
        "total_paid": 40000,
        "currency": "SGD",
        "allotment_date": "05/06/2026",
        "authority_date": "05/06/2026",
        "form24_required": "Auto",
        "certificate_no": "004",
        "generate_certificate": "Yes",
        "post_allotment_total_shares": 160000,
        "post_allotment_issued_share_capital": 200000,
        "post_allotment_paid_up_capital": 110000,
        "remarks": "Corporate allottee; fully paid new shares.",
    },
    {
        "generate": "Yes",
        "allottee_name": "TAN PARTLY PAID INVESTOR",
        "allottee_id_number": "S7654321C",
        "allottee_address": "20 CECIL STREET, #14-01 PLUS BUILDING, SINGAPORE 049705",
        "shares_allotted": 20000,
        "share_class": "Ordinary",
        "issued_share_capital": 60000,
        "amount_paid_per_share": 0.5,
        "total_paid": 10000,
        "currency": "SGD",
        "allotment_date": "05/06/2026",
        "authority_date": "05/06/2026",
        "form24_required": "Auto",
        "certificate_no": "",
        "generate_certificate": "Auto",
        "post_allotment_total_shares": 160000,
        "post_allotment_issued_share_capital": 200000,
        "post_allotment_paid_up_capital": 110000,
        "remarks": "Partly paid scenario; certificate number intentionally blank.",
    },
]


def set_kv(ws, key: str, value) -> None:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 3).value or "").strip() == key:
            ws.cell(row, 4).value = value
            return
    raise KeyError(f"Field key not found: {key}")


def find_section(ws, section_name: str) -> int:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() == section_name:
            return row
    raise KeyError(f"Section not found: {section_name}")


def header_keys(ws, header_row: int) -> dict[str, int]:
    keys = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if not value:
            continue
        key = str(value).splitlines()[-1].strip()
        keys[key] = col
    return keys


def clear_table_rows(ws, first_row: int, last_row: int, max_col: int) -> None:
    for row in range(first_row, last_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def main() -> None:
    wb = load_workbook(TEMPLATE)
    ws = wb.worksheets[0]
    for key, value in COMPANY_FIELDS.items():
        set_kv(ws, key, value)

    section_row = find_section(ws, "增资配股")
    header_row = section_row + 1
    first_data_row = header_row + 1
    keys = header_keys(ws, header_row)
    clear_table_rows(ws, first_data_row, ws.max_row, ws.max_column)

    for offset, data in enumerate(ALLOTMENT_ROWS):
        row = first_data_row + offset
        for key, value in data.items():
            if key in keys:
                ws.cell(row, keys[key]).value = value

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
