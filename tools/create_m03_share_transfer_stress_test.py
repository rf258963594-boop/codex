from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "outputs" / "AI适配_公司维护变更年审资料模板_v7_一页式快速业务单含快速年审_空白.xlsx"
OUTPUT = ROOT / "outputs" / "M03_share_transfer_stress_input.xlsx"


COMPANY_FIELDS = {
    "company_name": "RBIZ SHARE TRANSFER STRESS PTE. LTD.",
    "uen": "202612999Z",
    "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A, PENINSULA PLAZA, SINGAPORE 179098",
    "default_document_date": "05/06/2026",
    "director_signer_names": "ZHANG FIRST CLIENT DIRECTOR\nLEE SECOND CLIENT DIRECTOR\nONG THIRD CLIENT DIRECTOR",
    "member_signer_names": "ZHANG FIRST CLIENT SHAREHOLDER\nLEE SECOND CLIENT SHAREHOLDER",
    "client_signatory_name": "ZHANG FIRST CLIENT SHAREHOLDER",
    "business_order_id": "M03-STRESS-001",
    "source_type": "AI",
    "source_file_id": "BizFile + transfer instruction draft",
    "prepared_by": "Codex test",
    "change_registered_office_required": "No",
    "change_business_activity_required": "No",
    "change_fye_required": "No",
    "transfer_in_required": "No",
    "annual_review_required": "No",
    "notes": "M03 pressure test: two transfers, corporate party, long addresses, multi-director signatures.",
}

TRANSFER_ROWS = [
    {
        "generate": "Yes",
        "transferor_name": "ALPHA HOLDING PTE. LTD.",
        "transferor_id_number": "202355551M",
        "transferor_address": "8 MARINA BOULEVARD, #33-01 MARINA BAY FINANCIAL CENTRE TOWER 1, SINGAPORE 018981",
        "transferee_name": "ZHANG FIRST CLIENT SHAREHOLDER",
        "transferee_id_number": "E12345678",
        "transferee_address": "20 CECIL STREET, #14-01 PLUS BUILDING, SINGAPORE 049705",
        "shares_transferred": 50000,
        "share_class": "Ordinary",
        "transfer_date": "05/06/2026",
        "consideration_basis": "stamp_duty_higher_of_price_or_nav",
        "consideration_amount": "1",
        "currency": "SGD",
        "old_certificate_no": "001",
        "new_certificate_no": "003",
        "transferor_remaining_shares": 50000,
        "generate_new_certificate": "Yes",
        "stamp_duty_review": "Yes",
        "remarks": "Corporate transferor; NAV/stamp duty review should be flagged.",
    },
    {
        "generate": "Yes",
        "transferor_name": "LEE SECOND CLIENT SHAREHOLDER",
        "transferor_id_number": "G1234567A",
        "transferor_address": "50 RAFFLES PLACE, #37-00 SINGAPORE LAND TOWER, SINGAPORE 048623",
        "transferee_name": "BETA INVESTMENT PTE. LTD.",
        "transferee_id_number": "202366662N",
        "transferee_address": "3 TEMASEK AVENUE, #21-00 CENTENNIAL TOWER, SINGAPORE 039190",
        "shares_transferred": 25000,
        "share_class": "Ordinary",
        "transfer_date": "05/06/2026",
        "consideration_basis": "internal_paid_up_basis",
        "consideration_amount": "",
        "currency": "SGD",
        "old_certificate_no": "002",
        "new_certificate_no": "",
        "transferor_remaining_shares": 75000,
        "generate_new_certificate": "Auto",
        "stamp_duty_review": "Auto",
        "remarks": "New certificate number intentionally blank to test To be assigned logic.",
    },
]


def set_kv(ws, key: str, value) -> None:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 3).value or "").strip() == key:
            ws.cell(row, 4).value = value
            return
    raise KeyError(f"Field key not found: {key}")


def find_section(ws, section_name: str) -> int:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() == section_name:
            return row
    raise KeyError(f"Section not found: {section_name}")


def header_keys(ws, header_row: int) -> dict[str, int]:
    keys = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if not value:
            continue
        key = str(value).splitlines()[-1].strip()
        keys[key] = col
    return keys


def clear_table_rows(ws, first_row: int, last_row: int, max_col: int) -> None:
    for row in range(first_row, last_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def main() -> None:
    wb = load_workbook(TEMPLATE)
    ws = wb.worksheets[0]
    for key, value in COMPANY_FIELDS.items():
        set_kv(ws, key, value)

    section_row = find_section(ws, "股份转让")
    header_row = section_row + 1
    first_data_row = header_row + 1
    keys = header_keys(ws, header_row)
    next_section_row = find_section(ws, "增资配股")
    clear_table_rows(ws, first_data_row, next_section_row - 1, ws.max_column)

    for offset, data in enumerate(TRANSFER_ROWS):
        row = first_data_row + offset
        for key, value in data.items():
            if key in keys:
                ws.cell(row, keys[key]).value = value

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
