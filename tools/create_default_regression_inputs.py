from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = ROOT / "templates" / "import"
OUT_DIR = ROOT / "tests" / "fixtures" / "defaults"

P1_TEMPLATE = TEMPLATE_DIR / "P1_registration_blank_v3_1.xlsx"
P2_TEMPLATE = TEMPLATE_DIR / "P2_maintenance_annual_blank_v7.xlsx"


def clean(value: Any) -> str:
    return str(value or "").strip()


def find_header_row(ws, key_name: str = "field_key") -> tuple[int, list[str]]:
    for row_no in range(1, min(ws.max_row, 20) + 1):
        headers = [clean(cell.value) for cell in ws[row_no]]
        if key_name in headers:
            return row_no, headers
    raise ValueError(f"Cannot find {key_name} in {ws.title}")


def fill_single_value_sheet(wb, sheet_name: str, values: dict[str, Any]) -> None:
    ws = wb[sheet_name]
    header_row, headers = find_header_row(ws)
    key_col = headers.index("field_key") + 1
    value_col = next(idx + 1 for idx, header in enumerate(headers) if "Value" in header or "填写内容" in header)
    for row_no in range(header_row + 1, ws.max_row + 1):
        key = clean(ws.cell(row_no, key_col).value)
        if key in values:
            ws.cell(row_no, value_col).value = values[key]


def fill_transposed_sheet(wb, sheet_name: str, prefix: str, objects: list[dict[str, Any]]) -> None:
    ws = wb[sheet_name]
    header_row, headers = find_header_row(ws)
    key_col = headers.index("field_key") + 1
    object_cols = [idx for idx, header in enumerate(headers, start=1) if clean(header).startswith(prefix)]
    for row_no in range(header_row + 1, ws.max_row + 1):
        key = clean(ws.cell(row_no, key_col).value)
        if not key:
            continue
        for obj_idx, col_no in enumerate(object_cols):
            ws.cell(row_no, col_no).value = objects[obj_idx].get(key, "") if obj_idx < len(objects) else ""


def set_p2_value(ws, key: str, value: Any) -> None:
    for row_no in range(1, ws.max_row + 1):
        if clean(ws.cell(row_no, 3).value) == key:
            ws.cell(row_no, 4).value = value
            return
    raise KeyError(key)


def find_section(ws, name: str) -> int:
    for row_no in range(1, ws.max_row + 1):
        if clean(ws.cell(row_no, 1).value) == name:
            return row_no
    raise KeyError(name)


def header_keys(ws, header_row: int) -> dict[str, int]:
    keys: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = clean(ws.cell(header_row, col).value)
        if value:
            keys[value.splitlines()[-1].strip()] = col
    return keys


def clear_section(ws, section_name: str, next_section_name: str | None = None) -> tuple[int, dict[str, int]]:
    section_row = find_section(ws, section_name)
    header_row = section_row + 1
    first_row = header_row + 1
    last_row = ws.max_row if next_section_name is None else find_section(ws, next_section_name) - 1
    for row_no in range(first_row, last_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row_no, col).value = None
    return first_row, header_keys(ws, header_row)


def write_rows(ws, start_row: int, keys: dict[str, int], rows: list[dict[str, Any]]) -> None:
    for offset, data in enumerate(rows):
        row_no = start_row + offset
        for key, value in data.items():
            if key in keys:
                ws.cell(row_no, keys[key]).value = value


def new_p2_workbook(company_name: str, uen: str) -> tuple[Any, Any]:
    wb = load_workbook(P2_TEMPLATE)
    ws = wb["P2快速业务单"]
    for key, value in {
        "company_name": company_name,
        "uen": uen,
        "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A, PENINSULA PLAZA, SINGAPORE 179098",
        "default_document_date": "06/06/2026",
        "business_order_id": company_name.replace(" ", "-")[:40],
        "source_type": "Default regression",
        "prepared_by": "Codex",
    }.items():
        set_p2_value(ws, key, value)
    return wb, ws


def create_p1_sparse() -> Path:
    wb = load_workbook(P1_TEMPLATE)
    fill_single_value_sheet(
        wb,
        "公司信息",
        {
            "company_name": "DEFAULT P1 SPARSE PTE. LTD.",
            "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
            "incorporation_date": "06/06/2026",
            "business_activity_1": "MANAGEMENT CONSULTANCY SERVICES",
            "ssic_code_1": "70201",
            "remarks": "Sparse default regression: currency, share class, paid-up and FYE fields intentionally blank.",
        },
    )
    fill_transposed_sheet(
        wb,
        "人员信息",
        "人员",
        [
            {
                "source": "new",
                "full_name": "DEFAULT CLIENT DIRECTOR",
                "id_type": "Passport",
                "id_number": "PDEFAULT001",
                "nationality": "Chinese",
                "residential_address": "ROOM 1808, TEST BUILDING, SHANGHAI, CHINA",
                "email": "default.director@example.com",
                "phone": "+86 138 0000 0000",
                "is_director": "Yes",
            },
            {
                "source": "new",
                "full_name": "DEFAULT COMPANY SECRETARY",
                "id_type": "NRIC",
                "id_number": "S1234567A",
                "nationality": "Singaporean",
                "residential_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
                "is_secretary": "Yes",
            },
            {
                "source": "new",
                "full_name": "DEFAULT NOMINEE DIRECTOR",
                "id_type": "NRIC",
                "id_number": "S7654321B",
                "nationality": "Singaporean",
                "residential_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
                "is_director": "Yes",
                "is_local_resident_director": "Yes",
                "is_nominee_director": "Yes",
            },
        ],
    )
    fill_transposed_sheet(
        wb,
        "股东与股份",
        "股东",
        [
            {
                "shareholder_type": "person",
                "person_full_name": "DEFAULT CLIENT DIRECTOR",
                "person_id_number": "PDEFAULT001",
                "shares": 100,
            }
        ],
    )
    output = OUT_DIR / "P1_sparse_defaults.xlsx"
    wb.save(output)
    return output


def create_m01_sparse() -> Path:
    wb, ws = new_p2_workbook("DEFAULT ORDINARY RESOLUTION SPARSE PTE. LTD.", "202600001A")
    for key, value in {
        "new_registered_office_address": "20 CECIL STREET, #12-02 PLUS, SINGAPORE 049705",
        "new_primary_ssic": "62011",
        "new_primary_activity": "DEVELOPMENT OF SOFTWARE AND APPLICATIONS EXCEPT GAMES",
        "new_secondary_ssic": "63119",
        "new_secondary_activity": "DATA PROCESSING, HOSTING AND RELATED ACTIVITIES",
    }.items():
        set_p2_value(ws, key, value)
    output = OUT_DIR / "ordinary_dr_sparse_defaults.xlsx"
    wb.save(output)
    return output


def create_m02_sparse() -> Path:
    wb, ws = new_p2_workbook("DEFAULT TRANSFER IN SPARSE PTE. LTD.", "202600002B")
    set_p2_value(ws, "transfer_in_required", "Yes")
    output = OUT_DIR / "transfer_in_sparse_defaults.xlsx"
    wb.save(output)
    return output


def create_m03_sparse() -> Path:
    wb, ws = new_p2_workbook("DEFAULT SHARE TRANSFER SPARSE PTE. LTD.", "202600003C")
    first_row, keys = clear_section(ws, "股份转让", "增资配股")
    write_rows(
        ws,
        first_row,
        keys,
        [
            {
                "transferor_name": "DEFAULT TRANSFEROR",
                "transferee_name": "DEFAULT TRANSFEREE",
                "shares_transferred": 10,
            }
        ],
    )
    output = OUT_DIR / "share_transfer_sparse_defaults.xlsx"
    wb.save(output)
    return output


def create_m04_sparse() -> Path:
    wb, ws = new_p2_workbook("DEFAULT SHARE ALLOTMENT SPARSE PTE. LTD.", "202600004D")
    first_row, keys = clear_section(ws, "增资配股")
    write_rows(
        ws,
        first_row,
        keys,
        [
            {
                "allottee_name": "DEFAULT ALLOTTEE",
                "shares_allotted": 25,
            }
        ],
    )
    output = OUT_DIR / "share_allotment_sparse_defaults.xlsx"
    wb.save(output)
    return output


def create_m05_sparse() -> Path:
    wb, ws = new_p2_workbook("DEFAULT ANNUAL REVIEW SPARSE PTE. LTD.", "202600005E")
    set_p2_value(ws, "fye_date", "31/12/2025")
    annual = wb["快速年审"]
    for row_no in range(1, annual.max_row + 1):
        if clean(annual.cell(row_no, 3).value) in {"annual_review_required", "agm_date", "director_signer_name", "shareholder_signer_name"}:
            annual.cell(row_no, 4).value = None
    output = OUT_DIR / "annual_review_sparse_defaults.xlsx"
    wb.save(output)
    return output


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    outputs = [
        create_p1_sparse(),
        create_m01_sparse(),
        create_m02_sparse(),
        create_m03_sparse(),
        create_m04_sparse(),
        create_m05_sparse(),
    ]
    for path in outputs:
        print(path)


if __name__ == "__main__":
    main()
