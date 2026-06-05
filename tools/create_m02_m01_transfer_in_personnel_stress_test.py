from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "outputs" / "AI适配_公司维护变更年审资料模板_v7_一页式快速业务单含快速年审_空白.xlsx"
OUTPUT = ROOT / "outputs" / "M02_M01_transfer_in_registered_office_secretary_director_stress.xlsx"


def set_field(ws, key: str, value: object) -> None:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 3).value or "").strip() == key:
            ws.cell(row, 4).value = value
            return
    raise KeyError(key)


def find_section(ws, name: str) -> int:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() == name:
            return row
    raise KeyError(name)


def main() -> None:
    wb = load_workbook(INPUT)
    ws = wb["P2快速业务单"]

    values = {
        "company_name": "RBIZ TRANSFER-IN ADDRESS STRESS PTE. LTD.",
        "uen": "202698765R",
        "registered_office_address": "10 ANSON ROAD, #20-05, INTERNATIONAL PLAZA, SINGAPORE 079903",
        "default_document_date": "04/06/2026",
        "director_signer_names": "ZHANG FIRST CLIENT DIRECTOR\nLEE SECOND CLIENT DIRECTOR\nONG THIRD CLIENT DIRECTOR",
        "member_signer_names": "ZHANG FIRST CLIENT SHAREHOLDER\nLEE SECOND CLIENT SHAREHOLDER",
        "client_signatory_name": "ZHANG FIRST CLIENT SHAREHOLDER",
        "business_order_id": "M02-M01-ADDRESS-STRESS-001",
        "source_type": "Manual / AI assisted",
        "source_file_id": "BizFile + prior secretary handover test",
        "prepared_by": "RSIN TEST USER",
        "transfer_in_required": "Yes",
        "transfer_in_mode": "cooperative",
        "old_secretary_company": "PREVIOUS CORPORATE SECRETARIAL SERVICES PTE. LTD.",
        "new_secretary_company": "RSIN GROUP PTE. LTD.",
        "change_registered_office_required": "Yes",
        "new_registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A, PENINSULA PLAZA, SINGAPORE 179098",
        "generate_resignation_letter": "Yes",
        "notes": "Pressure test: transfer-in with registered office change, old secretary resignation, new secretary appointment, old nominee director resignation and new local director appointment.",
    }
    for key, value in values.items():
        set_field(ws, key, value)

    section_row = find_section(ws, "人员任免")
    header_row = section_row + 1
    rows = [
        (
            "Yes",
            "resign_secretary",
            "OLD SECRETARY TAN WEI LING",
            "04/06/2026",
            "Yes",
            "Old individual secretary resignation letter should be attached to M02 second PDF.",
        ),
        (
            "Yes",
            "appoint_secretary",
            "NEW SECRETARY LIM HUI MIN",
            "04/06/2026",
            "No",
            "New individual secretary appointment should appear in M01 DR.",
        ),
        (
            "Yes",
            "resign_director",
            "OLD NOMINEE DIRECTOR CHAN KOK MING",
            "04/06/2026",
            "Yes",
            "Old nominee director resignation letter should be attached to M02 second PDF.",
        ),
        (
            "Yes",
            "appoint_director",
            "NEW LOCAL DIRECTOR ONG JIA HAO",
            "04/06/2026",
            "No",
            "New local resident director appointment should appear in M01 DR.",
        ),
    ]
    start_row = header_row + 1
    for offset, row_values in enumerate(rows):
        row_no = start_row + offset
        for col_no, value in enumerate(row_values, start=1):
            cell = ws.cell(row_no, col_no)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
