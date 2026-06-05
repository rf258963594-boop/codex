from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"

HEADER_FILL = PatternFill("solid", fgColor="17324D")
SECTION_FILL = PatternFill("solid", fgColor="EAF2F8")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

YES_NO = '"Yes,No"'
APPROVAL_ROUTES = '"auto,DR,EGM,WR,EGM+DR,manual"'
TRANSFER_IN_MODES = '"auto,cooperative,non_cooperative"'
SHARE_ACTIONS = '"none,share_transfer,share_allotment"'
ANNUAL_ROUTES = '"ordinary_agm,exempt_private_company,dormant_company,manual"'


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            width = min(max(len(str(cell.value)) + 2, 10), 42)
            widths[cell.column] = max(widths.get(cell.column, 0), width)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_sheet(wb: Workbook, title: str, rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def add_validation(ws, col_name: str, values: str, start_row: int = 2, end_row: int = 250) -> None:
    headers = [cell.value for cell in ws[1]]
    if col_name not in headers:
        return
    col = headers.index(col_name) + 1
    dv = DataValidation(type="list", formula1=values, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}")


def add_validations(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for col in [
            "generate",
            "combine_in_dr",
            "resignation_letter",
            "manual_review_required",
            "stamp_duty_review",
            "form24_required",
            "is_director",
            "is_local_resident_director",
            "is_nominee_director",
            "is_secretary",
            "is_shareholder",
            "is_client_signatory",
        ]:
            add_validation(ws, col, YES_NO)
        add_validation(ws, "approval_route", APPROVAL_ROUTES)
        add_validation(ws, "approval_route_override", APPROVAL_ROUTES)
        add_validation(ws, "transfer_in_mode", TRANSFER_IN_MODES)
        add_validation(ws, "share_action", SHARE_ACTIONS)
        add_validation(ws, "agm_route", ANNUAL_ROUTES)


def build_maintenance(blank: bool) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb,
        "README",
        [
            ["项目 / Item", "中文说明", "English notes"],
            ["用途", "现有公司维护业务单：可同时放年审、普通变更、转入、转股、增资。", "Maintenance order for annual review and company changes."],
            ["主逻辑", "上传后网站自动判断文件包；网页选项只做覆盖和确认，不作为主数据。", "Upload first; the website detects packages from the workbook."],
            ["DR 合并", "同一 document_group 且 combine_in_dr=Yes 的普通变更会合并进一份 DR。", "Ordinary changes in the same document_group can be grouped into one DR."],
            ["转入", "统一作为 EGM/WR 转入包处理；Termination/Handover Letter 和辞职信作为包内可选文件，由 resignation_letter 等字段控制。", "Transfer-in is handled inside the EGM/WR takeover pack; handover and resignation letters are optional outputs."],
            ["转股", "股份转让不生成 Form 24；Form 24 只用于增资/配股。", "Share transfer does not use Form 24; allotment does."],
            ["EROM", "EROM 是 ACRA 系统登记结果，不作为文件模板生成。", "EROM is not generated as a document template."],
            ["审计师", "审计师相关文件暂不纳入本系统。", "Auditor-related documents are out of scope for this phase."],
        ],
    )

    company_rows = [
        ["field_key", "中文名称", "English name", "value", "required", "AI/人工备注"],
        ["task_type", "任务类型", "Task type", "maintenance", "Yes", "固定填写 maintenance"],
        ["business_order_id", "业务单编号", "Business order ID", "", "No", "可空；以后对接客户系统时使用。"],
        ["source_type", "数据来源", "Source type", "Excel", "No", "Excel / BizFile / AI / Manual。"],
        ["source_file_id", "来源文件编号", "Source file ID", "", "No", "可填 BizFile/护照/旧表等来源文件编号或备注。"],
        ["company_name", "公司名称", "Company name", "", "Yes", "BizFile 读取或人工填写。"],
        ["uen", "UEN", "UEN", "", "Recommended", "现有公司建议必填。"],
        ["registered_office_address", "当前注册地址", "Current registered office", "", "Recommended", ""],
        ["current_secretary_company", "当前秘书公司", "Current secretary firm", "", "No", "转入/交接时使用。"],
        ["default_document_date", "默认文件日期", "Default document date", "", "Recommended", "DD/MM/YYYY；留空网页端用当天。"],
        ["prepared_by", "经办人", "Prepared by", "", "No", ""],
        ["contact_person_id", "联系人", "Contact person_id", "", "No", "可空；引用 People.person_id。"],
        ["agent_person_id", "代理/代办人", "Agent person_id", "", "No", "可空；引用 People.person_id。"],
        ["client_signatory_person_id", "客户方默认签字人", "Default client signer person_id", "", "No", "留空默认第一个客户股东/董事。"],
        ["authorized_representative_person_id", "授权代表", "Authorised rep person_id", "", "No", "可空；引用 People.person_id。"],
        ["director_signer_person_id", "董事默认签字人", "Default director signer person_id", "", "No", "留空默认第一个非挂名客户董事。"],
        ["shareholder_signer_person_id", "股东默认签字人", "Default shareholder signer person_id", "", "No", "留空默认第一个股东。"],
        ["notes", "备注", "Notes", "", "No", "内部备注，不进入正式文件。"],
    ]

    people_rows = [
        [
            "person_id",
            "source",
            "common_person_name",
            "full_name",
            "id_type",
            "id_number",
            "nationality",
            "residential_address",
            "email",
            "phone",
            "current_roles",
            "new_roles",
            "is_director",
            "is_local_resident_director",
            "is_nominee_director",
            "is_secretary",
            "is_shareholder",
            "is_client_signatory",
            "effective_date",
            "remarks",
        ]
    ]

    shareholding_rows = [
        [
            "shareholder_id",
            "shareholder_type",
            "person_id",
            "shareholder_name",
            "id_or_reg_no",
            "share_class",
            "shares",
            "paid_amount",
            "currency",
            "certificate_no",
            "is_current",
            "remarks",
        ]
    ]

    change_event_rows = [
        [
            "event_id",
            "event_type",
            "event_name_cn",
            "generate",
            "effective_date",
            "approval_route",
            "document_group",
            "combine_in_dr",
            "target_person_id",
            "target_name",
            "old_value",
            "new_value",
            "resignation_letter",
            "manual_review_required",
            "remarks",
        ]
    ]

    transfer_rows = [
        [
            "transfer_id",
            "generate",
            "transferor_shareholder_id",
            "transferor_name",
            "transferee_shareholder_id",
            "transferee_name",
            "share_class",
            "shares_transferred",
            "transfer_date",
            "consideration_basis",
            "cash_consideration",
            "old_certificate_no",
            "new_certificate_no",
            "stamp_duty_review",
            "remarks",
        ]
    ]

    allotment_rows = [
        [
            "allotment_id",
            "generate",
            "allottee_person_id",
            "allottee_name",
            "share_class",
            "shares_allotted",
            "amount_paid_per_share",
            "total_paid",
            "currency",
            "allotment_date",
            "authority_date",
            "form24_required",
            "certificate_no",
            "remarks",
        ]
    ]

    annual_rows = [
        ["field_key", "中文名称", "English name", "value", "required", "AI/人工备注"],
        ["annual_review_required", "是否做年审", "Annual review required", "", "No", "Yes 时生成年审包。"],
        ["fye_date", "财政年结日", "Financial year end", "", "Recommended", "DD/MM/YYYY"],
        ["financial_year_start", "财年开始日", "Financial year start", "", "No", "可由系统/财报判断；不知道可留空。"],
        ["agm_date", "AGM 日期", "AGM date", "", "Recommended", "DD/MM/YYYY"],
        ["agm_time", "AGM 时间", "AGM time", "10.00 a.m.", "No", ""],
        ["agm_place", "AGM 地点", "AGM place", "", "No", "留空默认注册地址或线上/书面决议逻辑。"],
        ["agm_route", "年审方式", "AGM route", "ordinary_agm", "Recommended", "ordinary_agm / exempt_private_company / dormant_company / manual"],
        ["accounts_status", "财报状态", "Accounts status", "", "No", "dormant / non_dormant / unaudited / audited"],
        ["audit_exemption", "审计豁免/小公司", "Audit exemption", "", "No", "审计师文件暂不生成。"],
        ["directors_fee", "董事费", "Directors' fees", "0", "No", "默认 0。"],
        ["directors_remuneration", "董事薪酬", "Directors' remuneration", "0", "No", "默认 0。"],
        ["ar_authorized_signer_person_id", "AR 授权签字人", "AR authorised signer person_id", "", "No", "留空默认客户董事/股东。"],
        ["management_rep_letter", "是否生成 MRL", "Management representation letter", "Yes", "No", ""],
        ["cdd_form", "是否生成 CDD", "Client due diligence form", "No", "No", "可后续单独做 KYC 模块。"],
        ["remarks", "年审备注", "Annual review notes", "", "No", ""],
    ]

    output_rows = [
        ["option_key", "中文名称", "English name", "value", "说明 / Notes"],
        ["package", "文件包", "Package", "AUTO_MAINTENANCE", "网站自动判断多个文件包。"],
        ["output_format", "输出格式", "Output format", "docx_pdf_zip", "docx / pdf / docx_pdf_zip"],
        ["approval_route_override", "审批方式覆盖", "Approval route override", "auto", "默认 auto；特殊才填 DR/EGM/WR/manual。"],
        ["transfer_in_mode", "转入模式", "Transfer-in mode", "auto", "auto / cooperative / non_cooperative"],
        ["share_action", "股份动作", "Share action", "none", "none / share_transfer / share_allotment"],
        ["generate_internal_checklist", "生成内部核对表", "Generate internal checklist", "Yes", ""],
        ["include_full_pack", "完整包", "Include full pack", "Yes", "No 时后续可只出签字页。"],
        ["notes", "输出备注", "Output notes", "", ""],
    ]

    rules_rows = [
        ["分类", "文件/逻辑", "适用事件", "默认签字人", "说明"],
        ["普通 DR 包", "Combined Directors' Resolution", "地址、营业范围、人员资料、秘书任免", "董事", "可合并成一份 DR。"],
        ["转入包", "EGM/WR + DR + Termination/Handover Letter + Optional Resignation Letter", "transfer_in_cooperative / transfer_in_non_cooperative", "股东 + 董事 + 可选离任人", "股东授权更稳；交接信、辞职信是转入包内可选文件，不单独做母版。不配合场景不要自动生成辞职信。"],
        ["转股包", "Transfer DR + Instrument of Transfer + Share Certificate", "ShareTransfers", "董事 + 转让双方", "不生成 Form 24。"],
        ["增资配股包", "S161/EGM/WR + Allotment DR + Share Application + Form 24 + Share Certificate", "ShareAllotments", "股东 + 董事 + 认购人", "这里才生成 Form 24。"],
        ["公司更名", "Special Resolution / EGM or WR", "change_company_name", "股东", "独立文件包，人工复核。"],
        ["年审包", "AGM Notice + Shorter Notice + Attendance + Minutes + AR Authorization", "AnnualReview", "董事 + 股东", "以 SINGERIA 样本为最终版基础。"],
        ["注销包", "Strike-off Pack", "strike_off", "董事 + 股东", "高风险，必须人工复核。"],
    ]

    options_rows = [
        ["字段", "可选值", "说明"],
        ["event_type", "change_registered_office", "注册地址变更"],
        ["event_type", "change_business_activity", "营业范围/SSIC 变更"],
        ["event_type", "update_officer_particulars", "董事/股东/秘书身份地址资料更新"],
        ["event_type", "appoint_director", "委任董事"],
        ["event_type", "resign_director", "董事辞任"],
        ["event_type", "remove_director", "股东决议移除董事"],
        ["event_type", "appoint_secretary", "委任秘书"],
        ["event_type", "resign_secretary", "秘书辞任"],
        ["event_type", "transfer_in_cooperative", "配合转入"],
        ["event_type", "transfer_in_non_cooperative", "不配合转入"],
        ["event_type", "change_company_name", "公司更名"],
        ["event_type", "strike_off", "注销/Strike off"],
        ["approval_route", "DR / EGM / WR / EGM+DR / manual", "留空或 auto 时由网站判断。"],
        ["consideration_basis", "internal_paid_up_basis", "内部/ACRA 登记口径，通常默认。"],
        ["consideration_basis", "stamp_duty_higher_of_price_or_nav", "涉及印花税/旧公司资产价值时人工复核。"],
    ]

    if not blank:
        company_values = {
            "business_order_id": "MAINT-TEST-001",
            "source_type": "BizFile",
            "company_name": "SAMPLE MAINTENANCE PTE. LTD.",
            "uen": "202612345A",
            "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
            "current_secretary_company": "OLD SECRETARIAL FIRM PTE. LTD.",
            "default_document_date": "29/05/2026",
            "contact_person_id": "P001",
            "client_signatory_person_id": "P001",
            "director_signer_person_id": "P001",
            "shareholder_signer_person_id": "P001",
        }
        for row in company_rows[1:]:
            row[3] = company_values.get(row[0], row[3])
        people_rows.extend(
            [
                ["P001", "new", "", "ZHANG YI", "Passport", "E12345678", "Chinese", "Shanghai, China", "zhangyi@example.com", "+86 13800000001", "Director/Shareholder", "Director/Shareholder", "Yes", "No", "No", "No", "Yes", "Yes", "29/05/2026", "客户董事股东"],
                ["P002", "new", "", "ZHANG ER", "Passport", "E22345678", "Chinese", "Beijing, China", "zhanger@example.com", "+86 13800000002", "Shareholder", "Shareholder", "No", "No", "No", "No", "Yes", "No", "29/05/2026", "新受让股东"],
                ["P003", "common", "公司秘书 A", "", "", "", "", "", "", "", "", "Secretary", "No", "No", "No", "Yes", "No", "No", "29/05/2026", "常用秘书"],
                ["P004", "new", "", "OLD SECRETARY", "NRIC", "S1234567A", "Singaporean", "Singapore", "", "", "Secretary", "", "No", "No", "No", "Yes", "No", "No", "29/05/2026", "旧秘书"],
            ]
        )
        shareholding_rows.extend(
            [
                ["SH001", "person", "P001", "ZHANG YI", "E12345678", "Ordinary", 1000, 1000, "SGD", "001", "Yes", "当前股东"],
                ["SH002", "person", "P002", "ZHANG ER", "E22345678", "Ordinary", 0, 0, "SGD", "", "No", "转入后股东"],
            ]
        )
        change_event_rows.extend(
            [
                ["EV001", "transfer_in_cooperative", "配合转入", "Yes", "29/05/2026", "EGM+DR", "TAKEOVER-001", "No", "", "", "Old secretary firm", "RSIN GROUP PTE. LTD.", "Yes", "Yes", "生成转入包和辞职信选项"],
                ["EV002", "change_registered_office", "注册地址变更", "Yes", "29/05/2026", "DR", "DR-001", "Yes", "", "", "Old address", "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098", "No", "No", ""],
                ["EV003", "appoint_secretary", "委任秘书", "Yes", "29/05/2026", "DR", "DR-001", "Yes", "P003", "公司秘书 A", "", "Secretary", "No", "No", ""],
            ]
        )
        transfer_rows.append(["TR001", "Yes", "SH001", "ZHANG YI", "SH002", "ZHANG ER", "Ordinary", 300, "29/05/2026", "internal_paid_up_basis", "", "001", "002", "Yes", "股份转让示例，不生成 Form 24"])
        allotment_rows.append(["AL001", "No", "", "", "Ordinary", "", "", "", "SGD", "", "", "Yes", "", "如做增资，把 generate 改 Yes"])
        annual_values = {
            "annual_review_required": "Yes",
            "fye_date": "31/12/2025",
            "financial_year_start": "01/01/2025",
            "agm_date": "30/06/2026",
            "agm_place": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
            "accounts_status": "non_dormant",
            "ar_authorized_signer_person_id": "P001",
        }
        for row in annual_rows[1:]:
            row[3] = annual_values.get(row[0], row[3])

    write_sheet(wb, "Company", company_rows)
    write_sheet(wb, "People", people_rows)
    write_sheet(wb, "Shareholdings", shareholding_rows)
    write_sheet(wb, "ChangeEvents", change_event_rows)
    write_sheet(wb, "ShareTransfers", transfer_rows)
    write_sheet(wb, "ShareAllotments", allotment_rows)
    write_sheet(wb, "AnnualReview", annual_rows)
    write_sheet(wb, "OutputOptions", output_rows)
    write_sheet(wb, "规则说明_Rules", rules_rows)
    write_sheet(wb, "字段说明_Options", options_rows)
    add_validations(wb)
    return wb


def save_workbook(wb: Workbook, name: str) -> Path:
    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / name
    wb.save(path)
    return path


def main() -> None:
    paths = [
        save_workbook(build_maintenance(blank=True), "AI适配_公司维护变更年审资料模板_v3_空白.xlsx"),
        save_workbook(build_maintenance(blank=False), "AI适配_公司维护变更年审资料模板_v3_示例.xlsx"),
    ]
    for path in paths:
        print(path)


if __name__ == "__main__":
    main()
