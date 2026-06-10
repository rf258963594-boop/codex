from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FILES = [
    ROOT / "outputs" / "AI适配_新公司注册资料模板_v3.1_人性化竖排_Auto版_空白.xlsx",
    ROOT / "outputs" / "P1_v31_registration_sample_share_capital.xlsx",
]


def has_border(cell) -> bool:
    border = cell.border
    return any(
        side is not None and side.style
        for side in (border.left, border.right, border.top, border.bottom)
    )


def main() -> None:
    for path in FILES:
        wb = load_workbook(path)
        checks = []
        for sheet_name in ["填写说明", "公司信息", "人员信息", "股东与股份"]:
            ws = wb[sheet_name]
            bordered = 0
            total = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        total += 1
                        if has_border(cell):
                            bordered += 1
            checks.append(f"{sheet_name}:{bordered}/{total}")
            if total and bordered < total * 0.75:
                raise RuntimeError(f"{path.name} {sheet_name} border coverage too low: {bordered}/{total}")
        print(f"{path.name} OK " + " ".join(checks))


if __name__ == "__main__":
    main()
