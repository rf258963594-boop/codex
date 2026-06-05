from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "outputs" / "AI适配_新公司注册资料模板.xlsx"
OUTPUT = ROOT / "outputs" / "stress_multi_director_shareholder_full_fields.xlsx"


def update_company(wb) -> None:
    ws = wb["Company"]
    company = {
        "task_type": "incorporation",
        "company_name": "LONGFIELD TEST HOLDINGS PTE. LTD.",
        "company_type": "Private Company Limited by Shares",
        "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098, WITH EXTRA LONG UNIT DESCRIPTION FOR WRAP TESTING",
        "office_hours": "Monday to Friday, 9.00 a.m. to 5.00 p.m., excluding Singapore public holidays",
        "incorporation_date": "27/05/2026",
        "first_fye": "2027-12-31",
        "fye": "2026-12-31",
        "business_activity_1": "Management consultancy services and digital operations advisory with intentionally long description for field wrapping test",
        "ssic_code_1": "70201",
        "business_activity_2": "Development of software and applications except games and cybersecurity, used only as test text",
        "ssic_code_2": "62011",
        "currency": "SGD",
        "total_issued_shares": 1000,
        "paid_up_capital": 1000,
        "share_class_default": "Ordinary",
        "remarks": "Stress test company record with complete simulated information and long text.",
        "uen": "202612345Z",
        "register_location": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
        "incorporation_certificate_no": "TEST-CERT-2026-0001",
        "company_name_chinese": "长字段测试控股私人有限公司",
        "registered_office_address_chinese": "新加坡北桥路111号 Peninsula Plaza #29-06A，附加长地址用于测试换行",
    }
    headers = [c.value for c in ws[1]]
    key_col = headers.index("field_key") + 1
    val_col = headers.index("value") + 1
    existing = {}
    for row_no in range(2, ws.max_row + 1):
        key = ws.cell(row_no, key_col).value
        if key:
            existing[str(key)] = row_no
    for key, value in company.items():
        if key in existing:
            ws.cell(existing[key], val_col).value = value
        else:
            row_no = ws.max_row + 1
            ws.cell(row_no, 1).value = key
            ws.cell(row_no, 2).value = key
            ws.cell(row_no, 3).value = value
            ws.cell(row_no, 4).value = "No"
            ws.cell(row_no, 5).value = "Extra stress-test field"


def update_people(wb) -> None:
    ws = wb["People"]
    headers = [c.value for c in ws[1]]
    clear_body(ws)
    rows = [
        {
            "source": "new",
            "common_person_name": "",
            "full_name": "LI GUAMING (李挂名)",
            "id_type": "NRIC",
            "id_number": "S8123456A",
            "nationality": "Singapore Citizen",
            "date_of_birth": "1981-01-11",
            "residential_address": "9 RAFFLES PLACE, #55-01 REPUBLIC PLAZA, SINGAPORE 048619",
            "email": "nominee.director.long-test@example-rsin.sg",
            "phone": "+65 9123 0001",
            "is_director": "Yes",
            "is_local_resident_director": "Yes",
            "is_nominee_director": "Yes",
            "is_secretary": "No",
            "is_shareholder": "No",
            "is_authorized_rep": "No",
            "signing_required": "Yes",
            "appointment_date": "27/05/2026",
            "remarks": "Nominee director full-field simulated record.",
        },
        {
            "source": "new",
            "common_person_name": "",
            "full_name": "CHEN MISHU (陈秘书)",
            "id_type": "FIN",
            "id_number": "G7654321B",
            "nationality": "Singapore Permanent Resident",
            "date_of_birth": "1985-02-22",
            "residential_address": "20 CECIL STREET, #26-02 PLUS BUILDING, SINGAPORE 049705",
            "email": "company.secretary.very.long.alias@example-rsin.sg",
            "phone": "+65 9123 0002",
            "is_director": "No",
            "is_local_resident_director": "No",
            "is_nominee_director": "No",
            "is_secretary": "Yes",
            "is_shareholder": "No",
            "is_authorized_rep": "No",
            "signing_required": "Yes",
            "appointment_date": "27/05/2026",
            "remarks": "Company secretary full-field simulated record.",
        },
        {
            "source": "new",
            "common_person_name": "",
            "full_name": "ZHANG YI (张一)",
            "id_type": "Passport",
            "id_number": "P1000001A",
            "nationality": "Chinese",
            "date_of_birth": "03/03/1988",
            "residential_address": "ROOM 1808, BUILDING 1, INTERNATIONAL COMMERCE CENTRE, 999 VERY LONG AVENUE NAME, CHAOYANG DISTRICT, BEIJING 100020, PEOPLE'S REPUBLIC OF CHINA",
            "email": "zhang.yi.client.director.shareholder.with.long.email@example-long-domain.cn",
            "phone": "+86 138 0000 0001",
            "is_director": "Yes",
            "is_local_resident_director": "No",
            "is_nominee_director": "No",
            "is_secretary": "No",
            "is_shareholder": "Yes",
            "is_authorized_rep": "Yes",
            "signing_required": "Yes",
            "appointment_date": "27/05/2026",
            "remarks": "Shareholder 1; default client-side signatory; long contact fields.",
        },
        {
            "source": "new",
            "common_person_name": "",
            "full_name": "ZHANG ER (张二)",
            "id_type": "Passport",
            "id_number": "P2000002B",
            "nationality": "Chinese",
            "date_of_birth": "04/04/1990",
            "residential_address": "UNIT 3201, TOWER B, SHENZHEN BAY TECHNOLOGY ECOLOGY PARK, NANSHAN DISTRICT, SHENZHEN, GUANGDONG 518057, CHINA, ADDITIONAL ADDRESS LINE FOR WRAP TEST",
            "email": "zhang.er.second.director.shareholder.long-address-test@example-long-domain.cn",
            "phone": "+86 139 0000 0002",
            "is_director": "Yes",
            "is_local_resident_director": "No",
            "is_nominee_director": "No",
            "is_secretary": "No",
            "is_shareholder": "Yes",
            "is_authorized_rep": "No",
            "signing_required": "Yes",
            "appointment_date": "27/05/2026",
            "remarks": "Shareholder 2; RORC because 30 percent shareholding.",
        },
        {
            "source": "new",
            "common_person_name": "",
            "full_name": "ZHANG SAN (张三)",
            "id_type": "Passport",
            "id_number": "P3000003C",
            "nationality": "Chinese",
            "date_of_birth": "05/05/1992",
            "residential_address": "FLAT 12A, 88 QUEENS ROAD CENTRAL, CENTRAL, HONG KONG, WITH LONG BUILDING NAME AND MULTIPLE DESCRIPTOR SEGMENTS FOR LINE WRAPPING TEST",
            "email": "zhang.san.third.director.shareholder.long-email@example-long-domain.hk",
            "phone": "+852 5123 0003",
            "is_director": "Yes",
            "is_local_resident_director": "No",
            "is_nominee_director": "No",
            "is_secretary": "No",
            "is_shareholder": "Yes",
            "is_authorized_rep": "No",
            "signing_required": "Yes",
            "appointment_date": "27/05/2026",
            "remarks": "Shareholder 3; below 25 percent RORC threshold.",
        },
    ]
    write_rows(ws, headers, rows)


def update_shareholders(wb) -> None:
    ws = wb["Shareholders"]
    headers = [c.value for c in ws[1]]
    clear_body(ws)
    rows = [
        {
            "shareholder_type": "person",
            "person_full_name": "ZHANG YI (张一)",
            "person_id_number": "P1000001A",
            "share_class": "Ordinary",
            "shares": 600,
            "paid_amount": 600,
            "currency": "SGD",
            "signing_required": "Yes",
            "remarks": "股东1，默认客户代表，60%持股，字段较长测试",
        },
        {
            "shareholder_type": "person",
            "person_full_name": "ZHANG ER (张二)",
            "person_id_number": "P2000002B",
            "share_class": "Ordinary",
            "shares": 300,
            "paid_amount": 300,
            "currency": "SGD",
            "signing_required": "Yes",
            "remarks": "股东2，30%持股，触发RORC",
        },
        {
            "shareholder_type": "person",
            "person_full_name": "ZHANG SAN (张三)",
            "person_id_number": "P3000003C",
            "share_class": "Ordinary",
            "shares": 100,
            "paid_amount": 100,
            "currency": "SGD",
            "signing_required": "Yes",
            "remarks": "股东3，10%持股，不触发RORC",
        },
    ]
    write_rows(ws, headers, rows)


def update_generation(wb) -> None:
    ws = wb["Generation"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "prepared_by":
            row[2].value = "Codex stress test"
        if row[0].value == "signing_mode":
            row[2].value = "default"


def clear_body(ws) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def write_rows(ws, headers: list[str], rows: list[dict[str, object]]) -> None:
    for row_no, item in enumerate(rows, start=2):
        for col_no, header in enumerate(headers, start=1):
            ws.cell(row_no, col_no).value = item.get(header, "")


def main() -> None:
    wb = load_workbook(TEMPLATE)
    ws = wb["README"]
    ws["A8"] = "测试说明"
    ws["B8"] = "1名挂名董事、1名秘书、3名客户董事兼股东；含长地址、长邮箱、长备注，用于排版压力测试。"
    update_company(wb)
    update_people(wb)
    update_shareholders(wb)
    update_generation(wb)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
