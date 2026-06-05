from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "outputs" / "AI适配_新公司注册资料模板_v3.1_人性化竖排_Auto版_空白.xlsx"
OUTPUT = ROOT / "outputs" / "P1_v31_registration_sample_share_capital.xlsx"


def clean(value: Any) -> str:
    return str(value or "").strip()


def find_header_row(ws, key_name: str = "field_key") -> tuple[int, list[str]]:
    for row_no in range(1, min(ws.max_row, 15) + 1):
        headers = [clean(cell.value) for cell in ws[row_no]]
        if key_name in headers:
            return row_no, headers
    raise ValueError(f"Cannot find {key_name} in {ws.title}")


def fill_single_value_sheet(wb, sheet_name: str, values: dict[str, Any]) -> None:
    ws = wb[sheet_name]
    header_row, headers = find_header_row(ws)
    key_col = headers.index("field_key") + 1
    value_col = next(idx + 1 for idx, header in enumerate(headers) if "Value" in header or "填写内容" in header)
    for row_no in range(header_row + 1, ws.max_row + 1):
        key = clean(ws.cell(row_no, key_col).value)
        if key in values:
            ws.cell(row_no, value_col).value = values[key]


def fill_transposed_sheet(wb, sheet_name: str, prefix: str, objects: list[dict[str, Any]]) -> None:
    ws = wb[sheet_name]
    header_row, headers = find_header_row(ws)
    key_col = headers.index("field_key") + 1
    object_cols = [idx for idx, header in enumerate(headers, start=1) if clean(header).startswith(prefix)]
    for row_no in range(header_row + 1, ws.max_row + 1):
        key = clean(ws.cell(row_no, key_col).value)
        if not key:
            continue
        for obj_idx, col_no in enumerate(object_cols):
            ws.cell(row_no, col_no).value = objects[obj_idx].get(key, "") if obj_idx < len(objects) else ""


def main() -> None:
    wb = load_workbook(TEMPLATE)
    fill_single_value_sheet(
        wb,
        "公司信息",
        {
            "company_name": "CORECHANGE SAMPLE PTE. LTD.",
            "business_order_id": "INC-SAMPLE-001",
            "source_type": "Excel",
            "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
            "incorporation_date": "01/06/2026",
            "business_activity_1": "Management consultancy services",
            "ssic_code_1": "70201",
            "business_activity_2": "Development of software and applications except games",
            "ssic_code_2": "62011",
            "currency": "SGD",
            "share_class_default": "Ordinary",
            "client_signatory_person_id": "P001",
            "remarks": "Sample: issued share capital and paid-up share capital are filled at shareholder level.",
        },
    )
    fill_transposed_sheet(
        wb,
        "人员信息",
        "人员",
        [
            {
                "person_id": "P001",
                "source": "new",
                "full_name": "ZHANG YI",
                "id_type": "Passport",
                "id_number": "P1000001A",
                "nationality": "Chinese",
                "date_of_birth": "03/03/1988",
                "residential_address": "ROOM 1808, BUILDING 1, INTERNATIONAL COMMERCE CENTRE, BEIJING 100020, CHINA",
                "email": "zhang.yi@example.com",
                "phone": "+86 138 0000 0001",
                "is_director": "Yes",
            },
            {
                "person_id": "P002",
                "source": "new",
                "full_name": "ZHANG ER",
                "id_type": "Passport",
                "id_number": "P2000002B",
                "nationality": "Chinese",
                "date_of_birth": "04/04/1990",
                "residential_address": "UNIT 3201, TOWER B, SHENZHEN BAY TECHNOLOGY PARK, SHENZHEN, CHINA",
                "email": "zhang.er@example.com",
                "phone": "+86 139 0000 0002",
                "is_director": "Yes",
            },
            {
                "person_id": "P003",
                "source": "new",
                "full_name": "TAN NOMINEE DIRECTOR",
                "id_type": "NRIC",
                "id_number": "S1234567A",
                "nationality": "Singaporean",
                "residential_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
                "is_director": "Yes",
                "is_local_resident_director": "Yes",
                "is_nominee_director": "Yes",
            },
            {
                "person_id": "P004",
                "source": "new",
                "full_name": "LIM COMPANY SECRETARY",
                "id_type": "NRIC",
                "id_number": "S7654321B",
                "nationality": "Singaporean",
                "residential_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
                "is_secretary": "Yes",
            },
        ],
    )
    fill_transposed_sheet(
        wb,
        "股东与股份",
        "股东",
        [
            {
                "shareholder_type": "person",
                "person_full_name": "ZHANG YI",
                "person_id_number": "P1000001A",
                "share_class": "Ordinary",
                "shares": 700,
                "issued_share_capital": 700,
                "paid_amount": 700,
                "currency": "SGD",
                "remarks": "70% shareholder.",
            },
            {
                "shareholder_type": "person",
                "person_full_name": "ZHANG ER",
                "person_id_number": "P2000002B",
                "share_class": "Ordinary",
                "shares": 300,
                "issued_share_capital": 300,
                "paid_amount": 300,
                "currency": "SGD",
                "remarks": "30% shareholder.",
            },
        ],
    )
    fill_single_value_sheet(
        wb,
        "输出设置",
        {
            "output_format": "docx_pdf_zip",
            "prepared_by": "Sample",
            "client_signatory_logic": "shareholder_1",
        },
    )
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
