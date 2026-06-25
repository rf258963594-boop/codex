from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
P1_TEMPLATE = ROOT / "templates" / "import" / "P1_registration_blank_v3_1.xlsx"
DEFAULT_OUT = ROOT / "outputs" / f"online_signature_test_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
RSIN_ADDRESS = "111 NORTH BRIDGE ROAD, #29-06A, PENINSULA PLAZA, SINGAPORE 179098"


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def set_vertical_values(ws, values: dict[str, object]) -> None:
    for row in range(1, ws.max_row + 1):
        key = clean(ws.cell(row, 3).value)
        if key in values:
            ws.cell(row, 4).value = values[key]


def set_transposed_records(ws, records: list[dict[str, object]], start_col: int = 5) -> None:
    key_rows: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        key = clean(ws.cell(row, 3).value)
        if key:
            key_rows[key] = row
    for offset, record in enumerate(records):
        col = start_col + offset
        for key, value in record.items():
            if key in key_rows:
                ws.cell(key_rows[key], col).value = value


def build_workbook(out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(P1_TEMPLATE)
    company = wb.worksheets[1]
    people = wb.worksheets[2]
    shareholders = wb.worksheets[3]
    output = wb.worksheets[4]

    set_vertical_values(
        company,
        {
            "company_name": "SIGNATURE COMMON PEOPLE TEST PTE. LTD.",
            "business_order_id": "ONLINE-SIGNATURE-COMMON-001",
            "source_type": "Manual online signature test",
            "source_file_id": "Common people signature smoke test",
            "registered_office_address": RSIN_ADDRESS,
            "incorporation_date": "18/06/2026",
            "first_fye": "",
            "fye": "",
            "business_activity_1": "MANAGEMENT CONSULTANCY SERVICES",
            "ssic_code_1": "70201",
            "business_activity_2": "",
            "ssic_code_2": "",
            "currency": "",
            "share_class_default": "",
            "document_date": "",
            "client_signatory_person_id": "P001",
            "task_type": "incorporation",
            "remarks": "Signature smoke test for secretary and two nominee directors.",
        },
    )

    set_transposed_records(
        people,
        [
            {
                "person_id": "P001",
                "source": "new",
                "full_name": "CLIENT DIRECTOR SHAREHOLDER",
                "id_type": "Passport",
                "id_number": "P12345678",
                "nationality": "CHINESE",
                "date_of_birth": "01/01/1988",
                "residential_address": "ROOM 1801, TEST CLIENT RESIDENCE, 88 LONG ADDRESS ROAD, SHANGHAI, CHINA",
                "email": "client.signature.test@example.com",
                "phone": "+86 13800138000",
                "is_director": "Yes",
            },
            {
                "person_id": "N001",
                "source": "common",
                "common_person_name": "XIONG SAI",
            },
            {
                "person_id": "N002",
                "source": "common",
                "common_person_name": "LE THI NGOC TRANG",
            },
            {
                "person_id": "S001",
                "source": "common",
                "common_person_name": "FENDI CHANDRA TING S ING EE",
            },
        ],
    )

    set_transposed_records(
        shareholders,
        [
            {
                "shareholder_type": "person",
                "person_full_name": "CLIENT DIRECTOR SHAREHOLDER",
                "person_id_number": "P12345678",
                "share_class": "",
                "shares": 100,
                "issued_share_capital": "",
                "paid_amount": "",
                "currency": "",
            },
        ],
    )

    set_vertical_values(output, {"prepared_by": "Online signature test"})

    workbook_path = out_dir / "P1_signature_test_two_nominees_secretary.xlsx"
    wb.save(workbook_path)
    readme = out_dir / "README_signature_test.txt"
    readme.write_text(
        "\n".join(
            [
                "Signature smoke-test workbook.",
                "Purpose: verify backend common-person signatures in generated PDFs.",
                "Common people used:",
                "- XIONG SAI",
                "- LE THI NGOC TRANG",
                "- FENDI CHANDRA TING S ING EE",
            ]
        ),
        encoding="utf-8",
    )
    zip_path = out_dir / "signature_test_input.zip"
    with ZipFile(zip_path, "w", ZIP_DEFLATED) as zf:
        zf.write(workbook_path, workbook_path.name)
        zf.write(readme, readme.name)
    return [workbook_path, readme, zip_path]


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a common-person signature smoke-test workbook.")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()
    for path in build_workbook(args.out_dir):
        print(path)


if __name__ == "__main__":
    main()
