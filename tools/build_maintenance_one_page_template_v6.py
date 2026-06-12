from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"

TITLE_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FILL = PatternFill("solid", fgColor="1F4D78")
SECTION_FILL = PatternFill("solid", fgColor="EAF2F8")
INPUT_FILL = PatternFill("solid", fgColor="FFF8D6")
NOTE_FILL = PatternFill("solid", fgColor="F8FAFC")
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
KEY_FONT = Font(color="475569", size=9)
NOTE_FONT = Font(color="334155")
WRAP = Alignment(wrap_text=True, vertical="top")

YES_NO = '"Auto,Yes,No"'
PERSON_ACTIONS = '"appoint_director,resign_director,appoint_secretary,resign_secretary"'
FIELD_LABELS = '"ID type,ID number,Residential address,Email,Phone,Nationality,Name,Other"'
CONSIDERATION_BASIS = '"internal_paid_up_basis,acra_paid_up_capital_basis,stamp_duty_higher_of_price_or_nav"'


KV_FIELDS = [
    ("公司名称", "Company name", "company_name", "", "必填", "所有文件都会用。"),
    ("UEN / 注册号", "UEN", "uen", "", "建议必填", "现有公司变更建议填写。"),
    ("当前注册地址", "Current registered office", "registered_office_address", "", "建议必填", "多数文件会用；地址变更时作为旧地址。"),
    ("当前总股数", "Current total shares", "total_issued_shares", "", "M04 可空", "用于增资配股后资本推导；不知道可留空。"),
    ("当前已发行股本", "Current issued share capital", "issued_share_capital", "", "M04 可空", "对应 ACRA issued share capital；不知道可留空。"),
    ("当前实缴资本", "Current paid-up share capital", "paid_up_capital", "", "M04 可空", "对应 ACRA paid-up share capital；不知道可留空。"),
    ("币种", "Currency", "currency", "SGD", "默认", "通常 SGD；如 USD 等请填写。"),
    ("文件日期", "Document date", "default_document_date", "", "建议必填", "DD/MM/YYYY；留空后续可由网页补当天。"),
    ("董事签字人姓名", "Director signer name", "director_signer_name", "", "建议必填", "M01 默认用这个人签字；不填则文件签字栏留空。"),
    ("股东/成员签字人姓名", "Member/shareholder signer names", "member_signer_names", "", "转入建议填", "M02 股东/成员签署处使用；多人可换行、逗号或分号分隔。"),
    ("客户方签字人姓名", "Client signer name", "client_signatory_name", "", "可空", "交接信签字人；留空默认用第一个股东/成员签字人。"),
    ("业务单编号", "Business order ID", "business_order_id", "", "可空", "后台预留。"),
    ("数据来源", "Source type", "source_type", "AI", "可空", "AI / BizFile / Excel / Manual。"),
    ("来源文件编号", "Source file ID", "source_file_id", "", "可空", "可填 BizFile 或旧表文件名。"),
    ("经办人", "Prepared by", "prepared_by", "", "可空", "内部记录。"),
    ("是否变更注册地址", "Change registered office", "change_registered_office_required", "", "默认 Auto", "留空/Auto 时，填写新注册地址即可生成；填 No 则不生成。"),
    ("新注册地址", "New registered office", "new_registered_office_address", "", "地址变更时填", ""),
    ("是否变更营业范围", "Change business activity", "change_business_activity_required", "", "默认 Auto", "留空/Auto 时，填写新 SSIC/业务范围即可生成；填 No 则不生成。支持主业务和副业务。"),
    ("新主 SSIC", "New primary SSIC", "new_primary_ssic", "", "可空", ""),
    ("新主营业范围", "New primary activity", "new_primary_activity", "", "营业范围变更时填", ""),
    ("新副 SSIC", "New secondary SSIC", "new_secondary_ssic", "", "可空", ""),
    ("新副营业范围", "New secondary activity", "new_secondary_activity", "", "可空", ""),
    ("是否变更 FYE", "Change FYE", "change_fye_required", "", "默认 Auto", "留空/Auto 时，填写新 FYE 即可生成；填 No 则不生成。"),
    ("旧 FYE", "Old FYE", "old_fye", "", "FYE 变更时填", "DD/MM/YYYY。"),
    ("新 FYE", "New FYE", "new_fye", "", "FYE 变更时填", "DD/MM/YYYY。"),
    ("下个账期开始", "Next accounts period start", "next_accounts_period_start", "", "可空", ""),
    ("下个账期结束", "Next accounts period end", "next_accounts_period_end", "", "可空", ""),
    ("是否转入秘书公司", "Transfer-in required", "transfer_in_required", "", "需要时填 Yes", "转入需要时填 Yes；M01 正常出，M02 一律兜底。"),
    ("转入模式", "Transfer-in mode", "transfer_in_mode", "", "内部预留", "通常留空；系统不再要求区分配合/不配合。"),
    ("旧秘书公司", "Old secretary firm", "old_secretary_company", "", "可空", ""),
    ("新秘书公司", "New secretary firm", "new_secretary_company", "RSIN GROUP PTE. LTD.", "可空", ""),
    ("是否出辞职信", "Generate resignation letter", "generate_resignation_letter", "No", "可空", "只有需要旧人员单独签辞职信时填 Yes；一般留 No。"),
    ("是否做年审", "Annual review required", "annual_review_required", "", "默认 Auto", "留空/Auto 时，填写 FYE 或 AGM 日期即可生成年审包；填 No 则不生成。"),
    ("年审 FYE", "Annual review FYE", "fye_date", "", "年审时填", "DD/MM/YYYY。"),
    ("AGM 日期", "AGM date", "agm_date", "", "可空", "DD/MM/YYYY。"),
    ("年审备注", "Annual review remarks", "annual_review_remarks", "", "可空", "复杂情况写这里；详细字段可填“快速年审”页。"),
    ("备注", "Notes", "notes", "", "可空", "内部备注，不进正式文件。"),
]

ANNUAL_FIELDS = [
    ("是否做年审", "Annual review required", "annual_review_required", "", "默认 Auto", "留空/Auto 时，填写 FYE 或 AGM 日期即可生成年审包；填 No 则不生成。"),
    ("财年结束日", "Financial year end", "fye_date", "", "必填", "DD/MM/YYYY。"),
    ("AGM 日期", "AGM date", "agm_date", "", "建议填", "留空后续可由系统按规则推定。"),
    ("AGM 时间", "AGM time", "agm_time", "10.00 a.m.", "默认", "通常不用改。"),
    ("AGM 地点", "AGM place", "agm_place", "", "默认注册地址", "留空默认当前注册地址。"),
    ("年审方式", "AGM route", "agm_route", "ordinary_agm", "默认", "ordinary_agm / exempt_private_company / dormant_company / manual。"),
    ("财报状态", "Accounts status", "accounts_status", "non_dormant", "常用", "non_dormant / dormant / unaudited / audited。"),
    ("公司活动状态", "Company activity status", "company_activity_status", "Active", "默认", "Active / Dormant；留空按 Active。"),
    ("是否 ACRA 休眠相关公司", "ACRA dormant relevant company", "acra_dormant_relevant_company", "Auto", "默认 Auto", "Auto / Yes / No；休眠公司会影响是否需要财报。"),
    ("资产是否不超过 50 万", "Total assets under S$500k", "total_assets_under_500k", "Auto", "默认 Auto", "Auto / Yes / No；用于判断 dormant AGM exemption 风险。"),
    ("是否需要财报", "Financial statements required", "financial_statements_required", "Auto", "默认 Auto", "Auto / Yes / No；休眠且符合条件时可 No。"),
    ("财报类型", "Financial statements type", "financial_statements_type", "Auto", "默认 Auto", "Auto / Unaudited / Audited / Dormant no FS / Management accounts。"),
    ("审计/豁免状态", "Audit exemption status", "audit_exemption_status", "Auto", "默认 Auto", "Auto / Small company exempt / Audited / Dormant relevant / Manual review。"),
    ("AGM 状态", "AGM status", "agm_status", "Auto", "默认 Auto", "Auto / Held AGM / Dispensed with AGM / Exempt from AGM / Written resolutions。"),
    ("IRAS 税务状态", "IRAS tax status", "iras_tax_status", "Auto", "可空", "Auto / Active / Dormant / Dormant waiver granted / Manual review；只作复核提醒。"),
    ("财报日期", "Financial statement date", "financial_statement_date", "", "默认 FYE", "留空默认财年结束日。"),
    ("董事签字人", "Director signer", "director_signer_name", "", "常用", "留空默认首页董事签字人。"),
    ("股东/成员签字人", "Shareholder signer", "shareholder_signer_name", "", "可空", "需要股东签署的年审文件使用。"),
    ("AR 授权签字人", "AR authorised signer", "ar_authorized_signer_name", "", "默认董事签字人", "Annual Return 授权。"),
    ("董事费", "Directors' fee", "directors_fee", "0", "默认", "通常 0。"),
    ("董事薪酬", "Directors' remuneration", "directors_remuneration", "0", "默认", "通常 0。"),
    ("Shorter Notice Consent", "Shorter notice consent", "shorter_notice_consent", "Auto", "默认", "Auto / Yes / No。"),
    ("Management Representation Letter", "MRL", "management_rep_letter", "Yes", "默认", "Yes / No。"),
    ("备注", "Remarks", "remarks", "", "可空", "特殊情况写这里。"),
]

PERSON_HEADERS = [
    ("是否办理", "generate"),
    ("动作", "action_type"),
    ("人员姓名", "target_name"),
    ("生效日期", "effective_date"),
    ("是否出辞职信", "resignation_letter"),
    ("备注", "remarks"),
]

PARTICULAR_HEADERS = [
    ("是否办理", "generate"),
    ("人员姓名", "target_name"),
    ("变更字段", "field_label"),
    ("原资料", "old_value"),
    ("新资料", "new_value"),
    ("生效日期", "effective_date"),
    ("备注", "remarks"),
]

TRANSFER_HEADERS = [
    ("是否办理", "generate"),
    ("转让人", "transferor_name"),
    ("转让人证件/注册号", "transferor_id_number"),
    ("转让人地址", "transferor_address"),
    ("受让人", "transferee_name"),
    ("受让人证件/注册号", "transferee_id_number"),
    ("受让人地址", "transferee_address"),
    ("转让股数", "shares_transferred"),
    ("股份类别", "share_class"),
    ("转让日期", "transfer_date"),
    ("对价口径", "consideration_basis"),
    ("对价金额", "consideration_amount"),
    ("币种", "currency"),
    ("旧证书号", "old_certificate_no"),
    ("新证书号", "new_certificate_no"),
    ("转让人剩余股数", "transferor_remaining_shares"),
    ("是否生成新证书", "generate_new_certificate"),
    ("印花税复核", "stamp_duty_review"),
    ("备注", "remarks"),
]

ALLOTMENT_HEADERS = [
    ("是否办理", "generate"),
    ("认购人", "allottee_name"),
    ("认购人证件/注册号", "allottee_id_number"),
    ("认购人地址", "allottee_address"),
    ("配发股数", "shares_allotted"),
    ("股份类别", "share_class"),
    ("本次已发行股本", "issued_share_capital"),
    ("每股实缴", "amount_paid_per_share"),
    ("本次实缴资本", "total_paid"),
    ("币种", "currency"),
    ("配股日期", "allotment_date"),
    ("授权日期", "authority_date"),
    ("是否 Form 24", "form24_required"),
    ("证书号", "certificate_no"),
    ("是否生成证书", "generate_certificate"),
    ("配股后总股数", "post_allotment_total_shares"),
    ("配股后已发行股本", "post_allotment_issued_share_capital"),
    ("配股后实缴资本", "post_allotment_paid_up_capital"),
    ("备注", "remarks"),
]


def sample_values() -> dict[str, Any]:
    return {
        "company_name": "SAMPLE ONE PAGE PTE. LTD.",
        "uen": "202612345A",
        "registered_office_address": "10 ANSON ROAD, #20-01 INTERNATIONAL PLAZA, SINGAPORE 079903",
        "total_issued_shares": 100000,
        "issued_share_capital": 100000,
        "paid_up_capital": 100000,
        "currency": "SGD",
        "default_document_date": "03/06/2026",
        "director_signer_names": "ZHANG YI\nZHANG ER",
        "member_signer_names": "ZHANG YI\nZHANG ER",
        "client_signatory_name": "ZHANG YI",
        "business_order_id": "MAINT-V6-001",
        "source_type": "AI",
        "source_file_id": "BizFile draft",
        "prepared_by": "AI draft",
        "change_registered_office_required": "No",
        "change_business_activity_required": "No",
        "change_fye_required": "No",
        "transfer_in_required": "No",
        "annual_review_required": "No",
    }


def one_page_kv_fields() -> list[tuple[str, str, str, str, str, str]]:
    fields: list[tuple[str, str, str, str, str, str]] = []
    replacement = (
        "M01 directors to sign",
        "M01 director signer names",
        "director_signer_names",
        "",
        "Recommended",
        "Enter all current directors for M01. One per line, comma, or semicolon.",
    )
    inserted = False
    for item in KV_FIELDS:
        if item[2] == "director_signer_name":
            fields.append(replacement)
            inserted = True
        else:
            fields.append(item)
    if not inserted:
        fields.insert(4, replacement)
    return fields


def person_sample_rows() -> list[dict[str, Any]]:
    return [
        {"generate": "Yes", "action_type": "resign_director", "target_name": "ZHANG ER", "effective_date": "03/06/2026", "resignation_letter": "No", "remarks": "示例：董事辞任"},
        {"generate": "Yes", "action_type": "appoint_director", "target_name": "ZHANG SAN", "effective_date": "03/06/2026", "resignation_letter": "No", "remarks": "示例：委任董事"},
    ]


def transfer_sample_rows() -> list[dict[str, Any]]:
    return [
        {
            "generate": "No",
            "transferor_name": "ZHANG YI",
            "transferor_id_number": "E12345678",
            "transferor_address": "10 ANSON ROAD, #20-01 INTERNATIONAL PLAZA, SINGAPORE 079903",
            "transferee_name": "ZHANG ER",
            "transferee_id_number": "G1234567A",
            "transferee_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098",
            "shares_transferred": 100,
            "share_class": "Ordinary",
            "transfer_date": "03/06/2026",
            "consideration_basis": "internal_paid_up_basis",
            "consideration_amount": "",
            "currency": "SGD",
            "old_certificate_no": "001",
            "new_certificate_no": "002",
            "transferor_remaining_shares": "",
            "generate_new_certificate": "Auto",
            "stamp_duty_review": "Auto",
            "remarks": "转股时改 Yes",
        },
    ]


def allotment_sample_rows() -> list[dict[str, Any]]:
    return [
        {
            "generate": "No",
            "allottee_name": "ZHANG NEW INVESTOR",
            "allottee_id_number": "G7654321B",
            "allottee_address": "6 RAFFLES QUAY, #16-01, SINGAPORE 048580",
            "shares_allotted": 1000,
            "share_class": "Ordinary",
            "issued_share_capital": 1000,
            "amount_paid_per_share": 1,
            "total_paid": 1000,
            "currency": "SGD",
            "allotment_date": "03/06/2026",
            "authority_date": "03/06/2026",
            "form24_required": "Auto",
            "certificate_no": "003",
            "generate_certificate": "Auto",
            "post_allotment_total_shares": "",
            "post_allotment_issued_share_capital": "",
            "post_allotment_paid_up_capital": "",
            "remarks": "增资配股时改 Yes；资本汇总可留空让系统提示复核",
        },
    ]


def set_widths(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 12
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)) + 2, 42))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def add_validation(ws, range_ref: str, values: str) -> None:
    dv = DataValidation(type="list", formula1=values, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(range_ref)


def header_key(value: Any) -> str:
    return str(value or "").splitlines()[-1].strip()


def find_column(ws, row: int, key: str) -> int | None:
    for col in range(1, ws.max_column + 1):
        if header_key(ws.cell(row, col).value) == key:
            return col
    return None


def find_kv_row(ws, key: str) -> int | None:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 3).value or "").strip() == key:
            return row
    return None


def add_kv_validation(ws, key: str, values: str) -> None:
    row = find_kv_row(ws, key)
    if row:
        add_validation(ws, f"D{row}:D{row}", values)


def add_table_validations(ws, header_row: int, first_data_row: int, last_data_row: int, validations: dict[str, str]) -> None:
    for key, values in validations.items():
        col = find_column(ws, header_row, key)
        if col:
            add_validation(ws, f"{get_column_letter(col)}{first_data_row}:{get_column_letter(col)}{last_data_row}", values)


def cell_has_fill(cell) -> bool:
    fill = cell.fill
    rgb = getattr(fill.fgColor, "rgb", None) if fill and fill.fgColor else None
    return bool(fill and fill.fill_type and rgb not in (None, "00000000", "FFFFFFFF"))


def apply_template_borders(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, "") or cell_has_fill(cell):
                    cell.border = THIN_BORDER


def add_section(ws, row: int, title: str, headers: list[tuple[str, str]], rows: list[dict[str, Any]], blank_rows: int = 6) -> int:
    ws.cell(row, 1).value = title
    ws.cell(row, 1).fill = SECTION_FILL
    ws.cell(row, 1).font = BOLD
    header_row = row + 1
    for col, (label, key) in enumerate(headers, start=1):
        cell = ws.cell(header_row, col)
        cell.value = f"{label}\n{key}"
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = WRAP
    data_row = header_row + 1
    all_rows = rows + [{} for _ in range(blank_rows)]
    for r_offset, item in enumerate(all_rows):
        for col, (_, key) in enumerate(headers, start=1):
            cell = ws.cell(data_row + r_offset, col)
            cell.value = item.get(key, "")
            cell.fill = INPUT_FILL
            cell.alignment = WRAP
    return data_row + len(all_rows) + 1


def build_workbook(sample: bool) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "P2快速业务单"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "P2 一页式快速业务单 - 少填字段先生成文件"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    ws["A1"].alignment = WRAP
    ws.append([])
    ws.append(["项目", "English", "field_key", "填写内容 / Value", "必填", "说明"])
    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = WRAP
    sample_map = sample_values() if sample else {}
    row = 4
    for cn, en, key, default, required, note in one_page_kv_fields():
        values = [cn, en, key, sample_map.get(key, default), required, note]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col)
            cell.value = value
            cell.alignment = WRAP
            if col == 3:
                cell.fill = PatternFill("solid", fgColor="EEF2F7")
                cell.font = KEY_FONT
            elif col == 4:
                cell.fill = INPUT_FILL
            elif col == 6:
                cell.fill = NOTE_FILL
                cell.font = NOTE_FONT
        row += 1
    add_kv_validation(ws, "change_registered_office_required", YES_NO)
    add_kv_validation(ws, "change_business_activity_required", YES_NO)
    add_kv_validation(ws, "change_fye_required", YES_NO)
    add_kv_validation(ws, "transfer_in_required", YES_NO)
    add_kv_validation(ws, "generate_resignation_letter", YES_NO)
    add_kv_validation(ws, "annual_review_required", YES_NO)

    row += 2
    person_header_row = row + 1
    row = add_section(ws, row, "人员任免", PERSON_HEADERS, person_sample_rows() if sample else [], blank_rows=8)
    add_table_validations(ws, person_header_row, person_header_row + 1, row - 2, {"generate": YES_NO, "action_type": PERSON_ACTIONS, "resignation_letter": YES_NO})

    particular_header_row = row + 1
    particular_rows = [
        {"generate": "No", "target_name": "ZHANG YI", "field_label": "ID number", "old_value": "E12345678", "new_value": "G1234567A", "effective_date": "03/06/2026", "remarks": "资料更新时改 Yes"}
    ] if sample else []
    row = add_section(ws, row, "个人资料变更", PARTICULAR_HEADERS, particular_rows, blank_rows=8)
    add_table_validations(ws, particular_header_row, particular_header_row + 1, row - 2, {"generate": YES_NO, "field_label": FIELD_LABELS})

    transfer_header_row = row + 1
    row = add_section(ws, row, "股份转让", TRANSFER_HEADERS, transfer_sample_rows() if sample else [], blank_rows=8)
    add_table_validations(
        ws,
        transfer_header_row,
        transfer_header_row + 1,
        row - 2,
        {
            "generate": YES_NO,
            "consideration_basis": CONSIDERATION_BASIS,
            "generate_new_certificate": '"Auto,Yes,No"',
            "stamp_duty_review": '"Auto,Yes,No"',
        },
    )

    allotment_header_row = row + 1
    row = add_section(ws, row, "增资配股", ALLOTMENT_HEADERS, allotment_sample_rows() if sample else [], blank_rows=8)
    add_table_validations(
        ws,
        allotment_header_row,
        allotment_header_row + 1,
        row - 2,
        {
            "generate": YES_NO,
            "form24_required": '"Auto,Yes,No"',
            "generate_certificate": '"Auto,Yes,No"',
        },
    )

    ws.freeze_panes = "A4"
    set_widths(ws)
    build_annual_sheet(wb, sample)
    return wb


def build_annual_sheet(wb: Workbook, sample: bool) -> None:
    ws = wb.create_sheet("快速年审")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "快速年审 Annual Review - 常用默认值已内置"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    ws["A1"].alignment = WRAP
    ws.append([])
    ws.append(["项目", "English", "field_key", "填写内容 / Value", "必填", "说明"])
    annual_sample = {
        "annual_review_required": "Yes",
        "fye_date": "31/12/2025",
        "agm_date": "30/06/2026",
        "accounts_status": "unaudited",
        "company_activity_status": "Active",
        "financial_statements_type": "Unaudited",
        "financial_statements_required": "Yes",
        "audit_exemption_status": "Small company exempt",
        "agm_status": "Held AGM",
        "director_signer_name": "ZHANG YI",
        "ar_authorized_signer_name": "ZHANG YI",
    } if sample else {}
    row = 4
    for cn, en, key, default, required, note in ANNUAL_FIELDS:
        for col, value in enumerate([cn, en, key, annual_sample.get(key, default), required, note], start=1):
            cell = ws.cell(row, col)
            cell.value = value
            cell.alignment = WRAP
            if col == 3:
                cell.fill = PatternFill("solid", fgColor="EEF2F7")
                cell.font = KEY_FONT
            elif col == 4:
                cell.fill = INPUT_FILL
            elif col == 6:
                cell.fill = NOTE_FILL
                cell.font = NOTE_FONT
        row += 1
    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = WRAP
    add_kv_validation(ws, "annual_review_required", YES_NO)
    add_kv_validation(ws, "agm_route", '"ordinary_agm,exempt_private_company,dormant_company,written_resolutions,manual"')
    add_kv_validation(ws, "accounts_status", '"non_dormant,dormant,unaudited,audited,manual"')
    add_kv_validation(ws, "company_activity_status", '"Active,Dormant"')
    add_kv_validation(ws, "acra_dormant_relevant_company", YES_NO)
    add_kv_validation(ws, "total_assets_under_500k", YES_NO)
    add_kv_validation(ws, "financial_statements_required", YES_NO)
    add_kv_validation(ws, "financial_statements_type", '"Auto,Unaudited,Audited,Dormant no FS,Management accounts,Not applicable"')
    add_kv_validation(ws, "audit_exemption_status", '"Auto,Small company exempt,Audited,Dormant relevant,Manual review"')
    add_kv_validation(ws, "agm_status", '"Auto,Held AGM,Dispensed with AGM,Exempt from AGM,Written resolutions,Manual review"')
    add_kv_validation(ws, "iras_tax_status", '"Auto,Active,Dormant,Dormant waiver granted,Manual review"')
    add_kv_validation(ws, "shorter_notice_consent", YES_NO)
    add_kv_validation(ws, "management_rep_letter", YES_NO)
    ws.freeze_panes = "A4"
    set_widths(ws)


def save_workbook(wb: Workbook, name: str) -> Path:
    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / name
    apply_template_borders(wb)
    wb.save(path)
    return path


def main() -> None:
    paths = [
        save_workbook(build_workbook(False), "AI适配_公司维护变更年审资料模板_v7_一页式快速业务单含快速年审_空白.xlsx"),
        save_workbook(build_workbook(True), "AI适配_公司维护变更年审资料模板_v7_一页式快速业务单含快速年审_示例.xlsx"),
    ]
    for path in paths:
        print(path)


if __name__ == "__main__":
    main()
