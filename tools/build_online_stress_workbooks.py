from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
P1_TEMPLATE = ROOT / "templates" / "import" / "P1_registration_blank_v3_1.xlsx"
P2_TEMPLATE = ROOT / "templates" / "import" / "P2_maintenance_annual_blank_v7.xlsx"
DEFAULT_OUT = ROOT / "outputs" / f"online_stress_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

COMMON_NOMINEE = "LE THI NGOC TRANG"
COMMON_SECRETARY = "FENDI CHANDRA TING S ING EE"
COMMON_OLD_NOMINEE = "XIONG SAI"
RSIN_ADDRESS = "111 NORTH BRIDGE ROAD, #29-06A, PENINSULA PLAZA, SINGAPORE 179098"


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def header_key(value) -> str:
    raw = clean(value)
    if not raw:
        return ""
    parts = raw.replace("\r", "\n").split("\n")
    for part in reversed(parts):
        part = part.strip().strip("()[] ")
        if part.replace("_", "").isalnum() and "_" in part or part.isidentifier():
            return part
    return raw


def set_vertical_values(ws, values: dict[str, object]) -> None:
    for row in range(1, ws.max_row + 1):
        key = clean(ws.cell(row, 3).value)
        if key in values:
            ws.cell(row, 4).value = values[key]


def set_transposed_records(ws, records: list[dict[str, object]], start_col: int = 5) -> None:
    key_rows: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        key = clean(ws.cell(row, 3).value)
        if key:
            key_rows[key] = row
    for offset, record in enumerate(records):
        col = start_col + offset
        for key, value in record.items():
            if key in key_rows:
                ws.cell(key_rows[key], col).value = value


def find_table(ws, required_keys: set[str], used_header_rows: set[int] | None = None) -> tuple[int, list[str]]:
    used_header_rows = used_header_rows or set()
    for row in range(1, ws.max_row + 1):
        if row in used_header_rows:
            continue
        keys = [header_key(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if required_keys.issubset({key for key in keys if key}):
            return row, keys
    raise ValueError(f"Cannot find table headers: {sorted(required_keys)}")


def set_table(ws, required_keys: set[str], records: list[dict[str, object]], used_header_rows: set[int] | None = None) -> int:
    header_row, keys = find_table(ws, required_keys, used_header_rows)
    for offset, record in enumerate(records, start=1):
        row = header_row + offset
        for col, key in enumerate(keys, start=1):
            if key in record:
                ws.cell(row, col).value = record[key]
    return header_row


def p2_base_workbook():
    return load_workbook(P2_TEMPLATE)


def save_workbook(wb, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def build_p1(out_dir: Path) -> Path:
    wb = load_workbook(P1_TEMPLATE)
    company = wb.worksheets[1]
    people = wb.worksheets[2]
    shareholders = wb.worksheets[3]
    output = wb.worksheets[4]

    set_vertical_values(
        company,
        {
            "company_name": "APEX ONLINE STRESS HOLDINGS PTE. LTD.",
            "business_order_id": "ONLINE-P1-STRESS-001",
            "source_type": "AI + Manual online stress",
            "source_file_id": "Online registration simulation",
            "registered_office_address": RSIN_ADDRESS,
            "incorporation_date": "17/06/2026",
            "first_fye": "",
            "fye": "",
            "business_activity_1": "MANAGEMENT CONSULTANCY SERVICES",
            "ssic_code_1": "70201",
            "business_activity_2": "DEVELOPMENT OF SOFTWARE AND APPLICATIONS",
            "ssic_code_2": "62011",
            "currency": "SGD",
            "share_class_default": "Ordinary",
            "contact_person_id": "P001",
            "client_signatory_person_id": "",
            "task_type": "incorporation",
            "remarks": "Online stress test. FYE and document date intentionally left for system default.",
        },
    )
    set_transposed_records(
        people,
        [
            {
                "person_id": "P001",
                "source": "new",
                "full_name": "ALEX CHEN YI",
                "id_type": "Passport",
                "id_number": "E55881234",
                "nationality": "CHINESE",
                "date_of_birth": "18/03/1988",
                "residential_address": "ROOM 1802, BLOCK 7, RIVER VALLEY INTERNATIONAL RESIDENCE, 288 LONG ADDRESS ROAD, SHANGHAI, CHINA",
                "email": "alex.chen.online@example.com",
                "phone": "+86 13800138001",
                "is_director": "Yes",
            },
            {
                "person_id": "P002",
                "source": "new",
                "full_name": "BEATRICE LIN MEI",
                "id_type": "Passport",
                "id_number": "E66004567",
                "nationality": "CHINESE",
                "date_of_birth": "09/11/1990",
                "residential_address": "UNIT 1206, TOWER B, OCEAN FINANCE CENTRE, 88 HARBOUR AVENUE, SHENZHEN, GUANGDONG, CHINA",
                "email": "beatrice.lin.online@example.com",
                "phone": "+86 13900139002",
                "is_director": "Yes",
            },
            {
                "person_id": "P003",
                "source": "new",
                "full_name": "DAVID WONG JUN",
                "id_type": "Passport",
                "id_number": "K99881234",
                "nationality": "MALAYSIAN",
                "date_of_birth": "21/07/1985",
                "residential_address": "NO. 16, JALAN DAMAI, TAMAN CENTRAL, 50450 KUALA LUMPUR, MALAYSIA",
                "email": "david.wong.rep@example.com",
                "phone": "+60 121234567",
                "is_authorized_rep": "Yes",
            },
            {
                "person_id": "P004",
                "source": "common",
                "common_person_name": COMMON_NOMINEE,
            },
            {
                "person_id": "P005",
                "source": "common",
                "common_person_name": COMMON_SECRETARY,
            },
        ],
    )
    set_transposed_records(
        shareholders,
        [
            {
                "shareholder_type": "person",
                "person_full_name": "ALEX CHEN YI",
                "person_id_number": "E55881234",
                "share_class": "Ordinary",
                "shares": 60000,
                "issued_share_capital": 60000,
                "paid_amount": 60000,
                "currency": "SGD",
            },
            {
                "shareholder_type": "person",
                "person_full_name": "BEATRICE LIN MEI",
                "person_id_number": "E66004567",
                "share_class": "Ordinary",
                "shares": 20000,
                "issued_share_capital": 20000,
                "paid_amount": 2000,
                "currency": "SGD",
            },
            {
                "shareholder_type": "corporate",
                "corporate_name": "ORION GLOBAL INVESTMENTS LIMITED",
                "corporate_registration_country": "Hong Kong",
                "corporate_registration_number": "CR-88991234",
                "corporate_registered_address": "ROOM 2501, CENTRAL PLAZA, 18 HARBOUR ROAD, WAN CHAI, HONG KONG",
                "authorized_rep_full_name": "DAVID WONG JUN",
                "share_class": "Ordinary",
                "shares": 20000,
                "issued_share_capital": 20000,
                "paid_amount": 20000,
                "currency": "SGD",
            },
        ],
    )
    set_vertical_values(output, {"prepared_by": "Online stress tester"})
    return save_workbook(wb, out_dir / "P1_online_common_people_full.xlsx")


def fill_common_company(main, *, name: str, uen: str, address: str, order_id: str) -> None:
    set_vertical_values(
        main,
        {
            "company_name": name,
            "uen": uen,
            "registered_office_address": address,
            "default_document_date": "",
            "business_order_id": order_id,
            "source_type": "AI + Manual online stress",
            "source_file_id": "Online stress suite",
            "prepared_by": "Online stress tester",
            "director_signer_names": "ALEX CHEN YI\nBEATRICE LIN MEI",
            "member_signer_names": "ALEX CHEN YI\nBEATRICE LIN MEI\nORION GLOBAL INVESTMENTS LIMITED",
            "client_signatory_name": "ALEX CHEN YI",
            "new_secretary_company": "RSIN GROUP PTE. LTD.",
            "notes": "Default document date intentionally blank to test online default date.",
        },
    )


def build_m01(out_dir: Path) -> Path:
    wb = p2_base_workbook()
    main = wb.worksheets[0]
    fill_common_company(
        main,
        name="ORCHARD TECH SERVICES PTE. LTD.",
        uen="202612345N",
        address="8 TEMASEK BOULEVARD, #35-01 SUNTEC TOWER THREE, SINGAPORE 038988",
        order_id="ONLINE-M01-STRESS-001",
    )
    set_vertical_values(
        main,
        {
            "change_registered_office_required": "",
            "new_registered_office_address": RSIN_ADDRESS,
            "change_business_activity_required": "",
            "new_primary_ssic": "62011",
            "new_primary_activity": "DEVELOPMENT OF SOFTWARE AND APPLICATIONS",
            "new_secondary_ssic": "70201",
            "new_secondary_activity": "MANAGEMENT CONSULTANCY SERVICES",
            "change_fye_required": "",
            "old_fye": "31/12/2025",
            "new_fye": "30/06/2026",
            "next_accounts_period_start": "01/01/2026",
            "next_accounts_period_end": "30/06/2026",
            "transfer_in_required": "No",
            "annual_review_required": "No",
            "generate_resignation_letter": "Yes",
        },
    )
    set_table(
        main,
        {"generate", "action_type", "target_name"},
        [
            {"generate": "Yes", "action_type": "resign_director", "target_name": "OLD LOCAL DIRECTOR KOH", "effective_date": "17/06/2026", "resignation_letter": "Yes"},
            {"generate": "Yes", "action_type": "appoint_director", "target_name": COMMON_NOMINEE, "effective_date": "17/06/2026", "resignation_letter": "No"},
            {"generate": "Yes", "action_type": "resign_secretary", "target_name": "OLD SECRETARY LIM", "effective_date": "17/06/2026", "resignation_letter": "Yes"},
            {"generate": "Yes", "action_type": "appoint_secretary", "target_name": COMMON_SECRETARY, "effective_date": "17/06/2026", "resignation_letter": "No"},
        ],
    )
    set_table(
        main,
        {"generate", "field_label", "old_value", "new_value"},
        [
            {"generate": "Yes", "target_name": "ALEX CHEN YI", "field_label": "ID number", "old_value": "E55881234", "new_value": "S8123456A", "effective_date": "17/06/2026"},
            {"generate": "Yes", "target_name": "ALEX CHEN YI", "field_label": "Residential address", "old_value": "OLD OVERSEAS ADDRESS", "new_value": "12 MARINA VIEW, #20-01, ASIA SQUARE TOWER 2, SINGAPORE 018961", "effective_date": "17/06/2026"},
            {"generate": "Yes", "target_name": "BEATRICE LIN MEI", "field_label": "Email", "old_value": "old.email@example.com", "new_value": "beatrice.new@example.com", "effective_date": "17/06/2026"},
            {"generate": "Yes", "target_name": "BEATRICE LIN MEI", "field_label": "Phone", "old_value": "+86 13900139002", "new_value": "+65 81234567", "effective_date": "17/06/2026"},
        ],
    )
    return save_workbook(wb, out_dir / "P2_M01_online_ordinary_dr_common_people.xlsx")


def build_m02(out_dir: Path) -> Path:
    wb = p2_base_workbook()
    main = wb.worksheets[0]
    fill_common_company(
        main,
        name="EASTSEED ONLINE TRANSFER PTE. LTD.",
        uen="202512345K",
        address="10 ANSON ROAD, #20-05, INTERNATIONAL PLAZA, SINGAPORE 079903",
        order_id="ONLINE-M02-STRESS-001",
    )
    set_vertical_values(
        main,
        {
            "change_registered_office_required": "",
            "new_registered_office_address": RSIN_ADDRESS,
            "transfer_in_required": "Yes",
            "transfer_in_mode": "",
            "old_secretary_company": "",
            "generate_resignation_letter": "No",
            "change_business_activity_required": "No",
            "change_fye_required": "No",
            "annual_review_required": "No",
        },
    )
    set_table(
        main,
        {"generate", "action_type", "target_name"},
        [
            {"generate": "Yes", "action_type": "resign_director", "target_name": COMMON_OLD_NOMINEE, "effective_date": "17/06/2026", "resignation_letter": "No"},
            {"generate": "Yes", "action_type": "appoint_director", "target_name": COMMON_NOMINEE, "effective_date": "17/06/2026", "resignation_letter": "No"},
            {"generate": "Yes", "action_type": "appoint_secretary", "target_name": COMMON_SECRETARY, "effective_date": "17/06/2026", "resignation_letter": "No"},
        ],
    )
    return save_workbook(wb, out_dir / "P2_M02_online_transfer_in_common_people.xlsx")


def build_m03(out_dir: Path) -> Path:
    wb = p2_base_workbook()
    main = wb.worksheets[0]
    fill_common_company(
        main,
        name="RBIZ SHARE TRANSFER ONLINE PTE. LTD.",
        uen="202633333Z",
        address=RSIN_ADDRESS,
        order_id="ONLINE-M03-STRESS-001",
    )
    set_vertical_values(
        main,
        {
            "total_issued_shares": 100000,
            "issued_share_capital": 100000,
            "paid_up_capital": 100000,
            "transfer_in_required": "No",
            "change_registered_office_required": "No",
            "change_business_activity_required": "No",
            "change_fye_required": "No",
            "annual_review_required": "No",
        },
    )
    set_table(
        main,
        {"generate", "transferor_name", "transferee_name", "shares_transferred"},
        [
            {
                "generate": "",
                "transferor_name": "ALEX CHEN YI",
                "transferor_id_number": "S8123456A",
                "transferor_address": "12 MARINA VIEW, #20-01, ASIA SQUARE TOWER 2, SINGAPORE 018961",
                "transferee_name": "BEATRICE LIN MEI",
                "transferee_id_number": "E66004567",
                "transferee_address": "UNIT 1206, TOWER B, OCEAN FINANCE CENTRE, 88 HARBOUR AVENUE, SHENZHEN, CHINA",
                "shares_transferred": 15000,
                "share_class": "Ordinary",
                "consideration": "S$15,000",
                "transfer_date": "17/06/2026",
            },
            {
                "generate": "Yes",
                "transferor_name": "ORION GLOBAL INVESTMENTS LIMITED",
                "transferor_id_number": "CR-88991234",
                "transferor_address": "ROOM 2501, CENTRAL PLAZA, 18 HARBOUR ROAD, WAN CHAI, HONG KONG",
                "transferee_name": "ALEX CHEN YI",
                "transferee_id_number": "S8123456A",
                "transferee_address": "12 MARINA VIEW, #20-01, ASIA SQUARE TOWER 2, SINGAPORE 018961",
                "shares_transferred": 5000,
                "share_class": "Ordinary",
                "consideration": "S$5,000",
                "transfer_date": "17/06/2026",
            },
        ],
    )
    return save_workbook(wb, out_dir / "P2_M03_online_share_transfer.xlsx")


def build_m04(out_dir: Path) -> Path:
    wb = p2_base_workbook()
    main = wb.worksheets[0]
    fill_common_company(
        main,
        name="RBIZ SHARE ALLOTMENT ONLINE PTE. LTD.",
        uen="202644444D",
        address=RSIN_ADDRESS,
        order_id="ONLINE-M04-STRESS-001",
    )
    set_vertical_values(
        main,
        {
            "total_issued_shares": 80000,
            "issued_share_capital": 80000,
            "paid_up_capital": 8000,
            "currency": "SGD",
            "transfer_in_required": "No",
            "change_registered_office_required": "No",
            "change_business_activity_required": "No",
            "change_fye_required": "No",
            "annual_review_required": "No",
        },
    )
    set_table(
        main,
        {"generate", "allottee_name", "shares_allotted", "issued_share_capital"},
        [
            {
                "generate": "",
                "allottee_name": "BEATRICE LIN MEI",
                "allottee_id_number": "E66004567",
                "allottee_address": "UNIT 1206, TOWER B, OCEAN FINANCE CENTRE, 88 HARBOUR AVENUE, SHENZHEN, CHINA",
                "shares_allotted": 20000,
                "share_class": "Ordinary",
                "issued_share_capital": 100000,
                "amount_paid_per_share": 1,
                "total_paid": 20000,
                "currency": "SGD",
                "allotment_date": "17/06/2026",
                "authority_date": "17/06/2026",
            },
            {
                "generate": "Yes",
                "allottee_name": "ORION GLOBAL INVESTMENTS LIMITED",
                "allottee_id_number": "CR-88991234",
                "allottee_address": "ROOM 2501, CENTRAL PLAZA, 18 HARBOUR ROAD, WAN CHAI, HONG KONG",
                "shares_allotted": 50000,
                "share_class": "Ordinary",
                "issued_share_capital": 150000,
                "amount_paid_per_share": 0.1,
                "total_paid": 5000,
                "currency": "SGD",
                "allotment_date": "17/06/2026",
                "authority_date": "17/06/2026",
            },
        ],
    )
    return save_workbook(wb, out_dir / "P2_M04_online_share_allotment.xlsx")


def build_m05(out_dir: Path, dormant: bool = False) -> Path:
    wb = p2_base_workbook()
    main = wb.worksheets[0]
    annual = wb.worksheets[1]
    fill_common_company(
        main,
        name="RBIZ ANNUAL REVIEW ONLINE PTE. LTD." if not dormant else "RBIZ DORMANT ANNUAL ONLINE PTE. LTD.",
        uen="202655555H" if not dormant else "202666666R",
        address=RSIN_ADDRESS,
        order_id="ONLINE-M05-STRESS-001" if not dormant else "ONLINE-M05-DORMANT-STRESS-001",
    )
    set_vertical_values(
        main,
        {
            "total_issued_shares": 100000,
            "issued_share_capital": 100000,
            "paid_up_capital": 100000,
            "annual_review_required": "",
            "fye_date": "31/12/2025",
            "agm_date": "" if dormant else "17/06/2026",
            "change_registered_office_required": "No",
            "change_business_activity_required": "No",
            "change_fye_required": "No",
            "transfer_in_required": "No",
        },
    )
    set_vertical_values(
        annual,
        {
            "annual_review_required": "",
            "fye_date": "31/12/2025",
            "agm_date": "" if dormant else "17/06/2026",
            "agm_time": "",
            "agm_place": "",
            "agm_route": "exempt_private_company" if dormant else "ordinary_agm",
            "accounts_status": "dormant" if dormant else "active",
            "company_activity_status": "",
            "acra_dormant_relevant_company": "Yes" if dormant else "",
            "total_assets_under_500k": "Yes" if dormant else "",
            "financial_statements_required": "No" if dormant else "",
            "financial_statements_type": "",
            "audit_exemption_status": "",
            "agm_status": "Exempt from AGM" if dormant else "Held AGM",
            "director_signer_name": "ALEX CHEN YI",
            "shareholder_signer_name": "ALEX CHEN YI",
            "ar_authorized_signer_name": "",
            "directors_fee": 0,
            "directors_remuneration": 0,
            "shorter_notice_consent": "Auto",
            "management_rep_letter": "Yes",
            "remarks": "Online annual review stress test.",
        },
    )
    suffix = "dormant" if dormant else "ordinary"
    return save_workbook(wb, out_dir / f"P2_M05_online_annual_review_{suffix}.xlsx")


def blank_p1_defaults(company, output) -> None:
    set_vertical_values(
        company,
        {
            "business_order_id": "",
            "source_type": "",
            "source_file_id": "",
            "incorporation_date": "",
            "first_fye": "",
            "fye": "",
            "currency": "",
            "share_class_default": "",
            "issued_share_capital": "",
            "total_issued_shares": "",
            "paid_up_capital": "",
            "office_hours": "",
            "document_date": "",
            "uen": "",
            "register_location": "",
            "contact_person_id": "",
            "agent_person_id": "",
            "client_signatory_person_id": "",
            "authorized_representative_person_id": "",
            "remarks": "",
        },
    )
    set_vertical_values(
        output,
        {
            "prepared_by": "",
            "signing_mode": "",
            "client_signatory_logic": "",
            "form24_mode": "",
            "share_certificate_signing_director": "",
        },
    )


def build_p1_quick_start(out_dir: Path) -> Path:
    wb = load_workbook(P1_TEMPLATE)
    company = wb.worksheets[1]
    people = wb.worksheets[2]
    output = wb.worksheets[4]
    blank_p1_defaults(company, output)
    set_vertical_values(
        company,
        {
            "company_name": "",
            "registered_office_address": RSIN_ADDRESS,
            "business_activity_1": "",
            "ssic_code_1": "",
            "business_activity_2": "",
            "ssic_code_2": "",
            "task_type": "incorporation",
        },
    )
    set_transposed_records(
        people,
        [
            {},
            {},
            {
                "person_id": "COMMON_NOMINEE",
                "source": "common",
                "common_person_name": COMMON_NOMINEE,
            },
            {
                "person_id": "COMMON_SECRETARY",
                "source": "common",
                "common_person_name": COMMON_SECRETARY,
            },
        ],
    )
    return save_workbook(wb, out_dir / "P1_quick_start_common_info_blank_clients.xlsx")


def blank_p2_defaults(main, annual) -> None:
    set_vertical_values(
        main,
        {
            "company_name": "",
            "uen": "",
            "registered_office_address": "",
            "total_issued_shares": "",
            "issued_share_capital": "",
            "paid_up_capital": "",
            "currency": "",
            "default_document_date": "",
            "director_signer_names": "",
            "member_signer_names": "",
            "client_signatory_name": "",
            "business_order_id": "",
            "source_type": "",
            "source_file_id": "",
            "prepared_by": "",
            "change_registered_office_required": "",
            "new_registered_office_address": "",
            "change_business_activity_required": "",
            "new_primary_ssic": "",
            "new_primary_activity": "",
            "new_secondary_ssic": "",
            "new_secondary_activity": "",
            "change_fye_required": "",
            "old_fye": "",
            "new_fye": "",
            "next_accounts_period_start": "",
            "next_accounts_period_end": "",
            "transfer_in_required": "",
            "transfer_in_mode": "",
            "old_secretary_company": "",
            "new_secretary_company": "RSIN GROUP PTE. LTD.",
            "generate_resignation_letter": "",
            "annual_review_required": "",
            "fye_date": "",
            "agm_date": "",
            "annual_review_remarks": "",
            "notes": "",
        },
    )
    set_vertical_values(
        annual,
        {
            "annual_review_required": "",
            "fye_date": "",
            "agm_date": "",
            "agm_time": "",
            "agm_place": "",
            "agm_route": "",
            "accounts_status": "",
            "company_activity_status": "",
            "acra_dormant_relevant_company": "",
            "total_assets_under_500k": "",
            "financial_statements_required": "",
            "financial_statements_type": "",
            "audit_exemption_status": "",
            "agm_status": "",
            "iras_tax_status": "",
            "financial_statement_date": "",
            "director_signer_name": "",
            "shareholder_signer_name": "",
            "ar_authorized_signer_name": "",
            "directors_fee": "",
            "directors_remuneration": "",
            "shorter_notice_consent": "",
            "management_rep_letter": "",
            "remarks": "",
        },
    )


def build_p2_general_quick_start(out_dir: Path) -> Path:
    wb = p2_base_workbook()
    main = wb.worksheets[0]
    annual = wb.worksheets[1]
    blank_p2_defaults(main, annual)
    return save_workbook(wb, out_dir / "P2_quick_start_general_blank_clients.xlsx")


def build_p2_transfer_in_quick_start(out_dir: Path) -> Path:
    wb = p2_base_workbook()
    main = wb.worksheets[0]
    annual = wb.worksheets[1]
    blank_p2_defaults(main, annual)
    set_vertical_values(
        main,
        {
            "new_registered_office_address": RSIN_ADDRESS,
            "transfer_in_required": "Yes",
            "new_secretary_company": "RSIN GROUP PTE. LTD.",
        },
    )
    set_table(
        main,
        {"generate", "action_type", "target_name"},
        [
            {"generate": "", "action_type": "resign_director", "target_name": "", "effective_date": "", "resignation_letter": ""},
            {"generate": "Yes", "action_type": "appoint_director", "target_name": COMMON_NOMINEE, "effective_date": "", "resignation_letter": ""},
            {"generate": "Yes", "action_type": "appoint_secretary", "target_name": COMMON_SECRETARY, "effective_date": "", "resignation_letter": ""},
        ],
    )
    return save_workbook(wb, out_dir / "P2_quick_start_transfer_in_common_info.xlsx")


def build_p2_annual_quick_start(out_dir: Path) -> Path:
    wb = p2_base_workbook()
    main = wb.worksheets[0]
    annual = wb.worksheets[1]
    blank_p2_defaults(main, annual)
    set_vertical_values(main, {"registered_office_address": ""})
    return save_workbook(wb, out_dir / "P2_quick_start_annual_review_blank_clients.xlsx")


def build_quick_start_all(out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    files = [
        build_p1_quick_start(out_dir),
        build_p2_general_quick_start(out_dir),
        build_p2_transfer_in_quick_start(out_dir),
        build_p2_annual_quick_start(out_dir),
    ]
    readme = out_dir / "README_quick_start_templates.txt"
    readme.write_text(
        "\n".join(
            [
                "Quick-start workbooks for real daily use.",
                "",
                "Principles:",
                "- Client company, client people, shareholding, UEN, dates and financial data are intentionally blank.",
                "- Auto/default fields are intentionally blank so the website applies rules at generation time.",
                "- Internal common information is prefilled only where it does not create misleading client data.",
                "",
                "Prefilled common information:",
                f"- Registered office address for P1 registration: {RSIN_ADDRESS}",
                f"- Common nominee director: {COMMON_NOMINEE}",
                f"- Common secretary: {COMMON_SECRETARY}",
                "- New secretary company on P2: RSIN GROUP PTE. LTD.",
                "",
                "For transfer-in:",
                "- The transfer-in quick-start sheet pre-fills RSIN registered office and appoints the common nominee director / secretary.",
                "- Old director / old secretary company are left blank because they normally come from BizFile or client confirmation.",
            ]
        ),
        encoding="utf-8",
    )
    files.append(readme)
    zip_path = out_dir / "online_quick_start_templates_common_info_blank_clients.zip"
    with ZipFile(zip_path, "w", ZIP_DEFLATED) as zf:
        for path in files:
            zf.write(path, path.name)
    return files + [zip_path]


def build_all(out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    files = [
        build_p1(out_dir),
        build_m01(out_dir),
        build_m02(out_dir),
        build_m03(out_dir),
        build_m04(out_dir),
        build_m05(out_dir, dormant=False),
        build_m05(out_dir, dormant=True),
    ]
    readme = out_dir / "README_online_stress_inputs.txt"
    readme.write_text(
        "\n".join(
            [
                "Online stress and quick-start workbooks.",
                "Use these through the website upload page, not by editing server data directly.",
                "Common people names expected in admin backend:",
                f"- Nominee director: {COMMON_NOMINEE}",
                f"- Secretary: {COMMON_SECRETARY}",
                f"- Existing nominee director: {COMMON_OLD_NOMINEE}",
                "Some Auto/Yes/No and document date fields are intentionally blank to test defaults.",
            ]
        ),
        encoding="utf-8",
    )
    files.append(readme)
    zip_path = out_dir / "online_stress_and_quick_start_workbooks.zip"
    with ZipFile(zip_path, "w", ZIP_DEFLATED) as zf:
        for path in files:
            zf.write(path, path.name)
    return files + [zip_path]


def main() -> None:
    parser = argparse.ArgumentParser(description="Build online stress-test workbooks.")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT)
    parser.add_argument(
        "--mode",
        choices=("stress", "quick-start", "all"),
        default="stress",
        help="stress builds filled regression workbooks; quick-start builds daily-use blank client templates.",
    )
    args = parser.parse_args()
    if args.mode == "stress":
        files = build_all(args.out_dir)
    elif args.mode == "quick-start":
        files = build_quick_start_all(args.out_dir)
    else:
        files = build_all(args.out_dir / "stress") + build_quick_start_all(args.out_dir / "quick_start")
    for path in files:
        print(path)


if __name__ == "__main__":
    main()
