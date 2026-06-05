from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "outputs" / "AI适配_新公司注册资料模板.xlsx"


def main() -> None:
    wb = load_workbook(TEMPLATE)
    ws = wb["Company"]
    headers = [cell.value for cell in ws[1]]
    key_col = headers.index("field_key") + 1
    value_col = headers.index("value") + 1
    required_col = headers.index("required") + 1
    note_col = headers.index("AI备注") + 1

    notes = {
        "first_fye": "可空，系统按注册日期自动计算；特殊首个财年结束日才填写 DD/MM/YYYY",
        "fye": "可空，系统按注册日期自动取财年月份；特殊财年月份才填写 DD/MM/YYYY 或月份",
    }
    labels = {
        "first_fye": "首个财政年结日（可空自动计算）",
        "fye": "财政年度月份/年结日（可空自动计算）",
    }
    for row_no in range(2, ws.max_row + 1):
        key = ws.cell(row_no, key_col).value
        if key in notes:
            ws.cell(row_no, 2).value = labels[key]
            ws.cell(row_no, value_col).value = ""
            ws.cell(row_no, required_col).value = "Auto"
            ws.cell(row_no, note_col).value = notes[key]

    wb.save(TEMPLATE)
    print(TEMPLATE)


if __name__ == "__main__":
    main()
