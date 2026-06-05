from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "app" / "data" / "secretary_files.db"
TEMPLATE = ROOT / "outputs" / "AI适配_新公司注册资料模板_v3_人性化竖排_Auto版_空白.xlsx"
OUTPUT = ROOT / "outputs" / "p1_v3_human_vertical_common_people_stress_test.xlsx"


def clean(value: Any) -> str:
    return str(value or "").strip()


def pick_common_person(keyword: str, fallback: str) -> str:
    if not DB_PATH.exists():
        return fallback
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT display_name, default_role, notes
        FROM common_people
        WHERE active = 1 AND lower(default_role) LIKE ?
        ORDER BY
          CASE WHEN notes LIKE '示例%' THEN 1 ELSE 0 END,
          id
        """,
        (f"%{keyword.lower()}%",),
    ).fetchall()
    return rows[0]["display_name"] if rows else fallback


def find_header_row(ws, key_name: str = "field_key") -> tuple[int, list[str]]:
    for row_no in range(1, min(ws.max_row, 15) + 1):
        headers = [clean(cell.value) for cell in ws[row_no]]
        if key_name in headers:
            return row_no, headers
    raise ValueError(f"Cannot find {key_name} in {ws.title}")


def fill_single_value_sheet(wb, sheet_name: str, values: dict[str, Any]) -> None:
    ws = wb[sheet_name]
    header_row, headers = find_header_row(ws)
    key_col = headers.index("field_key") + 1
    value_col = headers.index("填写内容 / Value") + 1
    for row_no in range(header_row + 1, ws.max_row + 1):
        key = clean(ws.cell(row_no, key_col).value)
        if key in values:
            ws.cell(row_no, value_col).value = values[key]


def fill_transposed_sheet(wb, sheet_name: str, prefix: str, objects: list[dict[str, Any]]) -> None:
    ws = wb[sheet_name]
    header_row, headers = find_header_row(ws)
    key_col = headers.index("field_key") + 1
    object_cols = []
    for idx, header in enumerate(headers, start=1):
        if clean(header).startswith(prefix):
            object_cols.append(idx)
    for row_no in range(header_row + 1, ws.max_row + 1):
        key = clean(ws.cell(row_no, key_col).value)
        if not key:
            continue
        for obj_idx, col_no in enumerate(object_cols):
            ws.cell(row_no, col_no).value = objects[obj_idx].get(key, "") if obj_idx < len(objects) else ""


def main() -> None:
    nominee_name = pick_common_person("nominee director", "挂名董事 A")
    secretary_name = pick_common_person("secretary", "公司秘书 A")
    wb = load_workbook(TEMPLATE)
    company = {
        "company_name": "STRESS VERTICAL P1 COMMON PEOPLE TEST PTE. LTD.",
        "business_order_id": "INC-STRESS-001",
        "source_type": "Excel",
        "source_file_id": "stress-test-source",
        "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A PENINSULA PLAZA, SINGAPORE 179098, WITH EXTRA LONG UNIT DESCRIPTION FOR WRAP AND PDF LAYOUT TESTING",
        "incorporation_date": "01/06/2026",
        "business_activity_1": "Management consultancy services and corporate operations advisory with intentionally long business description for layout stress testing",
        "ssic_code_1": "70201",
        "business_activity_2": "Development of software and applications except games and cybersecurity, used as a secondary activity for stress testing",
        "ssic_code_2": "62011",
        "currency": "SGD",
        "share_class_default": "Ordinary",
        "total_issued_shares": 1000,
        "paid_up_capital": 1000,
        "contact_person_id": "P003",
        "client_signatory_person_id": "P003",
        "remarks": "P1 v3 human vertical stress test. Nominee director and secretary are pulled from website common people settings.",
    }
    people = [
        {
            "person_id": "P001",
            "source": "common",
            "common_person_name": nominee_name,
            "appointment_date": "01/06/2026",
            "remarks": "Only source + common_person_name is filled. Role and personal details should be inferred from website settings.",
        },
        {
            "person_id": "P002",
            "source": "common",
            "common_person_name": secretary_name,
            "appointment_date": "01/06/2026",
            "remarks": "Only source + common_person_name is filled. Role and personal details should be inferred from website settings.",
        },
        {
            "person_id": "P003",
            "source": "new",
            "full_name": "ZHANG YI",
            "id_type": "Passport",
            "id_number": "P1000001A",
            "nationality": "Chinese",
            "date_of_birth": "03/03/1988",
            "residential_address": "ROOM 1808, BUILDING 1, INTERNATIONAL COMMERCE CENTRE, 999 VERY LONG AVENUE NAME, CHAOYANG DISTRICT, BEIJING 100020, PEOPLE'S REPUBLIC OF CHINA",
            "email": "zhang.yi.client.director.shareholder.with.long.email@example-long-domain.cn",
            "phone": "+86 138 0000 0001",
            "is_director": "Yes",
            "appointment_date": "01/06/2026",
            "remarks": "Client director/shareholder 1; shareholder and client representative signer are left blank for auto inference.",
        },
        {
            "person_id": "P004",
            "source": "new",
            "full_name": "ZHANG ER",
            "id_type": "Passport",
            "id_number": "P2000002B",
            "nationality": "Chinese",
            "date_of_birth": "04/04/1990",
            "residential_address": "UNIT 3201, TOWER B, SHENZHEN BAY TECHNOLOGY ECOLOGY PARK, NANSHAN DISTRICT, SHENZHEN, GUANGDONG 518057, CHINA, ADDITIONAL ADDRESS LINE FOR WRAP TEST",
            "email": "zhang.er.second.director.shareholder.long-address-test@example-long-domain.cn",
            "phone": "+86 139 0000 0002",
            "is_director": "Yes",
            "appointment_date": "01/06/2026",
            "remarks": "Client director/shareholder 2; shareholder and client representative signer are left blank for auto inference.",
        },
        {
            "person_id": "P005",
            "source": "new",
            "full_name": "ZHANG SAN",
            "id_type": "Passport",
            "id_number": "P3000003C",
            "nationality": "Chinese",
            "date_of_birth": "05/05/1992",
            "residential_address": "FLAT 12A, 88 QUEENS ROAD CENTRAL, CENTRAL, HONG KONG, WITH LONG BUILDING NAME AND MULTIPLE DESCRIPTOR SEGMENTS FOR LINE WRAPPING TEST",
            "email": "zhang.san.third.director.shareholder.long-email@example-long-domain.hk",
            "phone": "+852 5123 0003",
            "is_director": "Yes",
            "appointment_date": "01/06/2026",
            "remarks": "Client director/shareholder 3; shareholder and client representative signer are left blank for auto inference.",
        },
    ]
    shareholders = [
        {
            "shareholder_type": "person",
            "person_full_name": "ZHANG YI",
            "person_id_number": "P1000001A",
            "share_class": "Ordinary",
            "shares": 600,
            "paid_amount": 600,
            "currency": "SGD",
            "remarks": "Shareholder 1, 60 percent. Default client representative.",
        },
        {
            "shareholder_type": "person",
            "person_full_name": "ZHANG ER",
            "person_id_number": "P2000002B",
            "share_class": "Ordinary",
            "shares": 300,
            "paid_amount": 300,
            "currency": "SGD",
            "remarks": "Shareholder 2, 30 percent.",
        },
        {
            "shareholder_type": "person",
            "person_full_name": "ZHANG SAN",
            "person_id_number": "P3000003C",
            "share_class": "Ordinary",
            "shares": 100,
            "paid_amount": 100,
            "currency": "SGD",
            "remarks": "Shareholder 3, 10 percent.",
        },
    ]
    output_options = {
        "output_format": "docx_pdf_zip",
        "prepared_by": "Codex stress test",
        "signing_mode": "default",
        "client_signatory_logic": "shareholder_1",
    }
    fill_single_value_sheet(wb, "公司信息", company)
    fill_transposed_sheet(wb, "人员信息", "人员", people)
    fill_transposed_sheet(wb, "股东与股份", "股东", shareholders)
    fill_single_value_sheet(wb, "输出设置", output_options)
    wb.save(OUTPUT)
    print(OUTPUT)
    print(f"common_nominee={nominee_name}")
    print(f"common_secretary={secretary_name}")


if __name__ == "__main__":
    main()
