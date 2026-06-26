from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "outputs" / "M05_annual_review_stress_input_v03.xlsx"


COMPANY_FIELDS = {
    "company_name": "RBIZ ANNUAL REVIEW STRESS PTE. LTD.",
    "uen": "202455551H",
    "registered_office_address": "111 NORTH BRIDGE ROAD, #29-06A, PENINSULA PLAZA, SINGAPORE 179098",
    "total_issued_shares": 100000,
    "issued_share_capital": 100000,
    "paid_up_capital": 100000,
    "currency": "SGD",
    "default_document_date": "30/06/2026",
    "director_signer_names": "ZHANG ANNUAL DIRECTOR\nLEE SECOND DIRECTOR",
    "member_signer_names": "ZHANG ANNUAL SHAREHOLDER\nGLOBAL HOLDING PTE. LTD.",
    "client_signatory_name": "ZHANG ANNUAL DIRECTOR",
    "business_order_id": "M05-STRESS-001",
    "source_type": "AI",
    "source_file_id": "BizFile + financial statements draft",
    "prepared_by": "Codex test",
    "change_registered_office_required": "No",
    "change_business_activity_required": "No",
    "change_fye_required": "No",
    "transfer_in_required": "No",
    "annual_review_required": "Yes",
    "fye_date": "31/12/2025",
    "agm_date": "30/06/2026",
    "annual_review_remarks": "M05 stress test: ordinary AGM, accounts_status=active, multiple directors and members.",
}

ANNUAL_FIELDS = {
    "annual_review_required": "Yes",
    "fye_date": "31/12/2025",
    "agm_date": "30/06/2026",
    "agm_time": "10.00 a.m.",
    "agm_place": "111 NORTH BRIDGE ROAD, #29-06A, PENINSULA PLAZA, SINGAPORE 179098",
    "agm_route": "ordinary_agm",
    "accounts_status": "active",
    "company_activity_status": "",
    "financial_statements_type": "",
    "financial_statements_required": "",
    "audit_exemption_status": "",
    "agm_status": "Held AGM",
    "acra_dormant_relevant_company": "",
    "total_assets_under_500k": "",
    "iras_tax_status": "",
    "financial_statement_date": "31/12/2025",
    "director_signer_name": "ZHANG ANNUAL DIRECTOR\nLEE SECOND DIRECTOR",
    "shareholder_signer_name": "ZHANG ANNUAL SHAREHOLDER\nGLOBAL HOLDING PTE. LTD.",
    "ar_authorized_signer_name": "ZHANG ANNUAL DIRECTOR",
    "directors_fee": "0",
    "directors_remuneration": "0",
    "shorter_notice_consent": "Auto",
    "management_rep_letter": "Yes",
    "remarks": "Use as common annual review test data.",
}


def template_path() -> Path:
    candidates = sorted((ROOT / "outputs").glob("*v7*空白.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError("Cannot find the P2 v7 blank template in outputs.")
    return candidates[0]


def set_kv_by_key(wb, key: str, value) -> bool:
    found = False
    for ws in wb.worksheets:
        for row in range(1, ws.max_row + 1):
            if str(ws.cell(row, 3).value or "").strip() == key:
                ws.cell(row, 4).value = value
                found = True
    return found


def main() -> None:
    wb = load_workbook(template_path())
    for key, value in COMPANY_FIELDS.items():
        if not set_kv_by_key(wb, key, value):
            raise KeyError(f"Field key not found: {key}")
    for key, value in ANNUAL_FIELDS.items():
        if not set_kv_by_key(wb, key, value):
            raise KeyError(f"Annual field key not found: {key}")
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
