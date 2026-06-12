from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


YES = {"yes", "y", "true", "1", "是", "有", "需要"}
NO = {"no", "n", "false", "0", "否", "不", "无需", "不用"}
AUTO = {"", "auto", "自动", "默认", "system", "系统"}
SINGAPORE_TZ = ZoneInfo("Asia/Singapore")
DEFAULT_PROVIDER_REGISTERED_ADDRESS = "111 NORTH BRIDGE ROAD, #29-06A, PENINSULA PLAZA, SINGAPORE 179098"


def clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def is_yes(value: Any) -> bool:
    return str(clean(value)).strip().lower() in YES


def is_no(value: Any) -> bool:
    return str(clean(value)).strip().lower() in NO


def is_auto(value: Any) -> bool:
    return str(clean(value)).strip().lower() in AUTO


def today_document_date() -> str:
    return datetime.now(SINGAPORE_TZ).strftime("%d/%m/%Y")


def apply_default_document_date(company: dict[str, Any]) -> None:
    if not clean(company.get("default_document_date")):
        company["default_document_date"] = today_document_date()


def read_kv_sheet(wb, sheet_name: str, value_col: str = "value") -> dict[str, Any]:
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    headers = [clean(c.value) for c in ws[1]]
    result: dict[str, Any] = {}
    try:
        key_idx = headers.index("field_key")
    except ValueError:
        key_idx = 0
    value_idx = headers.index(value_col) if value_col in headers else min(2, len(headers) - 1)
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = clean(row[key_idx] if key_idx < len(row) else "")
        if not key:
            continue
        result[str(key)] = clean(row[value_idx] if value_idx < len(row) else "")
    return result


def read_table(wb, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [str(clean(c.value)) for c in ws[1]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: clean(row[i] if i < len(row) else "") for i in range(len(headers)) if headers[i]}
        if any(v != "" for v in item.values()):
            rows.append(item)
    return rows


def header_field_key(value: Any) -> str:
    raw = str(clean(value)).strip()
    if not raw:
        return ""
    parts = re.split(r"[\r\n]+", raw)
    for part in reversed(parts):
        candidate = part.strip().strip("()（）[]【】")
        if re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", candidate):
            return candidate
    compact = re.sub(r"[^A-Za-z0-9_]", "", raw)
    if re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", compact):
        return compact
    return raw


def read_keyed_table(wb, sheet_name: str) -> list[dict[str, Any]]:
    """Read v5 human-facing tables whose headers show Chinese labels plus field keys."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    header_row = 1
    keys: list[str] = []
    for row_no in range(1, min(ws.max_row, 10) + 1):
        candidate = [header_field_key(c.value) for c in ws[row_no]]
        system_keys = [key for key in candidate if re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", key or "")]
        if len(system_keys) >= 2:
            header_row = row_no
            keys = candidate
            break
    if not keys:
        keys = [header_field_key(c.value) for c in ws[1]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item = {keys[i]: clean(row[i] if i < len(row) else "") for i in range(len(keys)) if keys[i]}
        if any(value != "" for value in item.values()):
            rows.append(item)
    return rows


def norm_header(value: Any) -> str:
    return str(clean(value)).strip().lower()


def find_header_row(ws, key_names: set[str]) -> tuple[int, list[Any]]:
    max_scan = min(ws.max_row, 15)
    lowered = {item.lower() for item in key_names}
    for row_no in range(1, max_scan + 1):
        headers = [clean(c.value) for c in ws[row_no]]
        if any(norm_header(header) in lowered for header in headers):
            return row_no, headers
    return 1, [clean(c.value) for c in ws[1]]


def header_index(headers: list[Any], names: set[str], fallback: int = 0) -> int:
    lowered = {name.lower() for name in names}
    for idx, header in enumerate(headers):
        if norm_header(header) in lowered:
            return idx
    return fallback


def read_vertical_kv_sheet(wb, sheet_name: str) -> dict[str, Any]:
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    header_row, headers = find_header_row(ws, {"field_key", "option_key"})
    key_idx = header_index(headers, {"field_key", "option_key"}, 2 if len(headers) > 2 else 0)
    value_idx = header_index(
        headers,
        {"value", "填写内容", "填写内容 / value", "填写内容/value", "current_value"},
        min(key_idx + 1, len(headers) - 1),
    )
    result: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        key = clean(row[key_idx] if key_idx < len(row) else "")
        if not key:
            continue
        result[str(key)] = clean(row[value_idx] if value_idx < len(row) else "")
    return result


def read_transposed_table(wb, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    header_row, headers = find_header_row(ws, {"field_key"})
    key_idx = header_index(headers, {"field_key"}, 2 if len(headers) > 2 else 0)
    metadata_headers = {
        "项目",
        "item",
        "english",
        "english name",
        "field_key",
        "说明",
        "notes",
        "note",
        "必填",
        "required",
    }
    object_cols = [
        idx
        for idx, header in enumerate(headers)
        if idx > key_idx and clean(header) and norm_header(header) not in metadata_headers
    ]
    rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    items: list[dict[str, Any]] = []
    for col_idx in object_cols:
        item: dict[str, Any] = {}
        for row in rows:
            key = clean(row[key_idx] if key_idx < len(row) else "")
            if not key:
                continue
            item[str(key)] = clean(row[col_idx] if col_idx < len(row) else "")
        if any(value != "" for value in item.values()):
            items.append(item)
    return items


def detect_task_type(wb) -> str:
    if "P2快速业务单" in wb.sheetnames:
        return "maintenance"
    if "P2公司信息" in wb.sheetnames:
        return "maintenance"
    if "维护公司信息" in wb.sheetnames and {"人员信息", "变更事项"} & set(wb.sheetnames):
        return "maintenance"
    if "公司信息" in wb.sheetnames and ({"人员信息", "股东与股份"} & set(wb.sheetnames)):
        company = read_vertical_kv_sheet(wb, "公司信息")
        task_type = str(company.get("task_type", "incorporation")).lower()
        if task_type in {"incorporation", "change", "maintenance"}:
            return task_type
        return "incorporation"
    if "Company" in wb.sheetnames:
        company = read_kv_sheet(wb, "Company", "value")
        task_type = str(company.get("task_type", "")).lower()
        if task_type in {"incorporation", "change", "maintenance"}:
            return task_type
    if {"ChangeEvents", "AnnualReview", "ShareTransfers", "ShareAllotments"} & set(wb.sheetnames):
        return "maintenance"
    if "Company_Current" in wb.sheetnames or "Company_Changes" in wb.sheetnames:
        return "change"
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            joined = " ".join(str(clean(v)).lower() for v in row if clean(v) != "")
            if "incorporation" in joined:
                return "incorporation"
            if "change" in joined or "变更" in joined:
                return "change"
    return "unknown"


def parse_excel(path: str | Path, common_people: dict[str, dict[str, Any]] | None = None) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False)
    task_type = detect_task_type(wb)
    if task_type == "incorporation":
        parsed = parse_registration(wb, common_people or {})
    elif task_type == "maintenance":
        parsed = parse_maintenance(wb, common_people or {})
    elif task_type == "change":
        parsed = parse_change(wb, common_people or {})
    else:
        parsed = parse_unknown(wb)
    parsed["task_type"] = task_type
    return parsed


def parse_registration(wb, common_people: dict[str, dict[str, Any]]) -> dict[str, Any]:
    if "公司信息" in wb.sheetnames:
        return parse_human_registration(wb, common_people)
    company = read_kv_sheet(wb, "Company", "value")
    people = resolve_people(read_table(wb, "People"), common_people)
    shareholders = read_table(wb, "Shareholders")
    generation = option_map(read_table(wb, "Generation"))
    return {
        "company": company,
        "people": people,
        "shareholders": shareholders,
        "generation": generation,
        "changes": [],
        "personal_changes": [],
        "share_changes": [],
    }


def parse_human_registration(wb, common_people: dict[str, dict[str, Any]]) -> dict[str, Any]:
    company = read_vertical_kv_sheet(wb, "公司信息")
    company.setdefault("task_type", "incorporation")
    shareholders = normalize_human_shareholders(read_transposed_table(wb, "股东与股份"), company)
    people = resolve_people(normalize_human_people(read_transposed_table(wb, "人员信息"), company), common_people)
    apply_human_registration_role_defaults(people, shareholders)
    generation = read_vertical_kv_sheet(wb, "输出设置")
    return {
        "company": company,
        "people": people,
        "shareholders": shareholders,
        "generation": generation,
        "changes": [],
        "personal_changes": [],
        "share_changes": [],
    }


def normalize_human_people(rows: list[dict[str, Any]], company: dict[str, Any]) -> list[dict[str, Any]]:
    default_date = company.get("incorporation_date", "")
    out = []
    for row in rows:
        item = dict(row)
        if not item.get("source"):
            item["source"] = "new"
        if default_date and not item.get("appointment_date"):
            item["appointment_date"] = default_date
        out.append(item)
    return out


def normalize_human_shareholders(rows: list[dict[str, Any]], company: dict[str, Any]) -> list[dict[str, Any]]:
    currency = company.get("currency") or "SGD"
    share_class = company.get("share_class_default") or "Ordinary"
    out = []
    for row in rows:
        item = dict(row)
        if not item.get("shareholder_type"):
            item["shareholder_type"] = "corporate" if item.get("corporate_name") else "person"
        if not item.get("currency"):
            item["currency"] = currency
        if not item.get("share_class"):
            item["share_class"] = share_class
        if item.get("shares") and not item.get("issued_share_capital"):
            item["issued_share_capital"] = item["shares"]
        if item.get("issued_share_capital") and not item.get("paid_amount"):
            item["paid_amount"] = item["issued_share_capital"]
        out.append(item)
    return out


def apply_human_registration_role_defaults(people: list[dict[str, Any]], shareholders: list[dict[str, Any]]) -> None:
    shareholder_names = {str(row.get("person_full_name", "")).strip().lower() for row in shareholders if row.get("person_full_name")}
    shareholder_ids = {str(row.get("person_id_number", "")).strip().lower() for row in shareholders if row.get("person_id_number")}
    boolean_keys = [
        "is_director",
        "is_local_resident_director",
        "is_nominee_director",
        "is_secretary",
        "is_shareholder",
        "is_authorized_rep",
        "signing_required",
    ]
    for person in people:
        name = str(person.get("full_name") or person.get("common_person_name") or "").strip().lower()
        ident = str(person.get("id_number") or "").strip().lower()
        matched_shareholder = bool((name and name in shareholder_names) or (ident and ident in shareholder_ids))
        if is_auto(person.get("is_shareholder")):
            person["is_shareholder"] = "Yes" if matched_shareholder else "No"
        for key in boolean_keys:
            if key != "signing_required" and is_auto(person.get(key)):
                person[key] = "No"
        if is_auto(person.get("signing_required")):
            person["signing_required"] = "Yes" if any(is_yes(person.get(key)) for key in ["is_director", "is_secretary", "is_shareholder"]) else "No"


def parse_change(wb, common_people: dict[str, dict[str, Any]]) -> dict[str, Any]:
    current = read_kv_sheet(wb, "Company_Current", "current_value")
    company_changes = read_table(wb, "Company_Changes")
    people = resolve_people(read_table(wb, "People"), common_people)
    personal_changes = read_table(wb, "Personal_Changes")
    shareholders_current = read_table(wb, "Shareholders_Current")
    share_changes = read_table(wb, "Share_Changes")
    generation = option_map(read_table(wb, "Generation"))
    return {
        "company": current,
        "people": people,
        "shareholders": shareholders_current,
        "generation": generation,
        "changes": company_changes,
        "personal_changes": personal_changes,
        "share_changes": share_changes,
    }


def parse_maintenance(wb, common_people: dict[str, dict[str, Any]]) -> dict[str, Any]:
    if "P2快速业务单" in wb.sheetnames:
        return parse_one_page_maintenance(wb, common_people)
    if "P2公司信息" in wb.sheetnames:
        return parse_quick_maintenance(wb, common_people)
    if "维护公司信息" in wb.sheetnames:
        return parse_human_maintenance(wb, common_people)
    company = read_kv_sheet(wb, "Company", "value")
    apply_default_document_date(company)
    people = resolve_people(read_table(wb, "People"), common_people)
    shareholdings = read_table(wb, "Shareholdings")
    change_events = read_table(wb, "ChangeEvents")
    share_transfers = read_table(wb, "ShareTransfers")
    share_allotments = read_table(wb, "ShareAllotments")
    annual_review = read_kv_sheet(wb, "AnnualReview", "value")
    output_options = option_map(read_table(wb, "OutputOptions"))
    return {
        "company": company,
        "people": people,
        "shareholders": shareholdings,
        "shareholdings": shareholdings,
        "generation": output_options,
        "output_options": output_options,
        "changes": change_events,
        "change_events": change_events,
        "personal_changes": [],
        "share_changes": share_transfers,
        "share_transfers": share_transfers,
        "share_allotments": share_allotments,
        "annual_review": annual_review,
    }


EVENT_NAME_CN = {
    "change_registered_office": "注册地址变更",
    "change_office_hours": "办公时间变更",
    "change_business_activity": "营业范围/SSIC 变更",
    "change_fye": "财政年结日变更",
    "update_officer_particulars": "人员资料更新",
    "appoint_director": "委任董事",
    "resign_director": "董事辞任",
    "appoint_secretary": "委任秘书",
    "resign_secretary": "秘书辞任",
    "transfer_in": "转入",
    "transfer_in_cooperative": "配合转入",
    "transfer_in_non_cooperative": "不配合转入",
}

PERSON_ACTIONS = {
    "委任董事": "appoint_director",
    "appoint director": "appoint_director",
    "appoint_director": "appoint_director",
    "董事辞任": "resign_director",
    "辞任董事": "resign_director",
    "resign director": "resign_director",
    "resign_director": "resign_director",
    "委任秘书": "appoint_secretary",
    "appoint secretary": "appoint_secretary",
    "appoint_secretary": "appoint_secretary",
    "秘书辞任": "resign_secretary",
    "辞任秘书": "resign_secretary",
    "resign secretary": "resign_secretary",
    "resign_secretary": "resign_secretary",
    }


def parse_one_page_maintenance(wb, common_people: dict[str, dict[str, Any]]) -> dict[str, Any]:
    ws = wb["P2快速业务单"]
    fields = read_one_page_kv(ws)
    apply_default_document_date(fields)
    annual_quick = read_vertical_kv_sheet(wb, "快速年审") if "快速年审" in wb.sheetnames else {}
    company = dict(fields)
    company.setdefault("task_type", "maintenance")
    company.setdefault("template_version", "p2_v6_one_page")
    apply_transfer_in_registered_office_defaults(company, fields, fields, {})
    one_page_sections = read_one_page_sections(ws)
    person_action_rows = one_page_sections.get("person_actions", [])
    person_particular_rows = one_page_sections.get("personal_particulars", [])
    people = one_page_people(fields)
    people.extend(one_page_referenced_people(person_action_rows + person_particular_rows, common_people))
    people = resolve_people(people, common_people)
    change_events: list[dict[str, Any]] = []
    change_events.extend(quick_company_change_events(company, fields, {}))
    change_events.extend(quick_person_action_events(read_one_page_table(ws, "人员任免"), company))
    change_events.extend(quick_person_particular_events(read_one_page_table(ws, "个人资料变更"), company))
    change_events.extend(quick_transfer_in_events(company, fields, {}))
    share_transfers = normalize_quick_rows(read_one_page_table(ws, "股份转让"), company)
    share_allotments = normalize_quick_rows(read_one_page_table(ws, "增资配股"), company)
    annual_review = normalize_quick_annual_review(company, fields, annual_quick)
    output_options = {
        "package": "AUTO_MAINTENANCE_ONE_PAGE",
        "output_format": "pdf_zip",
        "notes": fields.get("output_notes", ""),
    }
    return {
        "company": company,
        "people": people,
        "shareholders": [],
        "shareholdings": [],
        "generation": output_options,
        "output_options": output_options,
        "changes": change_events,
        "change_events": change_events,
        "personal_changes": [],
        "share_changes": share_transfers,
        "share_transfers": share_transfers,
        "share_allotments": share_allotments,
        "annual_review": annual_review,
    }


def normalize_quick_annual_review(company: dict[str, Any], fields: dict[str, Any], annual_quick: dict[str, Any]) -> dict[str, Any]:
    annual: dict[str, Any] = {
        "annual_review_required": annual_quick.get("annual_review_required") or fields.get("annual_review_required", "No"),
        "fye_date": annual_quick.get("fye_date") or fields.get("fye_date", ""),
        "agm_date": annual_quick.get("agm_date") or fields.get("agm_date", ""),
        "agm_time": annual_quick.get("agm_time") or "10.00 a.m.",
        "agm_place": annual_quick.get("agm_place") or company.get("registered_office_address", ""),
        "agm_route": annual_quick.get("agm_route") or fields.get("agm_route") or "ordinary_agm",
        "accounts_status": annual_quick.get("accounts_status") or "non_dormant",
        "company_activity_status": annual_quick.get("company_activity_status") or fields.get("company_activity_status", ""),
        "financial_statements_type": annual_quick.get("financial_statements_type") or fields.get("financial_statements_type", ""),
        "financial_statements_required": annual_quick.get("financial_statements_required") or fields.get("financial_statements_required", ""),
        "audit_exemption_status": annual_quick.get("audit_exemption_status") or fields.get("audit_exemption_status", ""),
        "agm_status": annual_quick.get("agm_status") or fields.get("agm_status", ""),
        "acra_dormant_relevant_company": annual_quick.get("acra_dormant_relevant_company") or fields.get("acra_dormant_relevant_company", ""),
        "total_assets_under_500k": annual_quick.get("total_assets_under_500k") or fields.get("total_assets_under_500k", ""),
        "iras_tax_status": annual_quick.get("iras_tax_status") or fields.get("iras_tax_status", ""),
        "financial_statement_date": annual_quick.get("financial_statement_date") or annual_quick.get("fye_date") or fields.get("fye_date", ""),
        "financial_year_start": annual_quick.get("financial_year_start") or "",
        "director_signer_name": annual_quick.get("director_signer_name") or company.get("director_signer_name", ""),
        "shareholder_signer_name": annual_quick.get("shareholder_signer_name") or company.get("client_signatory_name", ""),
        "ar_authorized_signer_name": annual_quick.get("ar_authorized_signer_name") or annual_quick.get("director_signer_name") or company.get("director_signer_name", ""),
        "directors_fee": annual_quick.get("directors_fee") or "0",
        "directors_remuneration": annual_quick.get("directors_remuneration") or "0",
        "shorter_notice_consent": annual_quick.get("shorter_notice_consent") or "Auto",
        "management_rep_letter": annual_quick.get("management_rep_letter") or "Yes",
        "remarks": annual_quick.get("remarks") or fields.get("annual_review_remarks", ""),
    }
    return annual


def read_one_page_kv(ws) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for row in range(1, ws.max_row + 1):
        key = clean(ws.cell(row, 3).value)
        if not isinstance(key, str) or key == "field_key":
            continue
        if re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", key):
            result[key] = clean(ws.cell(row, 4).value)
    return result


def read_one_page_table(ws, section_name: str) -> list[dict[str, Any]]:
    start_row = None
    for row in range(1, ws.max_row + 1):
        if clean(ws.cell(row, 1).value) == section_name:
            start_row = row
            break
    if not start_row:
        return []
    header_row = start_row + 1
    headers = [header_field_key(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    items: list[dict[str, Any]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        values = [clean(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if not any(values):
            break
        if values[0] in {"人员任免", "个人资料变更", "股份转让", "增资配股"}:
            break
        item = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values))) if headers[idx]}
        if any(value != "" for value in item.values()):
            items.append(item)
    return items


def read_one_page_sections(ws) -> dict[str, list[dict[str, Any]]]:
    sections: dict[str, list[dict[str, Any]]] = {}
    section_rules = [
        ("person_actions", {"generate", "action_type", "target_name"}),
        ("personal_particulars", {"generate", "field_label", "old_value", "new_value"}),
        ("share_transfers", {"generate", "transferor_name", "transferee_name", "shares_transferred"}),
        ("share_allotments", {"generate", "allottee_name", "shares_allotted"}),
    ]
    for row in range(1, ws.max_row):
        headers = [header_field_key(ws.cell(row + 1, col).value) for col in range(1, ws.max_column + 1)]
        header_set = {key for key in headers if key}
        section_key = ""
        for candidate, required_keys in section_rules:
            if required_keys.issubset(header_set):
                section_key = candidate
                break
        if not section_key or section_key in sections:
            continue
        items: list[dict[str, Any]] = []
        for data_row in range(row + 2, ws.max_row + 1):
            values = [clean(ws.cell(data_row, col).value) for col in range(1, ws.max_column + 1)]
            if not any(values):
                break
            item = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values))) if headers[idx]}
            if any(value != "" for value in item.values()):
                items.append(item)
        sections[section_key] = items
    return sections


def one_page_people(fields: dict[str, Any]) -> list[dict[str, Any]]:
    people: list[dict[str, Any]] = []
    director_signers = split_people_names(fields.get("director_signer_names") or fields.get("director_signer_name"))
    for index, signer in enumerate(director_signers, start=1):
        people.append(
            {
                "person_id": f"SIGNER{index}",
                "source": "new",
                "full_name": signer,
                "is_director": "Yes",
                "is_secretary": "No",
                "is_shareholder": "No",
                "is_client_signatory": "Yes" if index == 1 else "No",
            }
        )
    client_signer = clean(fields.get("client_signatory_name"))
    if client_signer and client_signer.lower() not in {name.lower() for name in director_signers}:
        people.append(
            {
                "person_id": "CLIENT_SIGNER1",
                "source": "new",
                "full_name": client_signer,
                "is_director": "No",
                "is_secretary": "No",
                "is_shareholder": "No",
                "is_client_signatory": "Yes",
            }
        )
    return people


def split_people_names(value: Any) -> list[str]:
    text_value = clean(value)
    if not text_value:
        return []
    return [item.strip() for item in re.split(r"[,;/\n]+", text_value) if item.strip()]


def one_page_referenced_people(rows: list[dict[str, Any]], common_people: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    people: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, row in enumerate(rows, start=1):
        name = clean(row.get("target_name") or row.get("person_name"))
        person_id = clean(row.get("target_person_id") or row.get("person_id"))
        if not name and not person_id:
            continue
        key = (person_id or name).lower()
        if key in seen:
            continue
        seen.add(key)
        matched_common = name in common_people
        people.append(
            {
                "person_id": person_id or f"ONEPAGE{index}",
                "source": "common" if matched_common else "new",
                "common_person_name": name if matched_common else "",
                "full_name": "" if matched_common else name,
                "id_type": clean(row.get("target_id_type") or row.get("id_type")),
                "id_number": clean(row.get("target_id_number") or row.get("id_number")),
                "residential_address": clean(row.get("target_address") or row.get("residential_address") or row.get("address")),
                "address": clean(row.get("target_address") or row.get("residential_address") or row.get("address")),
                "email": clean(row.get("target_email") or row.get("email")),
                "phone": clean(row.get("target_phone") or row.get("phone")),
                "is_director": "No",
                "is_secretary": "No",
                "is_shareholder": "No",
                "is_client_signatory": "No",
            }
        )
    return people


def parse_quick_maintenance(wb, common_people: dict[str, dict[str, Any]]) -> dict[str, Any]:
    company = read_vertical_kv_sheet(wb, "P2公司信息")
    company.setdefault("task_type", "maintenance")
    company.setdefault("template_version", "p2_v5")
    apply_default_document_date(company)

    task_rows = read_keyed_table(wb, "办理事项")
    task_flags = {str(row.get("task_key") or "").strip(): row for row in task_rows if row.get("task_key")}
    company_changes = read_vertical_kv_sheet(wb, "公司资料变更")
    transfer_in = read_vertical_kv_sheet(wb, "转入交接")
    apply_transfer_in_registered_office_defaults(company, company_changes, transfer_in, task_flags)
    output_options = read_vertical_kv_sheet(wb, "输出设置")

    shareholdings = read_transposed_table(wb, "股东现状")
    people = resolve_people(normalize_human_maintenance_people(read_transposed_table(wb, "人员信息"), company), common_people)
    apply_human_maintenance_role_defaults(people, shareholdings)

    change_events: list[dict[str, Any]] = []
    change_events.extend(quick_company_change_events(company, company_changes, task_flags))
    change_events.extend(quick_person_action_events(read_keyed_table(wb, "人员任免"), company))
    change_events.extend(quick_person_particular_events(read_keyed_table(wb, "个人资料变更"), company))
    change_events.extend(quick_transfer_in_events(company, transfer_in, task_flags))

    share_transfers = normalize_quick_rows(read_keyed_table(wb, "股份转让"), company)
    share_allotments = normalize_quick_rows(read_keyed_table(wb, "增资配股"), company)
    annual_review = read_vertical_kv_sheet(wb, "年审信息")
    return {
        "company": company,
        "people": people,
        "shareholders": shareholdings,
        "shareholdings": shareholdings,
        "generation": output_options,
        "output_options": output_options,
        "changes": change_events,
        "change_events": change_events,
        "personal_changes": [],
        "share_changes": share_transfers,
        "share_transfers": share_transfers,
        "share_allotments": share_allotments,
        "annual_review": annual_review,
    }


def quick_task_enabled(task_flags: dict[str, dict[str, Any]], key: str) -> bool:
    row = task_flags.get(key, {})
    return is_yes(row.get("required") or row.get("generate") or row.get("change_required"))


def quick_task_date(task_flags: dict[str, dict[str, Any]], key: str, company: dict[str, Any], fallback: Any = "") -> Any:
    row = task_flags.get(key, {})
    return row.get("effective_date") or fallback or company.get("default_document_date", "")


def transfer_in_enabled(fields: dict[str, Any], task_flags: dict[str, dict[str, Any]]) -> bool:
    if fields.get("transfer_in_required") not in (None, ""):
        return is_yes(fields.get("transfer_in_required"))
    return quick_task_enabled(task_flags, "transfer_in")


def normalized_address(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(clean(value)).lower())


def apply_transfer_in_registered_office_defaults(
    company: dict[str, Any],
    company_changes: dict[str, Any],
    transfer_in: dict[str, Any],
    task_flags: dict[str, dict[str, Any]],
) -> None:
    if not transfer_in_enabled(transfer_in, task_flags):
        return
    if is_no(company_changes.get("change_registered_office_required")):
        return

    new_address = clean(company_changes.get("new_registered_office_address") or company_changes.get("registered_office_new"))
    if not new_address:
        new_address = DEFAULT_PROVIDER_REGISTERED_ADDRESS
        company_changes["new_registered_office_address"] = new_address

    old_address = clean(company_changes.get("old_registered_office_address") or company.get("registered_office_address"))
    if old_address and not company_changes.get("old_registered_office_address"):
        company_changes["old_registered_office_address"] = old_address

    if company_changes.get("change_registered_office_required") in (None, ""):
        if normalized_address(old_address) != normalized_address(new_address):
            company_changes["change_registered_office_required"] = "Yes"

    if not clean(company_changes.get("registered_office_effective_date")):
        company_changes["registered_office_effective_date"] = transfer_in.get("effective_date") or quick_task_date(task_flags, "transfer_in", company)


def quick_company_change_events(company: dict[str, Any], fields: dict[str, Any], task_flags: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    if quick_detail_enabled(fields, "change_registered_office_required", task_flags, "change_registered_office", ["new_registered_office_address"]):
        new_address = fields.get("new_registered_office_address") or fields.get("registered_office_new")
        events.append(
            quick_event(
                "change_registered_office",
                quick_task_date(task_flags, "change_registered_office", company, fields.get("registered_office_effective_date")),
                old_value=fields.get("old_registered_office_address") or company.get("registered_office_address"),
                new_value=new_address,
            )
        )
    if quick_detail_enabled(
        fields,
        "change_business_activity_required",
        task_flags,
        "change_business_activity",
        ["new_primary_activity", "new_primary_ssic", "new_secondary_activity", "new_secondary_ssic"],
    ):
        events.append(
            quick_event(
                "change_business_activity",
                quick_task_date(task_flags, "change_business_activity", company, fields.get("business_activity_effective_date")),
                old_value=fields.get("old_primary_activity"),
                new_value=fields.get("new_primary_activity"),
                primary_ssic_old=fields.get("old_primary_ssic"),
                primary_activity_old=fields.get("old_primary_activity"),
                primary_ssic_new=fields.get("new_primary_ssic"),
                primary_activity_new=fields.get("new_primary_activity"),
                secondary_ssic_old=fields.get("old_secondary_ssic"),
                secondary_activity_old=fields.get("old_secondary_activity"),
                secondary_ssic_new=fields.get("new_secondary_ssic"),
                secondary_activity_new=fields.get("new_secondary_activity"),
            )
        )
    if quick_detail_enabled(fields, "change_fye_required", task_flags, "change_fye", ["new_fye"]):
        events.append(
            quick_event(
                "change_fye",
                quick_task_date(task_flags, "change_fye", company, fields.get("fye_effective_date")),
                old_value=fields.get("old_fye"),
                new_value=fields.get("new_fye"),
                next_accounts_period_start=fields.get("next_accounts_period_start"),
                next_accounts_period_end=fields.get("next_accounts_period_end") or fields.get("new_fye"),
            )
        )
    if quick_detail_enabled(fields, "change_office_hours_required", task_flags, "change_office_hours", ["new_office_hours"]):
        events.append(
            quick_event(
                "change_office_hours",
                quick_task_date(task_flags, "change_office_hours", company, fields.get("office_hours_effective_date")),
                new_value=fields.get("new_office_hours"),
            )
        )
    return events


def quick_detail_enabled(fields: dict[str, Any], field_key: str, task_flags: dict[str, dict[str, Any]], task_key: str, trigger_keys: list[str]) -> bool:
    value = fields.get(field_key)
    if value not in (None, ""):
        return is_yes(value)
    if task_key in task_flags:
        return quick_task_enabled(task_flags, task_key)
    return any(clean(fields.get(key)) for key in trigger_keys)


def quick_person_action_events(rows: list[dict[str, Any]], company: dict[str, Any]) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    default_date = company.get("default_document_date", "")
    for row in rows:
        if not is_yes(row.get("generate")):
            continue
        raw_action = str(row.get("action_type") or "").strip().lower()
        event_type = PERSON_ACTIONS.get(raw_action) or str(row.get("action_type") or "").strip()
        if not event_type:
            continue
        events.append(
            quick_event(
                event_type,
                row.get("effective_date") or default_date,
                target_person_id=row.get("target_person_id") or row.get("person_id"),
                target_name=row.get("target_name") or row.get("person_name"),
                target_id_type=row.get("target_id_type") or row.get("id_type"),
                target_id_number=row.get("target_id_number") or row.get("id_number"),
                target_address=row.get("target_address") or row.get("residential_address") or row.get("address"),
                new_value=row.get("new_capacity") or row.get("action_type"),
                resignation_letter=row.get("resignation_letter") or "No",
                remarks=row.get("remarks"),
            )
        )
    return events


def quick_person_particular_events(rows: list[dict[str, Any]], company: dict[str, Any]) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    default_date = company.get("default_document_date", "")
    for row in rows:
        if not is_yes(row.get("generate")):
            continue
        events.append(
            quick_event(
                "update_officer_particulars",
                row.get("effective_date") or default_date,
                target_person_id=row.get("target_person_id") or row.get("person_id"),
                target_name=row.get("target_name") or row.get("person_name"),
                old_value=row.get("old_value"),
                new_value=row.get("new_value"),
                field_label=row.get("field_label") or row.get("change_field"),
                remarks=row.get("remarks"),
            )
        )
    return events


def quick_transfer_in_events(company: dict[str, Any], fields: dict[str, Any], task_flags: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    if not transfer_in_enabled(fields, task_flags):
        return []
    return [
        quick_event(
            "transfer_in",
            fields.get("effective_date") or quick_task_date(task_flags, "transfer_in", company),
            approval_route="EGM+DR",
            document_group="TAKEOVER-001",
            combine_in_dr="No",
            old_value=fields.get("old_secretary_company") or company.get("current_secretary_company"),
            new_value=fields.get("new_secretary_company") or company.get("new_secretary_company") or "RSIN GROUP PTE. LTD.",
            resignation_letter=fields.get("generate_resignation_letter") or "No",
            manual_review_required="No",
            remarks=fields.get("remarks"),
        )
    ]


def quick_event(event_type: str, effective_date: Any, **extra: Any) -> dict[str, Any]:
    row = {
        "event_id": "",
        "event_type": event_type,
        "event_name_cn": EVENT_NAME_CN.get(event_type, event_type),
        "generate": "Yes",
        "effective_date": effective_date,
        "approval_route": extra.pop("approval_route", "DR"),
        "document_group": extra.pop("document_group", "DR-001"),
        "combine_in_dr": extra.pop("combine_in_dr", "Yes"),
        "resignation_letter": extra.pop("resignation_letter", "No"),
        "manual_review_required": extra.pop("manual_review_required", "No"),
    }
    row.update({key: clean(value) for key, value in extra.items()})
    return row


def normalize_quick_rows(rows: list[dict[str, Any]], company: dict[str, Any]) -> list[dict[str, Any]]:
    default_date = company.get("default_document_date", "")
    out = []
    for row in rows:
        item = dict(row)
        if not item.get("generate"):
            item["generate"] = "Auto"
        if default_date:
            for key in ["transfer_date", "allotment_date", "authority_date"]:
                if key in item and not item.get(key):
                    item[key] = default_date
        out.append(item)
    return out


def parse_human_maintenance(wb, common_people: dict[str, dict[str, Any]]) -> dict[str, Any]:
    company = read_vertical_kv_sheet(wb, "维护公司信息")
    company.setdefault("task_type", "maintenance")
    apply_default_document_date(company)
    shareholdings = read_transposed_table(wb, "股东现状")
    people = resolve_people(normalize_human_maintenance_people(read_transposed_table(wb, "人员信息"), company), common_people)
    apply_human_maintenance_role_defaults(people, shareholdings)
    change_events = read_transposed_table(wb, "变更事项")
    share_transfers = read_transposed_table(wb, "股份转让")
    share_allotments = read_transposed_table(wb, "增资配股")
    annual_review = read_vertical_kv_sheet(wb, "年审信息")
    output_options = read_vertical_kv_sheet(wb, "输出设置")
    return {
        "company": company,
        "people": people,
        "shareholders": shareholdings,
        "shareholdings": shareholdings,
        "generation": output_options,
        "output_options": output_options,
        "changes": change_events,
        "change_events": change_events,
        "personal_changes": [],
        "share_changes": share_transfers,
        "share_transfers": share_transfers,
        "share_allotments": share_allotments,
        "annual_review": annual_review,
    }


def normalize_human_maintenance_people(rows: list[dict[str, Any]], company: dict[str, Any]) -> list[dict[str, Any]]:
    default_date = company.get("default_document_date", "")
    out = []
    for row in rows:
        item = dict(row)
        if not item.get("source"):
            item["source"] = "new"
        if default_date and not item.get("effective_date"):
            item["effective_date"] = default_date
        out.append(item)
    return out


def apply_human_maintenance_role_defaults(people: list[dict[str, Any]], shareholdings: list[dict[str, Any]]) -> None:
    shareholder_person_ids = {str(row.get("person_id", "")).strip().lower() for row in shareholdings if row.get("person_id")}
    shareholder_names = {str(row.get("shareholder_name", "")).strip().lower() for row in shareholdings if row.get("shareholder_name")}
    boolean_keys = [
        "is_director",
        "is_local_resident_director",
        "is_nominee_director",
        "is_secretary",
        "is_shareholder",
        "is_client_signatory",
    ]
    for person in people:
        person_id = str(person.get("person_id") or "").strip().lower()
        name = str(person.get("full_name") or person.get("common_person_name") or "").strip().lower()
        matched_shareholder = bool((person_id and person_id in shareholder_person_ids) or (name and name in shareholder_names))
        if is_auto(person.get("is_shareholder")):
            person["is_shareholder"] = "Yes" if matched_shareholder else "No"
        for key in boolean_keys:
            if key != "is_client_signatory" and is_auto(person.get(key)):
                person[key] = "No"
        if is_auto(person.get("is_client_signatory")):
            person["is_client_signatory"] = "Yes" if any(is_yes(person.get(key)) for key in ["is_director", "is_shareholder"]) else "No"


def parse_unknown(wb) -> dict[str, Any]:
    sheets = {}
    for name in wb.sheetnames:
        sheets[name] = read_table(wb, name)
    return {
        "company": {},
        "people": [],
        "shareholders": [],
        "generation": {},
        "changes": [],
        "personal_changes": [],
        "share_changes": [],
        "raw_sheets": sheets,
    }


def normalized_common_person_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(clean(value)).lower())


def common_person_alias_values(display_name: str, common: dict[str, Any]) -> list[str]:
    values = [
        display_name,
        common.get("display_name"),
        common.get("full_name"),
        common.get("signature_text"),
    ]
    aliases = re.split(r"[,;，；\n/]+", str(common.get("aliases") or ""))
    values.extend(alias.strip() for alias in aliases if alias.strip())

    name_parts = re.findall(r"[A-Za-z0-9]+", str(display_name))
    if name_parts:
        values.append(name_parts[0])
        if len(name_parts) > 1:
            values.append(name_parts[-1])
    return [str(clean(value)) for value in values if str(clean(value))]


def common_person_lookup(common_people: dict[str, dict[str, Any]]) -> dict[str, str]:
    buckets: dict[str, set[str]] = {}
    for display_name, common in common_people.items():
        for value in common_person_alias_values(display_name, common):
            key = normalized_common_person_key(value)
            if len(key) >= 3:
                buckets.setdefault(key, set()).add(display_name)
    return {key: next(iter(names)) for key, names in buckets.items() if len(names) == 1}


def find_common_person(row: dict[str, Any], common_people: dict[str, dict[str, Any]]) -> tuple[str, dict[str, Any]] | tuple[str, None]:
    candidates = [
        row.get("common_person_name"),
        row.get("full_name"),
    ]
    for candidate in candidates:
        name = str(clean(candidate))
        if name in common_people:
            return name, common_people[name]

    lower_lookup = {str(name).strip().lower(): name for name in common_people}
    for candidate in candidates:
        key = str(clean(candidate)).lower()
        if key in lower_lookup:
            matched_name = lower_lookup[key]
            return matched_name, common_people[matched_name]

    normalized_lookup = common_person_lookup(common_people)
    for candidate in candidates:
        key = normalized_common_person_key(candidate)
        if key in normalized_lookup:
            matched_name = normalized_lookup[key]
            return matched_name, common_people[matched_name]
        if len(key) >= 3:
            prefix_matches = [
                name
                for lookup_key, name in normalized_lookup.items()
                if lookup_key.startswith(key)
            ]
            if len(set(prefix_matches)) == 1:
                matched_name = prefix_matches[0]
                return matched_name, common_people[matched_name]
    return "", None


def resolve_people(rows: list[dict[str, Any]], common_people: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    resolved = []
    for row in rows:
        source = str(clean(row.get("source"))).lower()
        item = dict(row)
        name, common = find_common_person(item, common_people) if source == "common" else ("", None)
        if common:
            item["common_person_name"] = item.get("common_person_name") or name
            item["common_person_id"] = item.get("common_person_id") or common.get("id") or ""
            item["common_person_display_name"] = item.get("common_person_display_name") or common.get("display_name") or name
            for key in ["full_name", "id_type", "id_number", "nationality", "residential_address", "email", "phone"]:
                item[key] = item.get(key) or common.get(key) or ""
            for key in ["signature_text", "signature_image_path", "auto_signature_enabled"]:
                item[key] = item.get(key) or common.get(key) or ""
            if not is_no(item.get("is_director")) and common.get("is_local_resident_director") and is_auto(item.get("is_local_resident_director")):
                item["is_local_resident_director"] = "Yes"
            apply_common_person_default_roles(item, common)
        resolved.append(item)
    return resolved


def apply_common_person_default_roles(item: dict[str, Any], common: dict[str, Any]) -> None:
    role_text = str(common.get("default_role", "") or "").lower()
    if "secretary" in role_text and is_auto(item.get("is_secretary")):
        item["is_secretary"] = "Yes"
    if "nominee director" in role_text:
        if not is_no(item.get("is_director")):
            if is_auto(item.get("is_director")):
                item["is_director"] = "Yes"
            if is_auto(item.get("is_nominee_director")):
                item["is_nominee_director"] = "Yes"
            if is_auto(item.get("is_local_resident_director")):
                item["is_local_resident_director"] = "Yes"
    elif "local resident director" in role_text:
        if not is_no(item.get("is_director")):
            if is_auto(item.get("is_director")):
                item["is_director"] = "Yes"
            if is_auto(item.get("is_local_resident_director")):
                item["is_local_resident_director"] = "Yes"
    elif "client director" in role_text and is_auto(item.get("is_director")):
        item["is_director"] = "Yes"
    if "client signatory" in role_text and is_auto(item.get("signing_required")):
        item["signing_required"] = "Yes"
    if "client signatory" in role_text and is_auto(item.get("is_client_signatory")):
        item["is_client_signatory"] = "Yes"


def option_map(rows: list[dict[str, Any]]) -> dict[str, Any]:
    result = {}
    for row in rows:
        key = row.get("option_key")
        if key:
            result[str(key)] = row.get("value", "")
    return result


def to_json(data: dict[str, Any]) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2, default=str)
